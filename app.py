
import os
import re
import time
import json
import hmac
import uuid
import smtplib
import requests
import razorpay
import pyotp
from datetime import datetime, timedelta
from functools import wraps
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from urllib.parse import urlparse

from flask import Flask, g, jsonify, redirect, render_template, request, session, url_for, flash, has_request_context
from werkzeug.routing import BuildError
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash, check_password_hash
from supabase import create_client, Client as SupabaseClient

from utils.shipping_manager import get_shipping_provider, ShiprocketProvider
from utils.credential_crypto import encrypt_credentials, decrypt_credentials
from utils.stock_ingestion import initialize_stock_tables_if_needed, sync_daily_data
from utils.stock_auth import (
    initialize_stocks_auth_if_needed,
    authenticate_stocks_admin,
    stocks_login_required,
    stocks_role_required,
    list_stocks_admin_users,
    create_child_admin,
    toggle_child_admin_active,
)
from utils.kite_client import KiteClient
from utils.kite_session import (
    initialize_kite_session_table_if_needed,
    get_kite_login_url,
    exchange_request_token,
    save_kite_access_token,
    get_kite_access_token,
    get_kite_session_status,
    IST,
)
from utils.kite_postback import (
    initialize_kite_postback_log_table_if_needed,
    verify_postback_checksum,
    log_postback,
)
import auth_providers
import io
from PIL import Image as PILImage


def load_env_file(env_path):
    if not os.path.exists(env_path):
        return
    with open(env_path, 'r', encoding='utf-8') as env_file:
        for raw_line in env_file:
            line = raw_line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


app = Flask(__name__)
# Render sits in front of the app behind a proxy; without this, request.remote_addr
# shows an internal 10.x.x.x hop IP (see Render logs) instead of the real visitor IP,
# which makes IP-based bot/rate-limit checks useless.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

@app.template_filter('fromjson')
def fromjson_filter(value):
    """Jinja2 filter to parse a JSON string into a Python object."""
    if not value:
        return []
    try:
        return json.loads(value)
    except Exception:
        return []
load_env_file(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'nari-nakhre-dev-secret')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
def normalize_supabase_url(raw_url):
    """Strip /rest/v1 suffix if accidentally included in env var."""
    base = (raw_url or '').strip().rstrip('/')
    for suffix in ['/rest/v1', '/rest/v1/']:
        if base.endswith(suffix.rstrip('/')):
            base = base[:-len(suffix.rstrip('/'))]
    return base.rstrip('/')

SUPABASE_URL = normalize_supabase_url(os.environ.get('SUPABASE_URL', ''))
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')

app.config['SHIPPING_PROVIDER'] = os.environ.get('SHIPPING_PROVIDER', 'mock')
app.config['DELHIVERY_API_KEY'] = os.environ.get('DELHIVERY_API_KEY', '')
app.config['WAREHOUSE_PIN'] = os.environ.get('WAREHOUSE_PIN', '482001')
app.config['RAZORPAY_KEY_ID'] = os.environ.get('RAZORPAY_KEY_ID', '')
app.config['RAZORPAY_KEY_SECRET'] = os.environ.get('RAZORPAY_KEY_SECRET', '')


# Supabase client for database operations
_supabase_client: SupabaseClient = None

def get_supabase():
    global _supabase_client
    if _supabase_client is None:
        _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client

DELHIVERY_API_TOKEN = os.environ.get('DELHIVERY_API_TOKEN', '')
DELHIVERY_CLIENT_NAME = os.environ.get('DELHIVERY_CLIENT_NAME', '')
DELHIVERY_PICKUP_LOCATION = os.environ.get('DELHIVERY_PICKUP_LOCATION', '')
# Nickname of a pickup address already registered in the Shiprocket dashboard
# (Settings -> Pickup Addresses) -- their API requires this exact string, not
# a raw address. Blank until one is registered there.
SHIPROCKET_PICKUP_LOCATION = os.environ.get('SHIPROCKET_PICKUP_LOCATION', '')
DELHIVERY_SELLER_GST = os.environ.get('DELHIVERY_SELLER_GST', '')
WAREHOUSE_CITY = os.environ.get('WAREHOUSE_CITY', 'Jabalpur')
WAREHOUSE_STATE = os.environ.get('WAREHOUSE_STATE', 'Madhya Pradesh')
WAREHOUSE_ADDRESS = os.environ.get('WAREHOUSE_ADDRESS', '')
WAREHOUSE_PHONE = os.environ.get('WAREHOUSE_PHONE', '')
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'mohinicosmetics.india@gmail.com')
# Order-related emails (confirmation, tracking updates) and general/support
# emails (welcome, campaigns, contact-form replies) each go out through
# their own Zeptomail Mail Agent, with its own verified sender identity and
# its own API token -- see send_contact_email() for how the matching token
# gets picked.
# Note the inherited casing on the "support" FROM var (lowercase "support"
# in SMTP_support_EMAIL_FROM but uppercase in the two SUPPORT_EMAIL_*
# credential vars below) -- that's exactly how these are named in Render,
# kept as-is here since env var names are case-sensitive.
ORDERS_FROM_EMAIL = os.environ.get('SMTP_ORDERS_FROM_EMAIL', 'orders-noreply@narinakhre.com')
SUPPORT_FROM_EMAIL = os.environ.get('SMTP_support_EMAIL_FROM', 'support-noreply@narinakhre.com')
# Shared secret checked by /cron/weekly-report -- that endpoint has no admin
# session to check (it's hit by a Render Cron Job, see render.yaml + the
# weekly_report_cron.py trigger script), so this header is its only guard.
# Unset by default so the endpoint stays disabled (returns 403) until an
# admin deliberately sets the same value here and on the cron service.
WEEKLY_REPORT_CRON_SECRET = os.environ.get('WEEKLY_REPORT_CRON_SECRET', '')
RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET')
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', '')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')
ADMIN_TOTP_SECRET = os.environ.get('ADMIN_TOTP_SECRET', '')
razorpay_client = razorpay.Client(auth=(os.environ.get("RAZORPAY_KEY_ID"), os.environ.get("RAZORPAY_KEY_SECRET")))

# Auth providers are pluggable -- see auth_providers/base.py. Adding another
# login option later (e.g. Facebook) means adding auth_providers/facebook.py
# and an init_provider() call here; no route handler changes required.
auth_providers.init_provider('google', app)


def upload_image_to_supabase(file_storage_object, filename):
    supabase_url = (os.environ.get('SUPABASE_URL') or '').rstrip('/')
    supabase_key = os.environ.get('SUPABASE_KEY')
    bucket_name = os.environ.get('SUPABASE_BUCKET_NAME', 'products')

    if not supabase_url or not supabase_key or not bucket_name:
        app.logger.error('Supabase configuration missing for image upload.')
        return None

    try:
        # Read raw bytes from the file-like object
        if hasattr(file_storage_object, 'stream') and hasattr(file_storage_object.stream, 'seek'):
            file_storage_object.stream.seek(0)
        elif hasattr(file_storage_object, 'seek'):
            file_storage_object.seek(0)
        raw_bytes = file_storage_object.read()

        # Convert & compress to WebP using PIL
        try:
            img = PILImage.open(io.BytesIO(raw_bytes))
            if img.mode in ('RGBA', 'P', 'LA'):
                img = img.convert('RGBA')
            else:
                img = img.convert('RGB')
            buf = io.BytesIO()
            img.save(buf, format='WEBP', quality=85, method=6)
            binary_payload = buf.getvalue()
            content_type = 'image/webp'
            # Always use .webp extension in the stored filename
            if not filename.lower().endswith('.webp'):
                filename = filename.rsplit('.', 1)[0] + '.webp'
        except Exception as pil_exc:
            app.logger.warning('WebP conversion failed, uploading original: %s', pil_exc)
            binary_payload = raw_bytes
            content_type = getattr(file_storage_object, 'mimetype', 'application/octet-stream')

        upload_url = f"{supabase_url}/storage/v1/object/{bucket_name}/{filename}"
        headers = {
            'Authorization': f'Bearer {supabase_key}',
            'apikey': supabase_key,
            'Content-Type': content_type,
            'x-upsert': 'true',
        }
        response = requests.put(upload_url, headers=headers, data=binary_payload, timeout=30)

        if response.status_code == 200:
            return f"{supabase_url}/storage/v1/object/public/{bucket_name}/{filename}"

        app.logger.error('Supabase upload failed: %s %s', response.status_code, response.text)
        return None
    except Exception as exc:
        app.logger.error('Supabase upload exception: %s', exc)
        return None


class SupabaseDB:
    """
    Wrapper around the Supabase REST API that mimics the sqlite3
    connection interface used throughout the app.
    Uses Supabase PostgREST for SELECT queries and
    direct SQL execution via the rpc/sql endpoint for
    INSERT, UPDATE, DELETE, CREATE TABLE operations.
    """

    def __init__(self, client):
        self._client = client
        self._pending = []

    def execute(self, sql, params=None):
        return SupabaseCursor(self._client, sql, params)

    def commit(self):
        pass  # Supabase REST is auto-commit

    def rollback(self):
        pass

    def close(self):
        pass


class SupabaseCursor:
    """
    Executes SQL via Supabase REST API.
    Uses the execute_sql RPC and correctly unwraps the JSONB response.
    """

    def __init__(self, client, sql, params=None):
        self._client = client
        self._rows = []
        self._rowcount = 0
        self._execute(sql.strip(), params or ())

    def _format_sql(self, sql, params):
        if not params:
            return sql
        parts = sql.split('?')
        if len(parts) - 1 != len(params):
            return sql
        result = ''
        for i, part in enumerate(parts):
            result += part
            if i < len(params):
                val = params[i]
                if val is None:
                    result += 'NULL'
                elif isinstance(val, bool):
                    result += '1' if val else '0'
                elif isinstance(val, (int, float)):
                    result += str(val)
                else:
                    escaped = str(val).replace("'", "''")
                    result += f"'{escaped}'"
        return result

    def _execute(self, sql, params):
        formatted = self._format_sql(sql, params)
        sql_upper = formatted.strip().upper()

        try:
            # For non-SELECT statements use execute_sql RPC
            if not sql_upper.startswith('SELECT') and not sql_upper.startswith('WITH'):
                self._client.rpc('execute_sql', {'query': formatted}).execute()
                self._rows = []
                return

            # For SELECT use execute_sql and parse the JSONB response
            result = self._client.rpc('execute_sql', {'query': formatted}).execute()
            raw = result.data

            if raw is None:
                self._rows = []
                return

            # Supabase returns: [{'execute_sql': '[{row1}, {row2}]'}]
            if isinstance(raw, list) and len(raw) > 0:
                first = raw[0]
                if isinstance(first, dict) and 'execute_sql' in first:
                    inner = first['execute_sql']
                    if inner is None:
                        self._rows = []
                    elif isinstance(inner, list):
                        self._rows = inner
                    elif isinstance(inner, str):
                        try:
                            parsed = json.loads(inner)
                            self._rows = parsed if isinstance(parsed, list) else []
                        except Exception:
                            self._rows = []
                    else:
                        self._rows = []
                    return

            # Fallback: raw is already a list of rows
            if isinstance(raw, list):
                self._rows = raw
            else:
                self._rows = []

        except Exception as e:
            app.logger.error(f'SupabaseCursor error: {e} | SQL: {formatted[:300]}')
            self._rows = []

    def fetchone(self):
        return self._rows[0] if self._rows else None

    def fetchall(self):
        return self._rows

    def __iter__(self):
        return iter(self._rows)


def get_db():
    if 'db' not in g:
        g.db = SupabaseDB(get_supabase())
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    g.pop('db', None)


def initialize_database_if_needed():
    """
    Create all tables in Supabase via the SQL editor RPC.
    This runs once on app startup. Tables are created only if
    they do not already exist so existing data is never touched.
    """
    tables_sql = [
        '''CREATE TABLE IF NOT EXISTS categories (
            id BIGSERIAL PRIMARY KEY,
            name TEXT
        )''',
        '''CREATE TABLE IF NOT EXISTS products (
            id BIGSERIAL PRIMARY KEY,
            sku TEXT NOT NULL UNIQUE,
            description TEXT,
            name TEXT,
            slug TEXT,
            category TEXT,
            sub_category TEXT,
            collection TEXT,
            size TEXT,
            retail_price FLOAT DEFAULT 0.0,
            mrp_price FLOAT DEFAULT 0.0,
            retail_discount_percent FLOAT DEFAULT 0.0,
            wholesale_price FLOAT DEFAULT 0.0,
            min_wholesale_qty INTEGER DEFAULT 0,
            sets_count INTEGER DEFAULT 0,
            image_field TEXT,
            quantity1 INTEGER DEFAULT 0,
            price1 FLOAT DEFAULT 0.0,
            quantity2 INTEGER DEFAULT 0,
            price2 FLOAT DEFAULT 0.0,
            quantity3 INTEGER DEFAULT 0,
            price3 FLOAT DEFAULT 0.0,
            purchase_cost FLOAT DEFAULT 0.0,
            making_charges FLOAT DEFAULT 0.0,
            weight_grams FLOAT DEFAULT 0.0,
            material TEXT,
            brand_name TEXT,
            pack_unit TEXT,
            pack_count INTEGER DEFAULT 0,
            hsn_code TEXT,
            gst_percent FLOAT DEFAULT 0.0,
            stock_total INTEGER DEFAULT 0,
            stock_alert_threshold INTEGER DEFAULT 5,
            box_packing_type TEXT,
            vendor_id TEXT,
            status TEXT,
            is_active INTEGER DEFAULT 1,
            is_featured INTEGER DEFAULT 0,
            category_id BIGINT REFERENCES categories(id),
            weight FLOAT DEFAULT 0.0,
            length FLOAT DEFAULT 0.0,
            breadth FLOAT DEFAULT 0.0,
            height FLOAT DEFAULT 0.0,
            key_features TEXT
        )''',
        '''CREATE TABLE IF NOT EXISTS quotes (
            id BIGSERIAL PRIMARY KEY,
            request_id TEXT UNIQUE,
            name TEXT,
            whatsapp TEXT,
            email TEXT,
            items_json TEXT,
            total_amount FLOAT DEFAULT 0.0,
            status TEXT DEFAULT 'New',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS users (
            id BIGSERIAL PRIMARY KEY
        )''',
        '''CREATE TABLE IF NOT EXISTS order_shipping (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT REFERENCES users(id),
            status TEXT NOT NULL DEFAULT 'pending',
            consignee_name TEXT NOT NULL,
            consignee_phone TEXT NOT NULL,
            consignee_address TEXT NOT NULL,
            consignee_city TEXT NOT NULL,
            consignee_state TEXT NOT NULL,
            consignee_pincode TEXT NOT NULL,
            internal_order_id TEXT NOT NULL UNIQUE,
            delhivery_waybill TEXT
        )''',
        '''CREATE TABLE IF NOT EXISTS coupons (
            id BIGSERIAL PRIMARY KEY,
            code TEXT NOT NULL UNIQUE,
            discount_percent FLOAT DEFAULT 0.0,
            min_order_amount FLOAT DEFAULT 0.0,
            category TEXT,
            sub_category TEXT,
            expiry_date DATE,
            is_active INTEGER DEFAULT 1,
            usage_limit INTEGER DEFAULT 0,
            times_used INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS email_campaigns (
            id BIGSERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            discount_percent FLOAT DEFAULT 0.0,
            max_discount_amount FLOAT DEFAULT 0.0,
            product_ids TEXT,
            status TEXT DEFAULT 'draft',
            recipient_group TEXT,
            recipient_count INTEGER DEFAULT 0,
            sent_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS email_campaign_recipients (
            id BIGSERIAL PRIMARY KEY,
            campaign_id BIGINT NOT NULL REFERENCES email_campaigns(id),
            user_id BIGINT REFERENCES users(id),
            email TEXT NOT NULL,
            name TEXT,
            coupon_code TEXT,
            sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS product_variants (
            id BIGSERIAL PRIMARY KEY,
            master_sku TEXT NOT NULL,
            variant_sku TEXT NOT NULL UNIQUE,
            size TEXT NOT NULL,
            stock_total INTEGER DEFAULT 0,
            stock_alert_threshold INTEGER DEFAULT 5,
            UNIQUE(master_sku, size)
        )''',
        '''CREATE TABLE IF NOT EXISTS credit_transactions (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL REFERENCES users(id),
            amount NUMERIC NOT NULL,
            reason TEXT NOT NULL,
            internal_order_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS product_events (
            id BIGSERIAL PRIMARY KEY,
            sku TEXT NOT NULL,
            event_type TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS page_views (
            id BIGSERIAL PRIMARY KEY,
            visitor_id TEXT,
            path TEXT,
            site_type TEXT,
            referrer TEXT,
            source TEXT,
            user_id BIGINT REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS admin_events (
            id BIGSERIAL PRIMARY KEY,
            event_type TEXT NOT NULL,
            title TEXT NOT NULL,
            detail TEXT,
            related_id TEXT,
            is_read INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS user_addresses (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL REFERENCES users(id),
            nickname TEXT NOT NULL,
            address_type TEXT NOT NULL DEFAULT 'Home',
            recipient_name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT,
            address_line TEXT NOT NULL,
            city TEXT NOT NULL,
            state TEXT NOT NULL,
            pincode TEXT NOT NULL,
            is_default INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        # Multi-courier support (Phase 1 — schema + admin CRUD only, no
        # Shiprocket/Delhivery API calls read these tables yet).
        '''CREATE TABLE IF NOT EXISTS delivery_partners (
            id BIGSERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE CHECK (name IN ('shiprocket', 'delhivery')),
            is_enabled INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''',
        '''CREATE TABLE IF NOT EXISTS delivery_partner_credentials (
            id BIGSERIAL PRIMARY KEY,
            partner_id BIGINT NOT NULL UNIQUE REFERENCES delivery_partners(id),
            environment TEXT NOT NULL DEFAULT 'production' CHECK (environment IN ('staging', 'production')),
            encrypted_credentials TEXT,
            token_cache TEXT,
            token_expires_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )'''
    ]
    client = get_supabase()
    for sql in tables_sql:
        try:
            client.rpc('execute_sql', {'query': sql}).execute()
        except Exception as e:
            app.logger.warning(f'Table init warning (may already exist): {e}')

    # Columns added after the initial table definitions — kept separate so
    # they can be applied to tables that already existed pre-migration.
    alter_sql = [
        'ALTER TABLE products ADD COLUMN IF NOT EXISTS key_features TEXT',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS google_sub TEXT UNIQUE',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS name TEXT',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS email TEXT',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS picture_url TEXT',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
        'ALTER TABLE quotes ADD COLUMN IF NOT EXISTS user_id BIGINT REFERENCES users(id)',
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS referral_code TEXT UNIQUE',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS referred_by BIGINT REFERENCES users(id)',
        'ALTER TABLE products ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
        "UPDATE products SET status='published' WHERE status IS NULL OR status=''",
        "ALTER TABLE products ALTER COLUMN status SET DEFAULT 'published'",
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS cod_collected_amount NUMERIC',
        # Which courier actually created the shipment for delhivery_waybill --
        # needed so cancel/track actions call the right courier's API.
        # Existing rows predate multi-courier support and were all Delhivery.
        "ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS courier_partner TEXT DEFAULT 'delhivery'",
        # ETA text quoted to the customer at checkout time (e.g. Shiprocket's
        # "Aug 15, 2026") -- stored so the confirmation page/emails show the
        # exact estimate that was quoted, not a possibly-different refetch.
        # NULL for couriers/quotes that didn't supply one (Delhivery today).
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS courier_eta TEXT',
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS cod_credit_awarded NUMERIC DEFAULT 0',
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS credits_redeemed NUMERIC DEFAULT 0',
        'ALTER TABLE products ADD COLUMN IF NOT EXISTS model_number TEXT UNIQUE',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS password_hash TEXT',
        'ALTER TABLE coupons ADD COLUMN IF NOT EXISTS is_public INTEGER DEFAULT 1',
        'ALTER TABLE coupons ADD COLUMN IF NOT EXISTS max_discount_amount NUMERIC',
        'ALTER TABLE coupons ADD COLUMN IF NOT EXISTS user_id BIGINT REFERENCES users(id)',
        'ALTER TABLE coupons ADD COLUMN IF NOT EXISTS campaign_id BIGINT REFERENCES email_campaigns(id)',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS welcome_email_sent_at TIMESTAMP',
        'ALTER TABLE email_campaigns ADD COLUMN IF NOT EXISTS min_order_amount NUMERIC DEFAULT 0',
        'ALTER TABLE email_campaign_recipients ADD COLUMN IF NOT EXISTS clicked_at TIMESTAMP',
        "ALTER TABLE email_campaign_recipients ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'sent'",
        'ALTER TABLE email_campaign_recipients ADD COLUMN IF NOT EXISTS error_detail TEXT',
        'ALTER TABLE product_events ADD COLUMN IF NOT EXISTS visitor_id TEXT',
        'ALTER TABLE product_events ADD COLUMN IF NOT EXISTS source TEXT',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS phone TEXT',
        'ALTER TABLE products ADD COLUMN IF NOT EXISTS pack_unit TEXT',
        'ALTER TABLE products ADD COLUMN IF NOT EXISTS pack_count INTEGER DEFAULT 0',
        'ALTER TABLE products ADD COLUMN IF NOT EXISTS brand_name TEXT',
        # Best-effort -- non-fatal if pre-existing rows already have duplicate
        # or blank emails (see the app-level check in email_signup as the
        # real guard against duplicate accounts).
        'ALTER TABLE users ADD CONSTRAINT users_email_unique UNIQUE (email)',
        # Customer-facing surcharge for picking a pricier courier than the
        # cheapest quote at checkout (see calculate_checkout_shipping /
        # checkout_process). The cheapest option always ships free; this is
        # the delta the customer actually paid on top of that, folded into
        # total_amount already -- stored separately too so admin/emails can
        # show the breakdown instead of just a mismatched total.
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS shipping_upgrade_charge NUMERIC DEFAULT 0',
        # Shiprocket's internal numeric shipment id (distinct from the AWB/
        # waybill) -- their label-generation API needs this, not the AWB.
        # NULL for Delhivery orders and for pre-existing Shiprocket orders
        # created before this was tracked.
        'ALTER TABLE order_shipping ADD COLUMN IF NOT EXISTS shiprocket_shipment_id TEXT',
    ]
    for sql in alter_sql:
        try:
            client.rpc('execute_sql', {'query': sql}).execute()
        except Exception as e:
            app.logger.warning(f'Column init warning (may already exist): {e}')

    # Seed the two supported courier partners, disabled until an admin
    # configures credentials and turns them on.
    seed_sql = [
        "INSERT INTO delivery_partners (name, is_enabled) VALUES ('shiprocket', 0) ON CONFLICT (name) DO NOTHING",
        "INSERT INTO delivery_partners (name, is_enabled) VALUES ('delhivery', 0) ON CONFLICT (name) DO NOTHING",
    ]
    for sql in seed_sql:
        try:
            client.rpc('execute_sql', {'query': sql}).execute()
        except Exception as e:
            app.logger.warning(f'Delivery partner seed warning: {e}')

    # Nari Nakhre Credits reservation -- these two functions run the
    # check-balance-then-insert as a single atomic Postgres statement (a
    # plpgsql function body executes strictly sequentially, unlike a CTE,
    # so pg_advisory_xact_lock genuinely blocks a concurrent caller until
    # this one's transaction ends), so two tabs/devices redeeming at the
    # same instant can't both succeed against the same balance.
    functions_sql = [
        '''CREATE OR REPLACE FUNCTION sweep_expired_credit_holds(p_user_id BIGINT)
           RETURNS VOID AS $func$
           BEGIN
               PERFORM pg_advisory_xact_lock(hashtext('nn_credits_' || p_user_id::text));
               INSERT INTO credit_transactions (user_id, amount, reason, internal_order_id)
               SELECT t.user_id, -t.amount, 'hold_expired', t.internal_order_id
               FROM credit_transactions t
               WHERE t.user_id = p_user_id
                 AND t.reason = 'hold_for_order'
                 AND t.created_at < NOW() - INTERVAL '30 minutes'
                 AND NOT EXISTS (
                     SELECT 1 FROM credit_transactions r
                     WHERE r.internal_order_id = t.internal_order_id
                       AND r.reason IN ('hold_expired', 'redeemed_at_checkout')
                 );
           END;
           $func$ LANGUAGE plpgsql''',
        '''CREATE OR REPLACE FUNCTION reserve_credits_atomic(
               p_user_id BIGINT, p_amount NUMERIC, p_order_id TEXT,
               p_reason TEXT DEFAULT 'hold_for_order'
           )
           RETURNS BOOLEAN AS $func$
           DECLARE
               v_available NUMERIC;
           BEGIN
               PERFORM pg_advisory_xact_lock(hashtext('nn_credits_' || p_user_id::text));
               PERFORM sweep_expired_credit_holds(p_user_id);
               SELECT COALESCE(SUM(amount), 0) INTO v_available
               FROM credit_transactions WHERE user_id = p_user_id;
               IF v_available >= p_amount THEN
                   INSERT INTO credit_transactions (user_id, amount, reason, internal_order_id)
                   VALUES (p_user_id, -p_amount, p_reason, p_order_id);
                   RETURN TRUE;
               ELSE
                   RETURN FALSE;
               END IF;
           END;
           $func$ LANGUAGE plpgsql''',
        '''CREATE OR REPLACE FUNCTION finalize_credit_hold(p_order_id TEXT)
           RETURNS BOOLEAN AS $func$
           DECLARE
               v_updated INT;
           BEGIN
               UPDATE credit_transactions
               SET reason = 'redeemed_at_checkout'
               WHERE internal_order_id = p_order_id AND reason = 'hold_for_order';
               GET DIAGNOSTICS v_updated = ROW_COUNT;
               RETURN v_updated > 0;
           END;
           $func$ LANGUAGE plpgsql''',
    ]
    for sql in functions_sql:
        try:
            client.rpc('execute_sql', {'query': sql}).execute()
        except Exception as e:
            app.logger.warning(f'Function init warning: {e}')

    app.logger.info('Database tables verified/created via Supabase RPC.')


def ensure_checkout_tables_exist():
    """No-op — all tables created in initialize_database_if_needed."""
    pass


initialize_database_if_needed()
ensure_checkout_tables_exist()
# Nari Nakhre Stocks -- separate feature, own tables, kept out of the
# e-commerce schema above. Same Supabase project only -- it has its own
# admin login (see utils/stock_auth.py), not the storefront's.
initialize_stock_tables_if_needed(get_supabase())
initialize_stocks_auth_if_needed(get_supabase())
initialize_kite_session_table_if_needed(get_supabase())
initialize_kite_postback_log_table_if_needed(get_supabase())


def calculate_inclusive_gst(display_cart, discount=0.0, full_subtotal=0.0):
    """Extract GST already included in retail prices (GST-inclusive pricing)."""
    db = get_db()
    total_gst = 0.0
    for item in display_cart:
        line_total = item.get('price', 0) * item.get('units', item.get('qty', 1))
        if line_total <= 0:
            continue
        gst_rate = 3.0
        sku = item.get('sku')
        if sku:
            try:
                prod = db.execute('SELECT gst_percent FROM products WHERE sku=?', (sku,)).fetchone()
                if prod and prod['gst_percent']:
                    gst_rate = float(prod['gst_percent'])
            except Exception:
                pass
        line_gst = line_total - (line_total / (1 + gst_rate / 100.0))
        total_gst += line_gst
    if discount and full_subtotal and full_subtotal > 0:
        discount_ratio = min(discount / full_subtotal, 1.0)
        total_gst = total_gst * (1 - discount_ratio)
    total_gst = round(total_gst, 2)
    half = round(total_gst / 2, 2)
    return {'total_gst': total_gst, 'cgst': half, 'sgst': round(total_gst - half, 2)}


def get_courier(partner_name=None):
    """
    Returns (partner_name, provider_instance).

    With no argument: resolves whichever courier is currently enabled in
    delivery_partners with credentials configured, falling back to the
    legacy SHIPPING_PROVIDER env var + DELHIVERY_API_TOKEN behavior when
    nothing is enabled via the admin panel yet -- so existing behavior
    doesn't change until an admin explicitly flips a partner on. If more
    than one partner is somehow enabled at once, the first by id wins --
    there's no rate-shopping/priority logic yet, that's later work.

    With a specific partner_name: resolves THAT partner's credentials
    regardless of is_enabled. Use this for tracking/cancelling an existing
    shipment -- a waybill belongs to whichever courier actually created it,
    even if the active courier has since been switched or disabled.
    """
    db = get_db()
    if partner_name:
        row = db.execute(
            '''SELECT dp.id, dp.name, dpc.environment, dpc.encrypted_credentials,
                      dpc.token_cache, dpc.token_expires_at
               FROM delivery_partners dp
               JOIN delivery_partner_credentials dpc ON dpc.partner_id = dp.id
               WHERE dp.name = ? AND dpc.encrypted_credentials IS NOT NULL
                     AND dpc.encrypted_credentials != ''
               LIMIT 1''',
            (partner_name,)
        ).fetchone()
    else:
        row = db.execute(
            '''SELECT dp.id, dp.name, dpc.environment, dpc.encrypted_credentials,
                      dpc.token_cache, dpc.token_expires_at
               FROM delivery_partners dp
               JOIN delivery_partner_credentials dpc ON dpc.partner_id = dp.id
               WHERE dp.is_enabled = 1 AND dpc.encrypted_credentials IS NOT NULL
                     AND dpc.encrypted_credentials != ''
               ORDER BY dp.id LIMIT 1'''
        ).fetchone()

    if not row:
        fallback_name = partner_name or app.config.get('SHIPPING_PROVIDER', 'mock')
        if fallback_name == 'delhivery' or not partner_name:
            token = app.config.get('DELHIVERY_API_KEY') or app.config.get('DELHIVERY_API_TOKEN', '')
            return fallback_name, get_shipping_provider(fallback_name if fallback_name != 'shiprocket' else 'mock', api_token=token)
        return fallback_name, get_shipping_provider('mock')

    return row['name'], _build_courier_provider(db, row)


def _build_courier_provider(db, row):
    """Constructs a provider instance from a delivery_partners + credentials
    row, decrypting credentials and persisting a refreshed Shiprocket token
    back to the DB when one gets issued. Shared by get_courier() (single
    partner) and get_enabled_couriers() (all enabled partners, for
    rate-shopping) so the token-refresh logic lives in exactly one place."""
    creds = decrypt_credentials(row['encrypted_credentials'])

    if row['name'] == 'delhivery':
        return get_shipping_provider('delhivery', api_token=creds.get('api_token', ''))

    if row['name'] == 'shiprocket':
        token_expires_at = None
        if row['token_expires_at']:
            try:
                token_expires_at = datetime.fromisoformat(str(row['token_expires_at']))
            except Exception:
                token_expires_at = None
        provider = get_shipping_provider(
            'shiprocket',
            email=creds.get('email', ''),
            password=creds.get('password', ''),
            cached_token=row['token_cache'],
            token_expires_at=token_expires_at,
        )
        if provider.token_refreshed:
            db.execute(
                'UPDATE delivery_partner_credentials SET token_cache=?, token_expires_at=?, updated_at=NOW() WHERE partner_id=?',
                (provider.token, provider.token_expires_at.isoformat(), row['id'])
            )
            db.commit()
        return provider

    # Unknown partner name shouldn't happen given the CHECK constraint.
    return get_shipping_provider('mock')


def get_active_courier():
    """Whichever courier is currently enabled -- see get_courier()."""
    return get_courier()


def get_enabled_couriers():
    """Returns [(partner_name, provider), ...] for every currently enabled,
    credentialed partner -- used for rate-shopping. Empty list if none are
    enabled (caller should fall back to get_active_courier() in that case,
    which covers the legacy env-var behavior)."""
    db = get_db()
    rows = db.execute(
        '''SELECT dp.id, dp.name, dpc.environment, dpc.encrypted_credentials,
                  dpc.token_cache, dpc.token_expires_at
           FROM delivery_partners dp
           JOIN delivery_partner_credentials dpc ON dpc.partner_id = dp.id
           WHERE dp.is_enabled = 1 AND dpc.encrypted_credentials IS NOT NULL
                 AND dpc.encrypted_credentials != ''
           ORDER BY dp.id'''
    ).fetchall()
    return [(row['name'], _build_courier_provider(db, row)) for row in rows]


def get_configured_courier_names():
    """Returns partner names with credentials saved, regardless of
    is_enabled. is_enabled gates live customer-checkout rate-shopping; the
    admin's manual order tool is a deliberate one-off action and shouldn't
    be limited to whatever's currently live -- an admin may want to test or
    use a courier there before ever flipping it on for real traffic."""
    db = get_db()
    rows = db.execute(
        '''SELECT dp.name FROM delivery_partners dp
           JOIN delivery_partner_credentials dpc ON dpc.partner_id = dp.id
           WHERE dpc.encrypted_credentials IS NOT NULL AND dpc.encrypted_credentials != ''
           ORDER BY dp.id'''
    ).fetchall()
    return [row['name'] for row in rows]


def get_all_courier_quotes(o_pin, d_pin, weight, mode="Prepaid"):
    """
    Rate-shops across every currently enabled courier and returns every
    working quote as [(partner_name, provider, rate_info), ...] -- rate_info
    is whatever that provider's get_rates() returned (rate, shipping_charge,
    cod_fee, and 'eta' when the courier supplies one; Delhivery's rate API
    doesn't return an ETA today, so rate_info['eta'] will be None for
    Delhivery quotes). Ordered cheapest first.

    With only one partner enabled, returns just that one (no comparison
    needed). With none enabled, falls back to get_active_courier()'s legacy
    behavior. A courier whose rate call fails/errors is skipped, not fatal
    -- if every enabled courier fails, falls back to whichever
    get_active_courier() would have picked anyway, so checkout never
    hard-fails just because one courier's API had a bad moment.
    """
    couriers = get_enabled_couriers()
    if not couriers:
        partner_name, provider = get_active_courier()
        couriers = [(partner_name, provider)]

    quotes = []
    for partner_name, provider in couriers:
        try:
            rate_info = provider.get_rates(o_pin, d_pin, weight, mode=mode)
            # Both providers only include 'msg' on a failed quote -- never
            # on success -- so its absence is a reliable success check.
            if 'msg' not in rate_info:
                quotes.append((partner_name, provider, rate_info))
        except Exception as e:
            app.logger.warning(f'Rate-shop: {partner_name} quote failed: {e}')

    if not quotes:
        # Every enabled courier failed to quote -- fall back to whichever
        # one get_active_courier() would pick, so checkout degrades instead
        # of breaking outright.
        partner_name, provider = get_active_courier()
        try:
            rate_info = provider.get_rates(o_pin, d_pin, weight, mode=mode)
        except Exception:
            rate_info = {"rate": 0, "shipping_charge": 0, "cod_fee": 0, "eta": None}
        return [(partner_name, provider, rate_info)]

    def _rate_of(q):
        info = q[2]
        return float(info.get('rate') if info.get('rate') is not None else info.get('shipping_charge') or 0)

    return sorted(quotes, key=_rate_of)


def get_best_courier_quote(o_pin, d_pin, weight, mode="Prepaid"):
    """The single cheapest quote -- see get_all_courier_quotes()."""
    return get_all_courier_quotes(o_pin, d_pin, weight, mode=mode)[0]


def _customer_shipping_charge(rate_info):
    """What a given quote's rate_info would cost -- 'rate' if the provider
    supplied one, else 'shipping_charge'. Same key precedence get_all_courier_quotes
    sorts by, so this always agrees with which quote is "cheapest"."""
    return float(rate_info.get('rate') if rate_info.get('rate') is not None else rate_info.get('shipping_charge') or 0)


def create_shiprocket_shipment(order_row, cart_items, cod_amount_override=None, provider=None):
    """Create a Shiprocket order + assign an AWB after payment confirmation.
    Returns (waybill, error_msg, shipment_id) -- shipment_id is Shiprocket's
    own internal numeric id (distinct from the AWB), which their label-
    generation API needs; the caller must persist it if real label fetching
    (see ShiprocketProvider.get_label) is to work later.

    UNVERIFIED against a real order -- only Shiprocket's login, serviceability,
    and rate endpoints have been tested live so far. Do not enable Shiprocket
    for real customers until this has been proven with one real test order.
    """
    if not SHIPROCKET_PICKUP_LOCATION:
        return None, "Shiprocket pickup location not configured (SHIPROCKET_PICKUP_LOCATION) -- register a pickup address in the Shiprocket dashboard first", None
    if provider is None:
        _, provider = get_active_courier()
    if not getattr(provider, 'token', None):
        return None, (provider.last_error if hasattr(provider, 'last_error') else "Shiprocket not authenticated"), None

    order_row_dict = dict(order_row) if not isinstance(order_row, dict) else order_row
    consignee_name = order_row_dict.get('consignee_name', '')
    phone = order_row_dict.get('consignee_phone', '')
    address = order_row_dict.get('consignee_address', '')
    city = order_row_dict.get('consignee_city', '')
    state = order_row_dict.get('consignee_state', '') or ''
    pincode = str(order_row_dict.get('consignee_pincode', ''))
    internal_order_id = order_row_dict.get('internal_order_id', '')
    payment_mode = order_row_dict.get('payment_mode', 'Prepaid')
    total_amount = float(order_row_dict.get('total_amount', 0) or 0)
    is_cod = (payment_mode == 'COD')
    # Shiprocket has no separate cod_amount override field like Delhivery --
    # it collects whatever sub_total says, so the rounded COD collection
    # amount (see round_cod_amount) becomes the sub_total itself here.
    sub_total = float(cod_amount_override) if cod_amount_override is not None else total_amount

    total_qty = max(sum(int(item.get('units', item.get('qty', 1))) for item in cart_items), 1) if cart_items else 1
    weight_kg = max(total_qty * 0.25, 0.25)

    order_items = [{
        "name": item.get('name') or item.get('sku') or 'Item',
        "sku": item.get('sku') or internal_order_id or 'SKU',
        "units": int(item.get('units', item.get('qty', 1))),
        "selling_price": float(item.get('price', 0) or 0),
        "hsn": 7117,
    } for item in (cart_items or [])]
    if not order_items:
        order_items = [{"name": "Order", "sku": internal_order_id or "SKU", "units": 1,
                         "selling_price": total_amount, "hsn": 7117}]

    payload = {
        "order_id": internal_order_id,
        "order_date": datetime.now().strftime('%Y-%m-%d %H:%M'),
        "pickup_location": SHIPROCKET_PICKUP_LOCATION,
        "billing_customer_name": consignee_name,
        "billing_last_name": "",
        "billing_address": address,
        "billing_city": city,
        "billing_pincode": pincode,
        "billing_state": state,
        "billing_country": "India",
        "billing_email": order_row_dict.get('consignee_email') or ADMIN_EMAIL,
        "billing_phone": phone,
        "shipping_is_billing": True,
        "order_items": order_items,
        "payment_method": "COD" if is_cod else "Prepaid",
        "sub_total": sub_total,
        "length": 20, "breadth": 15, "height": 10,
        "weight": weight_kg,
    }
    headers = {"Authorization": f"Bearer {provider.token}", "Content-Type": "application/json"}

    try:
        response = requests.post(
            f"{ShiprocketProvider.BASE_URL}/orders/create/adhoc",
            json=payload, headers=headers, timeout=30
        )
        resp_json = response.json()
        app.logger.info(f"Shiprocket create response: {resp_json}")
        shipment_id = resp_json.get('shipment_id')
        waybill = resp_json.get('awb_code')
        if not shipment_id:
            error_msg = resp_json.get('message') or str(resp_json)
            return None, f"Shiprocket order creation error: {error_msg}", None
        if waybill:
            return waybill, None, shipment_id
        # Order creation alone doesn't always assign a courier/AWB -- request
        # one explicitly, letting Shiprocket auto-pick the courier.
        assign_resp = requests.post(
            f"{ShiprocketProvider.BASE_URL}/courier/assign/awb",
            json={"shipment_id": shipment_id}, headers=headers, timeout=30
        )
        assign_json = assign_resp.json()
        app.logger.info(f"Shiprocket AWB assign response: {assign_json}")
        awb_data = (assign_json.get('response') or {}).get('data') or {}
        waybill = awb_data.get('awb_code')
        if waybill:
            return waybill, None, shipment_id
        return None, f"Order created (shipment_id={shipment_id}) but AWB assignment failed: {assign_json.get('message') or assign_json}", shipment_id
    except Exception as e:
        app.logger.error(f"Shiprocket shipment creation exception: {e}")
        return None, str(e), None


def create_courier_shipment(order_row, cart_items, cod_amount_override=None):
    """Dispatches shipment creation to whichever courier the order was
    already quoted with at checkout time (order_row['courier_partner'],
    set by the rate-shopping in checkout_process) -- so what actually ships
    matches what the customer was shown, even if the active/cheapest
    courier has changed since. Falls back to get_active_courier() only for
    orders with no stored courier_partner (pre-multi-courier legacy rows,
    or the admin-accept path for very old orders). Returns
    (waybill, error_msg, partner_name, shipment_id) -- partner_name lets
    callers do courier-specific follow-up (e.g. pickup scheduling)
    correctly; shipment_id is Shiprocket's internal shipment id (needed for
    real label fetching later), always None for Delhivery."""
    order_row_dict = dict(order_row) if not isinstance(order_row, dict) else order_row
    stored_partner = order_row_dict.get('courier_partner')
    if stored_partner:
        partner_name, provider = get_courier(stored_partner)
    else:
        partner_name, provider = get_active_courier()
    if partner_name == 'shiprocket':
        waybill, err, shipment_id = create_shiprocket_shipment(order_row, cart_items, cod_amount_override=cod_amount_override, provider=provider)
    else:
        waybill, err = create_delhivery_shipment(order_row, cart_items, cod_amount_override=cod_amount_override)
        partner_name = 'delhivery'
        shipment_id = None
    return waybill, err, partner_name, shipment_id


def create_delhivery_shipment(order_row, cart_items, cod_amount_override=None):
    """Create a Delhivery shipment after payment confirmation. Returns (waybill, error_msg).

    cod_amount_override, when given, is the whole-rupee amount actually
    collected in cash (see round_cod_amount) -- Delhivery gets told to
    collect this instead of the exact order total, which may have decimals
    or not be a multiple of 10.
    """
    if not DELHIVERY_API_TOKEN:
        return None, "Delhivery API token not configured"
    order_row_dict = dict(order_row) if not isinstance(order_row, dict) else order_row
    consignee_name = order_row_dict.get('consignee_name', '')
    phone = order_row_dict.get('consignee_phone', '')
    address = order_row_dict.get('consignee_address', '')
    city = order_row_dict.get('consignee_city', '')
    state = order_row_dict.get('consignee_state', '') or ''
    pincode = str(order_row_dict.get('consignee_pincode', ''))
    internal_order_id = order_row_dict.get('internal_order_id', '')
    payment_mode = order_row_dict.get('payment_mode', 'Prepaid')
    total_amount = float(order_row_dict.get('total_amount', 0) or 0)
    cod_amount = float(cod_amount_override) if cod_amount_override is not None else total_amount
    total_qty = max(sum(int(item.get('units', item.get('qty', 1))) for item in cart_items), 1) if cart_items else 1
    weight_grams = max(total_qty * 250, 250)
    delhivery_payment_mode = 'COD' if payment_mode == 'COD' else 'Prepaid'
    shipment = {
        'name': consignee_name, 'phone': phone, 'add': address,
        'city': city, 'state': state, 'pin': pincode, 'country': 'IN',
        'order': internal_order_id,
        'payment_mode': delhivery_payment_mode,
        'cod_amount': cod_amount if delhivery_payment_mode == 'COD' else 0,
        'weight': weight_grams,
        'shipment_width': 15, 'shipment_height': 10, 'shipment_length': 20,
        'quantity': total_qty, 'hsn_code': '7117',
        'seller_gst_tin': DELHIVERY_SELLER_GST,
        'client': DELHIVERY_CLIENT_NAME,
        'return_pin': app.config.get('WAREHOUSE_PIN', '482001'),
        'return_city': WAREHOUSE_CITY, 'return_state': WAREHOUSE_STATE,
        'return_add': WAREHOUSE_ADDRESS or address,
        'return_phone': WAREHOUSE_PHONE or phone,
        'return_name': DELHIVERY_CLIENT_NAME or consignee_name,
        'return_country': 'IN',
    }
    payload = {'shipments': [shipment], 'pickup_location': {'name': DELHIVERY_PICKUP_LOCATION}}
    try:
        response = requests.post(
            'https://track.delhivery.com/api/cmu/create.json',
            data={'format': 'json', 'data': json.dumps(payload)},
            headers={'Authorization': f'Token {DELHIVERY_API_TOKEN}',
                     'Content-Type': 'application/x-www-form-urlencoded'},
            timeout=30,
        )
        resp_json = response.json()
        app.logger.info(f"Delhivery create response: {resp_json}")
        packages = resp_json.get('packages', [])
        if packages and isinstance(packages, list):
            waybill = packages[0].get('waybill')
            if waybill:
                return waybill, None
        error_msg = resp_json.get('rmk') or resp_json.get('error') or str(resp_json)
        return None, f"Delhivery error: {error_msg}"
    except Exception as e:
        app.logger.error(f"Delhivery shipment creation exception: {e}")
        return None, str(e)


def send_contact_email(to_email, subject, body, html_body=None, from_email=None, capture_detail=False):
    """
    Send email via Zeptomail's HTTP API (NOT SMTP).
    This replaces the old smtplib-based sender: a raw SMTP socket can wedge
    indefinitely if Zeptomail's server accepts the TCP connection but doesn't
    respond cleanly (which is exactly what "Sender Org Blocked" can cause).
    An HTTPS POST via `requests` with an explicit timeout can't do that — it
    either succeeds, fails, or times out in 10s. No other behavior change.

    Credentials from Render environment variables:
        ZEPTOMAIL_API_URL          = https://api.zeptomail.in/v1.1/email  (default, .in region)
        SMTP_ORDERS_ZEPTO_PASSWORD = API token for the "orders" Mail Agent (orders-noreply@)
        SMTP_SUPPORT_EMAIL_PASSWORD = API token for the "support" Mail Agent (support-noreply@)
    Each Zeptomail Mail Agent has its own verified sender identity and its
    own API token -- a token from one agent can't send as the other agent's
    From address, which is exactly why sends from the old info@ address were
    being silently rejected. The From address picks which token gets used.
    The From address must be a verified sender in Zeptomail.

    Returns a plain bool by default (True/False), matching every existing
    call site. Pass capture_detail=True to instead get back (ok, detail) --
    detail is None on success, or Zeptomail's actual rejection reason (HTTP
    status + response body) on failure, so a caller that needs to show WHY a
    send failed (e.g. the campaign sender) doesn't have to go digging
    through server logs for it.
    """
    def result(ok, detail=None):
        return (ok, detail) if capture_detail else ok

    api_url = os.environ.get('ZEPTOMAIL_API_URL', 'https://api.zeptomail.in/v1.1/email')
    FROM_EMAIL = from_email or SUPPORT_FROM_EMAIL

    if FROM_EMAIL == ORDERS_FROM_EMAIL:
        api_key = os.environ.get('SMTP_ORDERS_ZEPTO_PASSWORD', '')
    else:
        api_key = os.environ.get('SMTP_SUPPORT_EMAIL_PASSWORD', '')

    if not api_key:
        msg = f'no Zeptomail API token configured for sender {FROM_EMAIL}'
        app.logger.warning(f'Email send skipped: {msg}')
        return result(False, msg)

    payload = {
        'from': {'address': FROM_EMAIL, 'name': 'Nari Nakhre'},
        'to': [{'email_address': {'address': to_email, 'name': to_email}}],
        'subject': subject,
        'textbody': body,
    }
    if html_body:
        payload['htmlbody'] = html_body

    headers = {
        'Authorization': f'Zoho-enczapikey {api_key}',
        'Content-Type': 'application/json',
        'Accept': 'application/json',
    }

    try:
        resp = requests.post(api_url, json=payload, headers=headers, timeout=10)
        if resp.status_code in (200, 201):
            app.logger.info(f'Email sent to {to_email}: {subject}')
            return result(True)
        detail = f'HTTP {resp.status_code}: {resp.text[:800]}'
        app.logger.error(f'Zeptomail API error ({resp.status_code}) sending to {to_email}: {resp.text[:500]}')
        return result(False, detail)
    except requests.exceptions.Timeout:
        app.logger.error(f'Zeptomail API call timed out (10s) sending to {to_email}')
        return result(False, 'Request to Zeptomail timed out after 10s')
    except Exception as e:
        app.logger.error(f'Email send failed to {to_email}: {type(e).__name__}: {e}')
        return result(False, f'{type(e).__name__}: {e}')


def send_contact_email_async(*args, **kwargs):
    """Fire-and-forget wrapper around send_contact_email. Use this on any
    customer-facing request path (order confirmation, etc.) so a slow or
    blocked SMTP server (see: Zeptomail 'Sender Org Blocked') can never hang
    the HTTP response the customer is waiting on."""
    t = threading.Thread(target=send_contact_email, args=args, kwargs=kwargs, daemon=True)
    t.start()
    return t


# --- CONTACT FORM ANTI-BOT HELPERS ---
# A prior bot attack (on PythonAnywhere hosting) hammered the contact forms with
# fake-but-regex-valid addresses; the resulting bounces/complaints got Zeptomail
# to block the whole sending org ("Sender Org Blocked"). These three checks —
# strict email format, a minimum human-fill-time trap, and per-IP throttling —
# run in addition to the existing honeypot field.
EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$')

# In-memory only — resets on deploy/restart and is per-worker-process, not shared
# across multiple gunicorn workers/dynos. Good enough to blunt simple bot floods;
# not a substitute for a real WAF/CAPTCHA if attacks continue.
_contact_last_submit = {}
CONTACT_RATE_LIMIT_SECONDS = 30
CONTACT_MIN_FILL_SECONDS = 3
CONTACT_MAX_FORM_AGE_SECONDS = 3600


def is_valid_email(email):
    return bool(email) and bool(EMAIL_RE.match(email)) and len(email) < 100


def contact_form_is_bot(form_rendered_at):
    """True if the timestamp hidden field is missing/malformed, filled in too
    fast to be a human, or is stale (replayed from a cached/old page)."""
    try:
        rendered_at = float(form_rendered_at)
    except (TypeError, ValueError):
        return True
    elapsed = time.time() - rendered_at
    return elapsed < CONTACT_MIN_FILL_SECONDS or elapsed > CONTACT_MAX_FORM_AGE_SECONDS


def contact_ip_is_rate_limited(ip):
    now = time.time()
    last = _contact_last_submit.get(ip)
    _contact_last_submit[ip] = now
    # Opportunistically trim so this dict doesn't grow forever
    if len(_contact_last_submit) > 5000:
        cutoff = now - CONTACT_RATE_LIMIT_SECONDS
        for k, v in list(_contact_last_submit.items()):
            if v < cutoff:
                del _contact_last_submit[k]
    return last is not None and (now - last) < CONTACT_RATE_LIMIT_SECONDS


# reCAPTCHA v3 config — set RECAPTCHA_SITE_KEY / RECAPTCHA_SECRET_KEY on Render.
# If unset, the check is skipped entirely (so the form still works before you've
# generated keys) and a one-time warning is logged.
RECAPTCHA_SITE_KEY = os.environ.get('RECAPTCHA_SITE_KEY', '')
RECAPTCHA_SECRET_KEY = os.environ.get('RECAPTCHA_SECRET_KEY', '')
RECAPTCHA_MIN_SCORE = 0.5
_recaptcha_unconfigured_warned = False


def verify_recaptcha(token, remote_ip=None, expected_action=None):
    """Returns True if the submission should be allowed through.
    Fails OPEN (allows the submission) if reCAPTCHA isn't configured yet, or if
    Google's API can't be reached — the timing/IP/honeypot checks still apply
    either way, so this is a defense layer, not the only one."""
    global _recaptcha_unconfigured_warned
    if not RECAPTCHA_SECRET_KEY:
        if not _recaptcha_unconfigured_warned:
            app.logger.warning('RECAPTCHA_SECRET_KEY not set — skipping reCAPTCHA checks on contact forms')
            _recaptcha_unconfigured_warned = True
        return True

    if not token:
        app.logger.warning('reCAPTCHA rejected: no token submitted')
        return False

    try:
        resp = requests.post(
            'https://www.google.com/recaptcha/api/siteverify',
            data={'secret': RECAPTCHA_SECRET_KEY, 'response': token, 'remoteip': remote_ip},
            timeout=5,
        )
        result = resp.json()
    except Exception as e:
        app.logger.error(f'reCAPTCHA verify request failed, allowing through: {type(e).__name__}: {e}')
        return True

    if not result.get('success'):
        app.logger.warning(f'reCAPTCHA rejected: {result.get("error-codes")}')
        return False
    if expected_action and result.get('action') != expected_action:
        app.logger.warning(f'reCAPTCHA action mismatch: expected {expected_action}, got {result.get("action")}')
        return False
    score = result.get('score', 0)
    if score < RECAPTCHA_MIN_SCORE:
        app.logger.warning(f'reCAPTCHA score too low: {score}')
        return False
    return True

@app.route('/retail/contact', methods=['GET', 'POST'])
def retail_contact():
    g.site_type = 'retail'
    if request.method == 'POST':
        # Honeypot: bots fill hidden fields, humans don't
        if (request.form.get('system_verification_token') or '').strip():
            app.logger.warning(f'Bot caught on retail contact (honeypot): {request.form.get("email")}')
            return redirect('/retail/thank_you')  # silent discard

        # Timing trap: bots submit instantly, or replay a stale cached page
        if contact_form_is_bot(request.form.get('form_rendered_at')):
            app.logger.warning(f'Bot caught on retail contact (timing): {request.form.get("email")}')
            return redirect('/retail/thank_you')

        client_ip = request.remote_addr or 'unknown'
        if contact_ip_is_rate_limited(client_ip):
            app.logger.warning(f'Retail contact rate-limited: ip={client_ip}')
            return redirect('/retail/thank_you')

        if not verify_recaptcha(request.form.get('recaptcha_token'), remote_ip=client_ip, expected_action='retail_contact'):
            app.logger.warning(f'Bot caught on retail contact (recaptcha): {request.form.get("email")}')
            return redirect('/retail/thank_you')

        name    = (request.form.get('name') or '').strip()
        whatsapp= (request.form.get('whatsapp') or '').strip()
        email   = (request.form.get('email') or '').strip()
        message = (request.form.get('message') or '').strip()

        # Basic validation — reject obviously empty/spam submissions
        if not name or not email or not message or len(message) < 5:
            return redirect('/retail/thank_you')

        # Only send to admin — don't auto-reply to unknown/spam emails
        # This prevents Zeptomail from being blocked for sending to bad addresses
        admin_body = (
            f"New retail contact form submission:\n\n"
            f"Name: {name}\nWhatsApp: {whatsapp}\nEmail: {email}\n"
            f"Message: {message}"
        )
        send_contact_email_async(
            os.environ.get('ADMIN_EMAIL', 'mohinicosmetics.india@gmail.com'),
            f'New Contact: {name} | Nari Nakhre',
            admin_body
        )
        # Only send customer reply if email passes strict format validation
        if is_valid_email(email):
            customer_body = (
                f"Dear {name},\n\n"
                f"Thank you for reaching out to Nari Nakhre! "
                f"We have received your message and will get back to you soon.\n\n"
                f"Best regards,\nNari Nakhre Team\ninfo@narinakhre.com"
            )
            send_contact_email_async(email, 'Thank you for contacting Nari Nakhre', customer_body)

        return redirect('/retail/thank_you')
    return render_template('retail/contact.html', form_rendered_at=time.time(), recaptcha_site_key=RECAPTCHA_SITE_KEY)


@app.route('/contact', methods=['GET', 'POST'])
@app.route('/wholesale/contact', methods=['GET', 'POST'])
def wholesale_contact():
    g.site_type = 'wholesale'
    if request.method == 'POST':
        # Honeypot defense: silently discard bot submissions.
        if (request.form.get('system_verification_token') or '').strip():
            app.logger.warning(f'Bot caught on wholesale contact (honeypot): {request.form.get("email")}')
            return redirect(url_for('wholesale_thank_you'))

        # Timing trap: bots submit instantly, or replay a stale cached page
        if contact_form_is_bot(request.form.get('form_rendered_at')):
            app.logger.warning(f'Bot caught on wholesale contact (timing): {request.form.get("email")}')
            return redirect(url_for('wholesale_thank_you'))

        client_ip = request.remote_addr or 'unknown'
        if contact_ip_is_rate_limited(client_ip):
            app.logger.warning(f'Wholesale contact rate-limited: ip={client_ip}')
            return redirect(url_for('wholesale_thank_you'))

        if not verify_recaptcha(request.form.get('recaptcha_token'), remote_ip=client_ip, expected_action='wholesale_contact'):
            app.logger.warning(f'Bot caught on wholesale contact (recaptcha): {request.form.get("email")}')
            return redirect(url_for('wholesale_thank_you'))

        name = (request.form.get('name') or '').strip()
        whatsapp = (request.form.get('whatsapp') or '').strip()
        email = (request.form.get('email') or '').strip()
        message = (request.form.get('message') or '').strip()

        # Reject obviously empty/spam submissions
        if not name or not message or len(message) < 5:
            app.logger.warning(f'Wholesale contact spam rejected: name={name}, email={email}')
            return redirect(url_for('wholesale_thank_you'))

        request_id = f"NN-QT-{datetime.now().strftime('%Y%m%d%H%M%S')}-{(whatsapp[-4:] if whatsapp else '0000')}"
        quote_payload = json.dumps({
            'source': 'wholesale_contact',
            'message': message,
        })

        db_conn = get_db()
        db_conn.execute(
            'INSERT INTO quotes (request_id, name, whatsapp, email, items_json, total_amount, user_id) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (request_id, name, whatsapp, email, quote_payload, None, session.get('user_id')),
        )
        db_conn.commit()

        # Only auto-reply if email passes strict format validation (not spam/bot)
        if is_valid_email(email):
            customer_subject = 'Thank you for your quote request - Nari Nakhre Wholesale'
            customer_body = (
                f"Dear {name},\n\n"
                'Your wholesale quote request has been received successfully. '
                'Our team will review and get in touch shortly.\n\n'
                f"Request ID: {request_id}\n\n"
                'Regards,\nNari Nakhre Wholesale Team'
            )
            send_contact_email_async(email, customer_subject, customer_body)

        admin_subject = f'New Wholesale Contact/Quote Request: {request_id}'
        admin_body = (
            f"Request ID: {request_id}\n"
            f"Name: {name}\n"
            f"WhatsApp: {whatsapp}\n"
            f"Email: {email}\n\n"
            f"Message:\n{message}"
        )
        send_contact_email_async('info@narinakhre.com', admin_subject, admin_body)

        session['wholesale_contact_user'] = {'name': name, 'email': email}
        session.modified = True
        return redirect(url_for('wholesale_thank_you'))

    return render_template('wholesale/contact.html', form_rendered_at=time.time(), recaptcha_site_key=RECAPTCHA_SITE_KEY)


@app.route('/wholesale/thank_you')
def wholesale_thank_you():
    g.site_type = 'wholesale'
    user = session.pop('wholesale_contact_user', None) or {'name': 'Customer', 'email': ''}
    session.modified = True
    return render_template('wholesale/thank_you.html', user=user, contact_only=True)

# --- SITE DETECTION ---
@app.before_request
def detect_site_type():
    host = request.host.lower()
    path = request.path.lower()
    if 'wholesale' in host:
        g.site_type = 'wholesale'
    elif 'retail' in host:
        g.site_type = 'retail'
    elif path.startswith('/retail'):
        g.site_type = 'retail'
    elif path.startswith('/wholesale'):
        g.site_type = 'wholesale'
    else:
        g.site_type = 'retail'  # default to retail for shared paths like /admin, /track


VISITOR_COOKIE_NAME = 'nn_vid'
VISITOR_COOKIE_MAX_AGE = 365 * 24 * 60 * 60  # 1 year


def classify_traffic_source(referrer):
    """Best-effort traffic source label from an HTTP Referer header.
    Many in-app browsers (notably WhatsApp, and often Instagram) simply
    don't send a Referer at all for privacy reasons -- those show up as
    "Direct / Unknown" here, which is a real limitation of referrer-based
    tracking, not a bug: there's no reliable way to distinguish "typed the
    URL in" from "tapped a WhatsApp link" once the browser withholds it."""
    if not referrer:
        return 'Direct / Unknown'
    try:
        netloc = urlparse(referrer).netloc.lower()
    except Exception:
        return 'Direct / Unknown'
    if not netloc:
        return 'Direct / Unknown'
    if 'narinakhre.com' in netloc:
        return 'Internal'
    if any(d in netloc for d in ('google.', 'bing.', 'yahoo.', 'duckduckgo.')):
        return 'Search Engine'
    if 'instagram.com' in netloc:
        return 'Instagram'
    if 'facebook.com' in netloc or netloc.endswith('fb.com'):
        return 'Facebook'
    if 'wa.me' in netloc or 'whatsapp.com' in netloc:
        return 'WhatsApp'
    if 'youtube.com' in netloc or 'youtu.be' in netloc:
        return 'YouTube'
    if 'twitter.com' in netloc or netloc == 'x.com' or 't.co' in netloc:
        return 'Twitter / X'
    return netloc


@app.before_request
def track_page_view():
    """Best-effort site-visit logging for the admin Visitors dashboard --
    never allowed to break the actual page. Skips admin/static/API paths
    and non-GET requests so it only counts real storefront page loads, not
    every AJAX call. Assigns a long-lived anonymous visitor cookie (read
    here, actually set on the response in stamp_visitor_cookie below) so
    repeat visits can be counted as the same person."""
    try:
        if request.method != 'GET':
            return
        path = request.path
        skip_prefixes = ('/static', '/admin', '/api', '/update-cart', '/apply_coupon', '/remove_coupon', '/c/')
        skip_exact = ('/favicon.ico', '/robots.txt', '/sitemap.xml')
        if path in skip_exact or any(path.startswith(p) for p in skip_prefixes):
            return

        vid = request.cookies.get(VISITOR_COOKIE_NAME)
        g._nn_new_visitor = not vid
        if not vid:
            vid = str(uuid.uuid4())
        g.visitor_id = vid

        referrer = (request.referrer or '')[:500]
        source = classify_traffic_source(referrer)
        g.traffic_source = source

        db = get_db()
        db.execute(
            'INSERT INTO page_views (visitor_id, path, site_type, referrer, source, user_id) VALUES (?,?,?,?,?,?)',
            (vid, path[:255], getattr(g, 'site_type', None), referrer, source, session.get('user_id'))
        )
        db.commit()
    except Exception as e:
        app.logger.warning(f'track_page_view failed: {e}')


@app.after_request
def stamp_visitor_cookie(response):
    try:
        if getattr(g, '_nn_new_visitor', False) and getattr(g, 'visitor_id', None):
            response.set_cookie(
                VISITOR_COOKIE_NAME, g.visitor_id, max_age=VISITOR_COOKIE_MAX_AGE,
                httponly=True, samesite='Lax'
            )
    except Exception as e:
        app.logger.warning(f'stamp_visitor_cookie failed: {e}')
    return response


def render_site(template_name, **kwargs):
    site_type = getattr(g, 'site_type', 'retail')
    db = get_db()
    # For retail, fetch categories from the products table's 'category' column
    try:
        cats = db.execute(
            "SELECT DISTINCT category FROM products WHERE is_active=1 AND category IS NOT NULL AND category != '' ORDER BY category"
        ).fetchall()
        categories = [c['category'] for c in cats if c['category']]
    except Exception:
        categories = []
    kwargs['categories'] = categories
    return render_template(f"{site_type}/{template_name}", **kwargs)

@app.route('/terms')
@app.route('/privacy')
@app.route('/privacy-policy')
def terms():
    """Combined Terms & Conditions / Privacy Policy page, shared by retail and wholesale.
    Also served at /privacy and /privacy-policy so it resolves regardless of which
    URL is registered as the Privacy Policy link on the Google OAuth consent screen."""
    return render_site('terms.html')

@app.context_processor
def inject_logged_in_user():
    # This context processor runs for every render_template() call in the
    # app, including ones fired from a background thread (campaign sends,
    # welcome-email backfill) that only has an app context pushed via
    # `with app.app_context():`, not a real request context. `session` is
    # request-context-bound, so touching it without checking first raises
    # "RuntimeError: Working outside of request context" -- there's no
    # logged-in visitor to report in that case anyway, so None is correct.
    ctx = {'recaptcha_site_key': RECAPTCHA_SITE_KEY}
    if not has_request_context() or not session.get('user_id'):
        ctx['logged_in_user'] = None
    else:
        ctx['logged_in_user'] = {'name': session.get('user_name'), 'email': session.get('user_email')}
    return ctx

def generate_referral_code(db_conn):
    """8-char, human-typeable, collision-checked referral code."""
    import random
    import string
    alphabet = string.ascii_uppercase + string.digits
    for _ in range(10):
        code = ''.join(random.choices(alphabet, k=8))
        if not db_conn.execute('SELECT id FROM users WHERE referral_code=?', (code,)).fetchone():
            return code
    # Astronomically unlikely, but never loop forever
    return ''.join(random.choices(alphabet, k=12))


def generate_welcome_coupon(db_conn, user_id):
    """Create a private, single-use 15%-off welcome coupon tied to one user.

    Each recipient gets their own unique code (not one shared code), so
    usage_limit=1 genuinely means "used once by this person" rather than
    "the first of everyone who got the email".
    """
    import random
    import string
    alphabet = string.ascii_uppercase + string.digits
    code = 'WELCOME' + ''.join(random.choices(alphabet, k=6))
    for _ in range(10):
        if not db_conn.execute('SELECT id FROM coupons WHERE code=?', (code,)).fetchone():
            break
        code = 'WELCOME' + ''.join(random.choices(alphabet, k=6))

    expiry_date = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    db_conn.execute(
        "INSERT INTO coupons (code, discount_percent, min_order_amount, max_discount_amount,"
        " expiry_date, usage_limit, is_public, is_active, user_id)"
        " VALUES (?,?,?,?,?,?,0,1,?)",
        (code, 15, 299, 75, expiry_date, 1, user_id)
    )
    db_conn.commit()
    return {'code': code, 'expiry_date': expiry_date, 'discount_percent': 15,
            'max_discount_amount': 75, 'min_order_amount': 299}


def get_popular_products(db_conn, limit=4):
    """Best-selling in-stock products by units ordered (all-time), for
    marketing emails. Falls back to recently-added in-stock products if
    there isn't enough order history yet, so the email is never empty."""
    orders = db_conn.execute(
        "SELECT cart_items_json FROM order_shipping WHERE status != 'cancelled'"
    ).fetchall()
    tally = {}
    for o in orders:
        try:
            items = json.loads(o['cart_items_json'] or '[]')
        except Exception:
            items = []
        for item in items:
            sku = item.get('sku') or ''
            if not sku:
                continue
            tally[sku] = tally.get(sku, 0) + int(item.get('units', 1) or 1)
    top_skus = [sku for sku, _ in sorted(tally.items(), key=lambda kv: -kv[1])]

    products = []
    seen_skus = set()
    for sku in top_skus:
        p = db_conn.execute(
            'SELECT id, sku, name, mrp_price, retail_price, price1, image_field, stock_total'
            ' FROM products WHERE sku=? AND is_active=1', (sku,)
        ).fetchone()
        if p and (p.get('stock_total') or 0) > 0:
            products.append(dict(p))
            seen_skus.add(sku)
        if len(products) >= limit:
            break

    if len(products) < limit:
        fallback = db_conn.execute(
            'SELECT id, sku, name, mrp_price, retail_price, price1, image_field, stock_total'
            ' FROM products WHERE is_active=1 AND stock_total > 0 ORDER BY id DESC LIMIT 20'
        ).fetchall()
        for p in fallback:
            if p['sku'] in seen_skus:
                continue
            products.append(dict(p))
            seen_skus.add(p['sku'])
            if len(products) >= limit:
                break

    for p in products:
        imgs = get_product_images(p)
        p['image'] = imgs[0] if imgs else ''
    return products[:limit]


def get_trending_products(db_conn, pool_size=8, limit=4):
    """Trending in-stock products (discount % + recency, same scoring as the
    homepage shelf), for use in marketing campaign emails. Returns a random
    sample from the top pool each call, so "regenerate" in the campaign
    builder gives a genuinely different picks rather than the same 4 every
    time."""
    import random
    rows = db_conn.execute(
        'SELECT id, sku, name, mrp_price, retail_price, price1, image_field, stock_total'
        ' FROM products WHERE is_active=1 AND stock_total > 0'
    ).fetchall()
    products = [dict(r) for r in rows]
    if not products:
        return []

    max_id = max((p['id'] for p in products), default=1) or 1

    def trend_score(p):
        mrp = float(p.get('mrp_price') or 0)
        rp = float(p.get('retail_price') or p.get('price1') or 0)
        disc_pct = ((mrp - rp) / mrp * 100) if mrp and mrp > rp else 0
        recency = (p['id'] / max_id) * 100
        return disc_pct * 0.6 + recency * 0.4

    pool = sorted(products, key=trend_score, reverse=True)[:pool_size]
    picks = random.sample(pool, min(limit, len(pool)))
    for p in picks:
        imgs = get_product_images(p)
        p['image'] = imgs[0] if imgs else ''
    return picks


def get_offers_carousel_data(db):
    """Public coupons + a random sample of 3 discounted trending products --
    the same content shown in the homepage offers carousel, reused for the
    compact version on category and product-detail pages."""
    import random
    public_coupons = get_public_coupons(db)
    pool = get_trending_products(db, pool_size=8, limit=8)
    discounted = [
        p for p in pool
        if p.get('mrp_price') and p['mrp_price'] > (p.get('retail_price') or p.get('price1') or 0)
    ]
    carousel_products = random.sample(discounted, min(3, len(discounted)))
    for p in carousel_products:
        mrp = float(p['mrp_price'])
        rp = float(p.get('retail_price') or p.get('price1') or 0)
        p['disc_pct'] = round((mrp - rp) / mrp * 100)
        p['current_price'] = rp
        p['thumb'] = p.get('image') or '/static/assets/products/default.jpg'
    return public_coupons, carousel_products


def generate_personal_coupon_code(db_conn, first_name, discount_percent):
    """Per-recipient, human-readable coupon code for campaign emails, e.g.
    "NNADITI15" for a 15%-off campaign sent to Aditi. Falls back to "USER"
    if the name is blank, and appends a numeric suffix on the rare collision
    (e.g. two "Aditi"s in the same campaign)."""
    first = (first_name or '').strip().split(' ')[0] or 'USER'
    base = 'NN' + re.sub(r'[^A-Za-z0-9]', '', first)[:5].upper() + str(int(discount_percent))
    code = base
    suffix = 2
    while db_conn.execute('SELECT id FROM coupons WHERE code=?', (code,)).fetchone():
        code = f'{base}{suffix}'
        suffix += 1
    return code


def send_welcome_email(db_conn, user_id, name, email, async_send=True):
    """Send a first-time welcome email with a private one-time 15% coupon.
    Used for both fresh signups and the admin-triggered backfill for
    existing users; marks users.welcome_email_sent_at so it's never sent
    twice to the same account.

    async_send=True (the default, for live signups) fires the send in a
    detached thread so the signup HTTP response isn't held up waiting on
    Zeptomail -- welcome_email_sent_at is set immediately since we won't
    know the real outcome in time anyway. async_send=False (used by the
    backfill, which already runs in its own background thread) sends
    synchronously and only marks the user as emailed if Zeptomail actually
    confirms the send, so a rejected send can be retried on the next run
    instead of being silently marked done."""
    if not email:
        return False
    coupon = generate_welcome_coupon(db_conn, user_id)
    popular = get_popular_products(db_conn, limit=4)
    first_name = (name or 'there').strip().split(' ')[0] or 'there'

    html = render_template('retail/email_welcome.html',
                            first_name=first_name, coupon=coupon, products=popular)
    text = (
        f"Hi {first_name},\n\nWelcome to Nari Nakhre! Here's a gift for you: "
        f"use code {coupon['code']} for 15% off (up to Rs.75) on orders above Rs.299.\n"
        f"Valid until {coupon['expiry_date']}.\n\nShop now: https://narinakhre.com\n"
    )
    subject = "Welcome to Nari Nakhre — here's 15% off your first order \U0001F381"
    if async_send:
        send_contact_email_async(email, subject, text, html_body=html)
        ok = True
    else:
        ok = send_contact_email(email, subject, text, html_body=html)

    if ok:
        db_conn.execute('UPDATE users SET welcome_email_sent_at=CURRENT_TIMESTAMP WHERE id=?', (user_id,))
        db_conn.commit()
    return ok


MODEL_NUMBER_PREFIX = 'N'


def _next_model_number_suffix(suffix):
    """'AA001' -> 'AA002' ... 'AA999' -> 'AB001' ... 'ZZ999' raises (space exhausted)."""
    letters, digits = suffix[:2], suffix[2:]
    num = int(digits) + 1
    if num > 999:
        num = 1
        first, second = letters[0], letters[1]
        if second == 'Z':
            if first == 'Z':
                raise ValueError('Model number space exhausted (ZZ999 reached)')
            first, second = chr(ord(first) + 1), 'A'
        else:
            second = chr(ord(second) + 1)
        letters = first + second
    return f"{letters}{num:03d}"


def generate_model_number(db_conn):
    """Product-level identifier, one per product row, shared across all its
    size variants (bangles' per-size rows live in product_variants under one
    master sku, so this never needs to special-case sizes). Format NAA001:
    'N' is a fixed prefix; the rest is a 2-letter + 3-digit counter, fixed
    width, so plain string ordering already matches numeric order -- the
    highest existing code is just MAX(model_number)."""
    row = db_conn.execute(
        "SELECT model_number FROM products WHERE model_number IS NOT NULL "
        "ORDER BY model_number DESC LIMIT 1"
    ).fetchone()
    if not row or not row['model_number']:
        return f"{MODEL_NUMBER_PREFIX}AA001"
    suffix = _next_model_number_suffix(row['model_number'][len(MODEL_NUMBER_PREFIX):])
    return f"{MODEL_NUMBER_PREFIX}{suffix}"


# Nari Nakhre Credits -- an in-app store-credit wallet, redeemable at
# checkout. Balance is never stored as a column (avoids it drifting out of
# sync with reality); it's always the sum of this ledger, computed fresh.
COD_ROUNDING_CREDIT = 10  # flat bonus credited on top of the actual rounding excess when COD rounding goes against the customer


def round_cod_amount(amount):
    """Round a COD cash-collection amount to the nearest ₹10.

    Delhivery's COD delivery agents collect physical cash and can't
    practically make exact change, so the amount they're told to collect
    has to be a whole multiple of 10, not the exact order total.

    Returns (rounded_amount, rounded_up). rounded_up is True only when the
    customer ends up paying MORE cash than the order actually costs -- that
    gap (rounded_amount - original amount), plus a flat COD_ROUNDING_CREDIT
    bonus, is what earns them a Nari Nakhre Credits award instead of
    physical change. Rounding DOWN (customer pays less) needs no special
    handling -- it's simply a smaller cash collection, i.e. an automatic
    discount.
    """
    amount = float(amount or 0)
    base = int(amount // 10) * 10
    remainder = round(amount - base, 2)
    if remainder >= 5:
        return float(base + 10), True
    return float(base), False


def get_credit_balance(db_conn, user_id):
    if not user_id:
        return 0.0
    # Release any checkout-time holds that expired (order was never
    # confirmed) before computing the balance, so an abandoned cart doesn't
    # permanently lock those credits away.
    try:
        db_conn.execute('SELECT sweep_expired_credit_holds(?)', (user_id,))
    except Exception as e:
        app.logger.warning(f'Credit hold sweep failed for user {user_id}: {e}')
    row = db_conn.execute(
        'SELECT COALESCE(SUM(amount), 0) as bal FROM credit_transactions WHERE user_id=?',
        (user_id,)
    ).fetchone()
    return float(row['bal']) if row and row['bal'] is not None else 0.0


def award_credits(db_conn, user_id, amount, reason, internal_order_id=None):
    """Add a positive ledger entry. Caller is responsible for conn.commit()."""
    if not user_id or amount <= 0:
        return
    db_conn.execute(
        'INSERT INTO credit_transactions (user_id, amount, reason, internal_order_id) VALUES (?,?,?,?)',
        (user_id, amount, reason, internal_order_id),
    )


def reserve_credits(db_conn, user_id, amount, internal_order_id, reason='hold_for_order'):
    """Atomically reserve `amount` credits against a not-yet-confirmed order.

    Returns True if the reservation succeeded, False if the user doesn't
    actually have enough available (e.g. a concurrent checkout in another
    tab/device already claimed them). The check-then-insert happens inside
    a single Postgres function call (reserve_credits_atomic) guarded by an
    advisory lock, so two simultaneous callers for the same user_id are
    genuinely serialized by the database rather than racing on a
    read-then-write from Python. Caller is responsible for conn.commit().
    """
    if not user_id or amount <= 0:
        return True
    row = db_conn.execute(
        "SELECT reserve_credits_atomic(?, ?, ?, ?) as ok",
        (user_id, amount, internal_order_id, reason),
    ).fetchone()
    return bool(row['ok']) if row else False


def finalize_credit_redemption(db_conn, user_id, amount, internal_order_id):
    """Convert a checkout-time hold into a final, confirmed debit.

    Normally just relabels the existing 'hold_for_order' ledger row created
    by reserve_credits() at checkout time -- no new row, so this never
    double-debits. Falls back to a fresh atomic reservation (tagged as
    already-redeemed) only if that hold already expired, which can happen
    if confirmation took longer than the hold's window; if the balance
    isn't there anymore at that point, the credits are simply not honored
    for this order rather than blocking the confirmation. Caller is
    responsible for conn.commit().
    """
    if not user_id or amount <= 0:
        return
    row = db_conn.execute("SELECT finalize_credit_hold(?) as ok", (internal_order_id,)).fetchone()
    if row and row['ok']:
        return
    ok = reserve_credits(db_conn, user_id, amount, internal_order_id, reason='redeemed_at_checkout')
    if not ok:
        app.logger.warning(
            f"Could not finalize {amount} credits for order {internal_order_id}: "
            f"hold had already expired and insufficient balance remains."
        )


@app.route('/auth/google/login')
def google_login():
    site_home = '/wholesale' if g.site_type == 'wholesale' else '/retail'
    session['post_login_redirect'] = request.referrer or site_home
    ref_code = (request.args.get('ref') or '').strip()
    if ref_code:
        session['pending_referral_code'] = ref_code
    # redirect_uri is built per-request (not a fixed APP_BASE_URL env var)
    # because this one Flask app serves four different domains
    # (narinakhre.com, wholesale.narinakhre.com, and the two test-* hosts);
    # url_for(_external=True) picks the right one automatically.
    provider = auth_providers.get_auth_provider('google')
    redirect_uri = url_for('google_callback', _external=True)
    return redirect(provider.get_auth_url(redirect_uri))

@app.route('/auth/google/callback')
def google_callback():
    site_home = '/wholesale' if g.site_type == 'wholesale' else '/retail'
    provider = auth_providers.get_auth_provider('google')
    try:
        token = provider.exchange_code()
        userinfo = provider.get_user_info(token)
    except Exception as e:
        app.logger.warning(f'Google OAuth callback failed: {e}')
        flash('Sign-in was cancelled or failed. Please try again.', 'error')
        return redirect(session.pop('post_login_redirect', site_home))
    google_sub = userinfo.get('sub')
    name = userinfo.get('name') or ''
    email = userinfo.get('email') or ''
    picture = userinfo.get('picture') or ''

    if not google_sub:
        app.logger.warning('Google callback missing sub claim; aborting login.')
        return redirect(session.pop('post_login_redirect', url_for('index')))

    db_conn = get_db()
    existing = db_conn.execute('SELECT id FROM users WHERE google_sub=?', (google_sub,)).fetchone()
    if existing:
        db_conn.execute(
            'UPDATE users SET name=?, email=?, picture_url=? WHERE google_sub=?',
            (name, email, picture, google_sub),
        )
    else:
        # New account -- assign a referral code, and credit whoever referred
        # them, if a valid ?ref= code was carried through from google_login.
        referred_by_id = None
        pending_ref = session.pop('pending_referral_code', None)
        if pending_ref:
            referrer = db_conn.execute('SELECT id FROM users WHERE referral_code=?', (pending_ref,)).fetchone()
            if referrer:
                referred_by_id = referrer['id']
        new_code = generate_referral_code(db_conn)
        db_conn.execute(
            'INSERT INTO users (google_sub, name, email, picture_url, referral_code, referred_by) VALUES (?,?,?,?,?,?)',
            (google_sub, name, email, picture, new_code, referred_by_id),
        )
    db_conn.commit()
    user_row = db_conn.execute('SELECT id FROM users WHERE google_sub=?', (google_sub,)).fetchone()

    session['user_id'] = user_row['id']
    session['user_name'] = name
    session['user_email'] = email

    if not existing:
        try:
            send_welcome_email(db_conn, user_row['id'], name, email)
        except Exception as e:
            app.logger.warning(f'Welcome email failed for {email}: {e}')
        notify_admin_new_user(db_conn, name, email, 'Google')

    return redirect(session.pop('post_login_redirect', '/'))

@app.route('/auth/logout')
def logout():
    session.pop('user_id', None)
    session.pop('user_name', None)
    session.pop('user_email', None)
    site_home = '/wholesale' if g.site_type == 'wholesale' else '/retail'
    return redirect(request.referrer or site_home)


EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def _log_in_user(db_conn, user_id):
    user_row = db_conn.execute('SELECT id, name, email FROM users WHERE id=?', (user_id,)).fetchone()
    session['user_id'] = user_row['id']
    session['user_name'] = user_row['name']
    session['user_email'] = user_row['email']


@app.route('/auth/email/signup', methods=['POST'])
def email_signup():
    # Not restricted to retail -- Google sign-in already works on both
    # storefronts against the same users table, so email sign-in matches it.
    data = request.get_json(silent=True) or request.form or {}
    full_name = (data.get('full_name') or '').strip()
    phone = (data.get('phone') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not verify_recaptcha(data.get('recaptcha_token'), remote_ip=request.remote_addr, expected_action='email_signup'):
        return jsonify({'status': 'error', 'message': 'Verification failed. Please refresh and try again.'}), 400

    if not (full_name and phone and email and password):
        return jsonify({'status': 'error', 'message': 'Please fill in all fields.'}), 400
    if not EMAIL_RE.match(email):
        return jsonify({'status': 'error', 'message': 'Please enter a valid email address.'}), 400
    if len(password) < 8:
        return jsonify({'status': 'error', 'message': 'Password must be at least 8 characters.'}), 400

    db_conn = get_db()
    existing = db_conn.execute('SELECT id, google_sub, password_hash FROM users WHERE LOWER(email)=?', (email,)).fetchone()
    if existing:
        if existing['password_hash']:
            return jsonify({'status': 'error', 'message': 'An account with this email already exists. Please sign in instead.'}), 400
        if existing['google_sub']:
            return jsonify({'status': 'error', 'message': 'This email is registered via Google Sign-In. Please use "Continue with Google" instead.'}), 400

    password_hash = generate_password_hash(password)

    referred_by_id = None
    pending_ref = session.pop('pending_referral_code', None)
    if pending_ref:
        referrer = db_conn.execute('SELECT id FROM users WHERE referral_code=?', (pending_ref,)).fetchone()
        if referrer:
            referred_by_id = referrer['id']
    new_code = generate_referral_code(db_conn)

    db_conn.execute(
        'INSERT INTO users (name, phone, email, password_hash, referral_code, referred_by) VALUES (?,?,?,?,?,?)',
        (full_name, phone, email, password_hash, new_code, referred_by_id),
    )
    db_conn.commit()
    user_row = db_conn.execute('SELECT id FROM users WHERE LOWER(email)=?', (email,)).fetchone()
    _log_in_user(db_conn, user_row['id'])

    try:
        send_welcome_email(db_conn, user_row['id'], full_name, email)
    except Exception as e:
        app.logger.warning(f'Welcome email failed for {email}: {e}')
    notify_admin_new_user(db_conn, full_name, email, 'Email')

    redirect_url = session.pop('post_login_redirect', None) or url_for('index')
    return jsonify({'status': 'success', 'redirect': redirect_url})


@app.route('/auth/email/login', methods=['POST'])
def email_login():
    data = request.get_json(silent=True) or request.form or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not verify_recaptcha(data.get('recaptcha_token'), remote_ip=request.remote_addr, expected_action='email_login'):
        return jsonify({'status': 'error', 'message': 'Verification failed. Please refresh and try again.'}), 400

    if not (email and password):
        return jsonify({'status': 'error', 'message': 'Please enter your email and password.'}), 400

    db_conn = get_db()
    user_row = db_conn.execute('SELECT id, password_hash, google_sub FROM users WHERE LOWER(email)=?', (email,)).fetchone()
    if not user_row or not user_row['password_hash']:
        if user_row and user_row['google_sub']:
            return jsonify({'status': 'error', 'message': 'This email is registered via Google Sign-In. Please use "Continue with Google" instead.'}), 400
        return jsonify({'status': 'error', 'message': 'Incorrect email or password.'}), 400
    if not check_password_hash(user_row['password_hash'], password):
        return jsonify({'status': 'error', 'message': 'Incorrect email or password.'}), 400

    _log_in_user(db_conn, user_row['id'])
    redirect_url = session.pop('post_login_redirect', None) or url_for('index')
    return jsonify({'status': 'success', 'redirect': redirect_url})


@app.route('/profile')
def profile():
    if g.site_type != 'retail':
        return redirect('/')
    if not session.get('user_id'):
        session['post_login_redirect'] = url_for('profile')
        return redirect(url_for('google_login'))

    db_conn = get_db()
    user_id = session['user_id']
    user_row = db_conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()

    # Backfill a referral code for accounts created before this feature existed.
    if user_row and not user_row['referral_code']:
        new_code = generate_referral_code(db_conn)
        db_conn.execute('UPDATE users SET referral_code=? WHERE id=?', (new_code, user_id))
        db_conn.commit()
        user_row = db_conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()

    orders = db_conn.execute(
        'SELECT * FROM order_shipping WHERE user_id=? ORDER BY created_at DESC',
        (user_id,),
    ).fetchall()

    # "Money saved" = coupon/discount savings actually applied at checkout.
    # This does not include MRP-vs-selling-price markdowns, which aren't
    # captured per-order, so it's a conservative (understated), exact figure
    # rather than an approximation.
    total_saved = sum(float(o['discount_amount'] or 0) for o in orders)

    referral_row = db_conn.execute(
        'SELECT COUNT(*) as c FROM users WHERE referred_by=?', (user_id,)
    ).fetchone()
    referral_count = referral_row['c'] if referral_row else 0

    referral_code = user_row['referral_code'] if user_row else ''
    referral_link = f"{request.url_root.rstrip('/')}/retail?ref={referral_code}" if referral_code else ''

    credit_balance = get_credit_balance(db_conn, user_id)
    credit_history = db_conn.execute(
        'SELECT * FROM credit_transactions WHERE user_id=? ORDER BY created_at DESC LIMIT 20',
        (user_id,),
    ).fetchall()

    addresses = db_conn.execute(
        'SELECT * FROM user_addresses WHERE user_id=? ORDER BY is_default DESC, created_at DESC',
        (user_id,),
    ).fetchall()
    active_tab = request.args.get('tab', 'orders')
    if active_tab not in ('orders', 'addresses'):
        active_tab = 'orders'

    return render_site(
        'profile.html',
        user=user_row,
        orders=orders,
        total_saved=total_saved,
        referral_code=referral_code,
        referral_link=referral_link,
        referral_count=referral_count,
        credit_balance=credit_balance,
        credit_history=credit_history,
        addresses=addresses,
        active_tab=active_tab,
    )


@app.route('/profile/addresses/add', methods=['POST'])
def add_address():
    if not session.get('user_id'):
        return redirect(url_for('google_login'))
    db_conn = get_db()
    user_id = session['user_id']

    nickname = (request.form.get('nickname') or '').strip()
    address_type = (request.form.get('address_type') or 'Home').strip()
    recipient_name = (request.form.get('recipient_name') or '').strip()
    phone = (request.form.get('phone') or '').strip()
    email = (request.form.get('email') or '').strip()
    address_line = (request.form.get('address_line') or '').strip()
    city = (request.form.get('city') or '').strip()
    state = (request.form.get('state') or '').strip()
    pincode = (request.form.get('pincode') or '').strip()
    make_default = request.form.get('is_default') == 'on'

    if not (nickname and recipient_name and phone and address_line and city and state and pincode):
        flash('Please fill in all required address fields.')
        return redirect(url_for('profile', tab='addresses'))

    existing_count = db_conn.execute(
        'SELECT COUNT(*) as c FROM user_addresses WHERE user_id=?', (user_id,)
    ).fetchone()['c']
    is_default = 1 if (make_default or existing_count == 0) else 0
    if is_default:
        db_conn.execute('UPDATE user_addresses SET is_default=0 WHERE user_id=?', (user_id,))

    db_conn.execute(
        '''INSERT INTO user_addresses
           (user_id, nickname, address_type, recipient_name, phone, email,
            address_line, city, state, pincode, is_default)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (user_id, nickname, address_type, recipient_name, phone, email or None,
         address_line, city, state, pincode, is_default)
    )
    db_conn.commit()
    flash(f'Address "{nickname}" added.')
    return redirect(url_for('profile', tab='addresses'))


@app.route('/profile/addresses/<int:address_id>/edit', methods=['POST'])
def edit_address(address_id):
    if not session.get('user_id'):
        return redirect(url_for('google_login'))
    db_conn = get_db()
    user_id = session['user_id']

    owned = db_conn.execute(
        'SELECT id FROM user_addresses WHERE id=? AND user_id=?', (address_id, user_id)
    ).fetchone()
    if not owned:
        flash('Address not found.')
        return redirect(url_for('profile', tab='addresses'))

    nickname = (request.form.get('nickname') or '').strip()
    address_type = (request.form.get('address_type') or 'Home').strip()
    recipient_name = (request.form.get('recipient_name') or '').strip()
    phone = (request.form.get('phone') or '').strip()
    email = (request.form.get('email') or '').strip()
    address_line = (request.form.get('address_line') or '').strip()
    city = (request.form.get('city') or '').strip()
    state = (request.form.get('state') or '').strip()
    pincode = (request.form.get('pincode') or '').strip()
    make_default = request.form.get('is_default') == 'on'

    if not (nickname and recipient_name and phone and address_line and city and state and pincode):
        flash('Please fill in all required address fields.')
        return redirect(url_for('profile', tab='addresses'))

    if make_default:
        db_conn.execute('UPDATE user_addresses SET is_default=0 WHERE user_id=?', (user_id,))

    db_conn.execute(
        '''UPDATE user_addresses SET nickname=?, address_type=?, recipient_name=?, phone=?, email=?,
           address_line=?, city=?, state=?, pincode=?, is_default=?
           WHERE id=? AND user_id=?''',
        (nickname, address_type, recipient_name, phone, email or None,
         address_line, city, state, pincode, 1 if make_default else 0, address_id, user_id)
    )
    db_conn.commit()
    flash(f'Address "{nickname}" updated.')
    return redirect(url_for('profile', tab='addresses'))


@app.route('/profile/addresses/<int:address_id>/delete', methods=['POST'])
def delete_address(address_id):
    if not session.get('user_id'):
        return redirect(url_for('google_login'))
    db_conn = get_db()
    db_conn.execute(
        'DELETE FROM user_addresses WHERE id=? AND user_id=?', (address_id, session['user_id'])
    )
    db_conn.commit()
    flash('Address removed.')
    return redirect(url_for('profile', tab='addresses'))


@app.route('/profile/addresses/<int:address_id>/set-default', methods=['POST'])
def set_default_address(address_id):
    if not session.get('user_id'):
        return redirect(url_for('google_login'))
    db_conn = get_db()
    user_id = session['user_id']
    owned = db_conn.execute(
        'SELECT id FROM user_addresses WHERE id=? AND user_id=?', (address_id, user_id)
    ).fetchone()
    if owned:
        db_conn.execute('UPDATE user_addresses SET is_default=0 WHERE user_id=?', (user_id,))
        db_conn.execute('UPDATE user_addresses SET is_default=1 WHERE id=?', (address_id,))
        db_conn.commit()
    return redirect(url_for('profile', tab='addresses'))

@app.route('/my-orders')
def my_orders():
    return redirect(url_for('profile'))

@app.route('/my-quotes')
def my_quotes():
    if g.site_type != 'wholesale':
        return redirect('/')
    if not session.get('user_id'):
        return redirect(url_for('google_login'))
    db_conn = get_db()
    quotes = db_conn.execute(
        'SELECT * FROM quotes WHERE user_id=? ORDER BY created_at DESC',
        (session['user_id'],),
    ).fetchall()
    return render_site('history.html', quotes=quotes)

# --- IMAGE HELPERS ---
def get_supabase_image_urls(sku):
    """Build Supabase image URLs for a SKU — _1 through _9."""
    base = (os.environ.get('SUPABASE_URL') or '').rstrip('/')
    if not base:
        return []
    bucket = 'products'
    return [f"{base}/storage/v1/object/public/{bucket}/{sku}_{i}.webp" for i in range(1, 10)]


def get_product_images(p_dict):
    """
    Return image URL list for a product.
    Images are stored in Supabase as {SKU}_1.webp, {SKU}_2.webp etc.
    Uses image_field as the primary/first image, then fills in the
    rest from the SKU pattern. The template uses onerror to hide
    broken images, so returning extra URLs that don't exist is safe.
    """
    sku = p_dict.get('sku', '')
    image_field = (p_dict.get('image_field') or '').strip()

    # Get all SKU-based URLs (_1 through _9)
    sku_urls = get_supabase_image_urls(sku) if sku else []

    if image_field.startswith('http'):
        # Put image_field first, then add remaining SKU urls
        others = [u for u in sku_urls if u != image_field]
        return [image_field] + others

    if sku_urls:
        return sku_urls

    return ['/static/assets/products/default.jpg']


def get_product_tiers(p_dict):
    """Extract wholesale tier pricing from a product dict."""
    tiers = []
    for i in range(1, 4):
        qty = p_dict.get(f'quantity{i}')
        price = p_dict.get(f'price{i}')
        if qty and price:
            try:
                if int(qty) > 0 and float(price) > 0:
                    tiers.append({'qty': int(qty), 'price': float(price)})
            except Exception:
                continue
    return tiers if tiers else [{'qty': 1, 'price': 0}]


BANGLE_SIZES = ['2.4', '2.6', '2.8']


def is_bangle_product(p_dict):
    """Same detection rule used in templates: category/sub_category/name mentions 'bangle'."""
    for field in ('category', 'sub_category', 'name'):
        val = (p_dict.get(field) or '').lower()
        if 'bangle' in val:
            return True
    return False


def log_product_event(db_conn, sku, event_type, visitor_id=None, source=None):
    """Best-effort analytics ping -- must never break the page it's called from.
    visitor_id/source default to None for backward compatibility, but call
    sites on a real request should pass g.visitor_id/g.traffic_source
    (set by the track_page_view before_request hook) so the admin Product
    Visits page can report unique visitors and traffic sources per SKU."""
    if not sku:
        return
    try:
        db_conn.execute(
            'INSERT INTO product_events (sku, event_type, visitor_id, source) VALUES (?,?,?,?)',
            (sku, event_type, visitor_id, source)
        )
        db_conn.commit()
    except Exception as e:
        app.logger.warning(f'log_product_event failed for {sku}/{event_type}: {e}')


def log_admin_event(db_conn, event_type, title, detail=None, related_id=None):
    """Best-effort admin Inbox notification -- must never break the page or
    background job that triggered it. Feeds /admin/events (new user
    registrations, new orders, orders stuck unaccepted past a threshold,
    deliveries)."""
    try:
        db_conn.execute(
            'INSERT INTO admin_events (event_type, title, detail, related_id) VALUES (?,?,?,?)',
            (event_type, title, detail, str(related_id) if related_id is not None else None)
        )
        db_conn.commit()
    except Exception as e:
        app.logger.warning(f'log_admin_event failed for {event_type}: {e}')


def notify_admin_new_user(db_conn, name, email, method):
    """Logs an admin Inbox event AND emails the admin whenever someone new
    registers (both explicitly requested) -- separate try/excepts so a
    failure in one never blocks the other, and neither ever blocks login."""
    try:
        log_admin_event(db_conn, 'new_user', f'New user registered: {name or email}',
                         detail=f'{email} — signed up via {method}')
    except Exception as e:
        app.logger.warning(f'notify_admin_new_user event log failed: {e}')
    try:
        send_contact_email_async(
            ADMIN_EMAIL, f'New user registered: {name or email}',
            f'A new user just created an account on Nari Nakhre.\n\n'
            f'Name: {name}\nEmail: {email}\nSign-up method: {method}\n'
        )
    except Exception as e:
        app.logger.warning(f'Admin new-user email failed: {e}')


def get_variant_sku(master_sku, size):
    return f"{master_sku}-{size.replace('.', '')}"


def get_bangle_size_stock(db, master_sku):
    """
    Return {size: stock_total} for a bangle's per-size variants.
    On first call for a given master SKU (no variant rows yet), lazily creates the
    3 size variants by splitting the product's current total stock evenly across
    2.4/2.6/2.8 — so no separate one-off migration is needed for existing products.
    """
    rows = db.execute(
        'SELECT size, stock_total FROM product_variants WHERE master_sku=?', (master_sku,)
    ).fetchall()
    if rows:
        return {r['size']: (r['stock_total'] or 0) for r in rows}

    prod = db.execute('SELECT stock_total FROM products WHERE sku=?', (master_sku,)).fetchone()
    total = (prod['stock_total'] or 0) if prod else 0
    base, remainder = divmod(total, len(BANGLE_SIZES))
    size_map = {}
    for i, size in enumerate(BANGLE_SIZES):
        stock = base + (1 if i < remainder else 0)
        variant_sku = get_variant_sku(master_sku, size)
        db.execute(
            'INSERT INTO product_variants (master_sku, variant_sku, size, stock_total) '
            'VALUES (?, ?, ?, ?) ON CONFLICT (master_sku, size) DO NOTHING',
            (master_sku, variant_sku, size, stock)
        )
        size_map[size] = stock
    db.commit()
    return size_map


def _tokenize_search_text(text):
    return re.findall(r'[a-z0-9]+', (text or '').lower())


def search_products(db, q, limit=None):
    """Word-aware product search, ranked by relevance.

    A plain `LIKE '%term%'` substring search (the old approach) makes a short
    word like "red" match almost everything, since it also matches inside
    unrelated words such as "embroidered" or "bordered". Here every query
    word must appear as a whole word in the product's name/category/
    description/key_features (or as a substring specifically within the SKU
    or model number, where partial-code matches are actually useful), which
    is what makes results for common words precise.
    """
    query_tokens = _tokenize_search_text(q)
    if not query_tokens:
        return []
    q_low = q.strip().lower()

    rows = db.execute(
        "SELECT id, sku, model_number, name, category, sub_category, description,"
        " key_features, retail_price, mrp_price, image_field, stock_total, size"
        " FROM products WHERE is_active = 1"
    ).fetchall()

    scored = []
    for r in rows:
        p = dict(r)
        name = (p.get('name') or '').lower()
        sku = (p.get('sku') or '').lower()
        model = (p.get('model_number') or '').lower()
        category = (p.get('category') or '').lower()
        sub_category = (p.get('sub_category') or '').lower()
        description = (p.get('description') or '').lower()
        key_features = (p.get('key_features') or '').lower()

        name_tokens = set(_tokenize_search_text(name))
        cat_tokens = set(_tokenize_search_text(category)) | set(_tokenize_search_text(sub_category))
        other_tokens = set(_tokenize_search_text(description)) | set(_tokenize_search_text(key_features))
        word_tokens = name_tokens | cat_tokens | other_tokens

        matched = sum(1 for qt in query_tokens if qt in word_tokens or qt in sku or qt in model)
        if matched < len(query_tokens):
            continue

        score = 0
        if name == q_low:
            score += 1000
        elif name.startswith(q_low):
            score += 500
        score += sum(50 for qt in query_tokens if qt in name_tokens)
        score += sum(20 for qt in query_tokens if qt in cat_tokens)
        if q_low and q_low in sku:
            score += 30
        if q_low and q_low in model:
            score += 30
        score += sum(5 for qt in query_tokens if qt in other_tokens)
        scored.append((score, p))

    scored.sort(key=lambda item: -item[0])
    results = [p for _, p in scored]
    return results[:limit] if limit else results


def apply_sort_and_filters(products, sort=None, size_filter=None, in_stock_only=False):
    """Shared sort/filter logic for category and search result listings.

    Handles both sizing schemes used across the catalogue: bangles carry
    per-size stock in `size_stock` (from get_bangle_size_stock), while other
    sized products use a plain comma-separated `size` column.
    """
    def eff_price(p):
        return float(p.get('retail_price') or p.get('price1') or 0)

    def product_sizes(p):
        if p.get('size_stock'):
            return set(p['size_stock'].keys())
        if p.get('size'):
            return {s.strip() for s in p['size'].split(',') if s.strip()}
        return set()

    def in_stock(p):
        if p.get('size_stock'):
            return any((v or 0) > 0 for v in p['size_stock'].values())
        return bool(p.get('stock_total') and p['stock_total'] > 0)

    all_sizes = sorted({s for p in products for s in product_sizes(p)})

    filtered = products
    if in_stock_only:
        filtered = [p for p in filtered if in_stock(p)]
    if size_filter:
        filtered = [p for p in filtered if size_filter in product_sizes(p)]

    if sort == 'name_asc':
        filtered = sorted(filtered, key=lambda p: (p.get('name') or '').lower())
    elif sort == 'price_asc':
        filtered = sorted(filtered, key=eff_price)
    elif sort == 'price_desc':
        filtered = sorted(filtered, key=eff_price, reverse=True)

    return filtered, all_sizes


def get_public_coupons(db):
    """Active, non-expired, non-maxed-out coupons marked public — safe to surface to shoppers."""
    today_str = datetime.now().strftime('%Y-%m-%d')
    return db.execute(
        "SELECT * FROM coupons WHERE is_active=1 AND is_public=1"
        " AND (expiry_date IS NULL OR expiry_date >= ?)"
        " AND (usage_limit IS NULL OR usage_limit=0 OR times_used < usage_limit)"
        " ORDER BY discount_percent DESC, id DESC",
        (today_str,)
    ).fetchall()


def get_random_hero_images(db, count=4):
    """Pick random product images from Supabase for hero banners."""
    rows = db.execute(
        "SELECT image_field, sku FROM products WHERE image_field IS NOT NULL AND image_field LIKE 'http%' ORDER BY RANDOM() LIMIT ?",
        (count,)
    ).fetchall()
    images = [r['image_field'] for r in rows if r['image_field']]
    # If not enough from DB, build from SKUs
    if len(images) < count:
        skus = db.execute("SELECT sku FROM products ORDER BY RANDOM() LIMIT ?", (count,)).fetchall()
        for row in skus:
            url = get_supabase_image_urls(row['sku'])
            if url:
                images.append(url[0])
            if len(images) >= count:
                break
    return images[:count]


# --- KEEP-ALIVE: Ping Supabase to prevent free plan pausing ---
import threading
import time as _time

def _supabase_keepalive():
    """Background thread that pings Supabase every 3 days to keep the project active."""
    _time.sleep(30)  # Wait 30 seconds after startup before first ping
    while True:
        try:
            client = get_supabase()
            client.rpc('execute_sql', {'query': 'SELECT 1'}).execute()
            app.logger.info('Supabase keep-alive ping sent.')
        except Exception as e:
            app.logger.warning(f'Supabase keep-alive failed: {e}')
        _time.sleep(3 * 24 * 60 * 60)  # 3 days

# Start keepalive only in the main Gunicorn worker process
if os.environ.get('SERVER_SOFTWARE', '').startswith('gunicorn') or True:
    _t = threading.Thread(target=_supabase_keepalive, daemon=True)
    _t.start()


# --- ORDER WATCHDOG: notify admin when orders sit unaccepted too long, and
# detect delivery. There's no Delhivery webhook wired up (confirmed -- see
# api_track_shipment, which only fetches live status on demand), so polling
# is the only way this app finds out an order has shipped/delivered. ---
ORDER_UNACCEPTED_THRESHOLDS_HOURS = [2, 6, 12, 24]
ORDER_WATCHDOG_INTERVAL_SECONDS = 15 * 60  # 15 minutes


def _parse_db_timestamp(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    try:
        return datetime.fromisoformat(str(value).replace('Z', '+00:00')).replace(tzinfo=None)
    except Exception:
        return None


def _check_unaccepted_orders(db):
    """Fires an admin Inbox event (once per order per threshold, checked via
    a prior admin_events row) for any order that's been sitting unaccepted
    longer than each of the 2h/6h/12h/24h thresholds."""
    now = datetime.now()
    unaccepted = db.execute(
        "SELECT internal_order_id, consignee_name, total_amount, created_at FROM order_shipping "
        "WHERE status IN ('pending', 'paid', 'cod_confirmed')"
    ).fetchall()
    for order in unaccepted:
        created_at = _parse_db_timestamp(order.get('created_at'))
        if not created_at:
            continue
        age_hours = (now - created_at).total_seconds() / 3600
        for threshold in ORDER_UNACCEPTED_THRESHOLDS_HOURS:
            if age_hours < threshold:
                continue
            event_type = f'order_not_accepted_{threshold}h'
            already = db.execute(
                'SELECT id FROM admin_events WHERE event_type=? AND related_id=?',
                (event_type, order['internal_order_id'])
            ).fetchone()
            if already:
                continue
            log_admin_event(
                db, event_type,
                f"Order {order['internal_order_id']} not accepted for {threshold}+ hours",
                detail=f"{order.get('consignee_name') or ''} — ₹{float(order.get('total_amount') or 0):.0f}",
                related_id=order['internal_order_id']
            )


def _check_delivered_orders(db):
    """Polls each order's own courier (Delhivery or Shiprocket -- whichever
    actually created that shipment, not just whoever's active right now) for
    any waybill that isn't already marked delivered/cancelled, and fires an
    admin Inbox event + updates status the moment the courier reports it
    delivered."""
    providers_by_partner = {}
    in_transit = db.execute(
        "SELECT internal_order_id, consignee_name, delhivery_waybill, courier_partner FROM order_shipping "
        "WHERE delhivery_waybill IS NOT NULL AND delhivery_waybill != '' "
        "AND status NOT IN ('delivered', 'cancelled')"
    ).fetchall()
    for order in in_transit:
        waybill = order.get('delhivery_waybill')
        if not waybill:
            continue
        partner_name = order.get('courier_partner') or 'delhivery'
        if partner_name not in providers_by_partner:
            providers_by_partner[partner_name] = get_courier(partner_name)[1]
        provider = providers_by_partner[partner_name]
        try:
            result = provider.track_shipment(waybill)
        except Exception as e:
            app.logger.warning(f'Order watchdog: tracking fetch failed for {waybill}: {e}')
            continue
        if not result or not result.get('status'):
            continue
        current_status = (result.get('current_status') or '').lower()
        status_type = (result.get('status_type') or '').upper()
        if 'deliver' in current_status or status_type == 'DL':
            db.execute("UPDATE order_shipping SET status='delivered' WHERE internal_order_id=?",
                       (order['internal_order_id'],))
            db.commit()
            log_admin_event(
                db, 'order_delivered', f"Order delivered: {order['internal_order_id']}",
                detail=order.get('consignee_name') or '', related_id=order['internal_order_id']
            )


SHIPROCKET_AUTH_ALERT_INTERVAL_HOURS = 24


def _check_shiprocket_auth(db):
    """If Shiprocket is enabled but its stored credentials are failing to
    authenticate, emails the admin -- repeating once every 24 hours until a
    login succeeds again. Note this is NOT about the short-lived bearer
    token expiring -- get_courier()/get_enabled_couriers() already refresh
    that transparently on every use. This only fires when even a *fresh*
    login attempt fails (wrong/changed password, Shiprocket account issue,
    etc.), since that's the one Shiprocket auth failure the app can't
    self-heal from.

    Re-alert timing is tracked via the most recent 'shiprocket_auth_failed'
    admin_events row, mirroring how _check_unaccepted_orders tracks what's
    already been alerted on -- no separate schedule table needed."""
    partner = db.execute(
        "SELECT dp.id FROM delivery_partners dp WHERE dp.name='shiprocket' AND dp.is_enabled=1"
    ).fetchone()
    if not partner:
        return  # not enabled -- nothing to monitor

    _name, provider = get_courier('shiprocket')
    if getattr(provider, 'token', None):
        return  # authenticated fine

    last_alert = db.execute(
        "SELECT created_at FROM admin_events WHERE event_type='shiprocket_auth_failed' "
        "ORDER BY created_at DESC LIMIT 1"
    ).fetchone()
    if last_alert:
        last_at = _parse_db_timestamp(last_alert.get('created_at'))
        if last_at and (datetime.now() - last_at).total_seconds() < SHIPROCKET_AUTH_ALERT_INTERVAL_HOURS * 3600:
            return  # already alerted within the last day

    error_msg = getattr(provider, 'last_error', None) or 'Login failed for an unknown reason'
    log_admin_event(db, 'shiprocket_auth_failed', 'Shiprocket login is failing', detail=error_msg)
    try:
        send_contact_email_async(
            ADMIN_EMAIL, '⚠️ Shiprocket login is failing — courier may be unusable',
            f"Shiprocket is enabled in the admin panel, but the stored credentials are "
            f"failing to authenticate.\n\nError: {error_msg}\n\n"
            f"This usually means the Shiprocket account password was changed, or there's "
            f"an issue with the account on Shiprocket's side. Until this is fixed, checkout "
            f"will fall back to Delhivery (if enabled) or fail to fetch shipping rates.\n\n"
            f"Update the credentials at: https://narinakhre.com/admin/delivery-partners\n\n"
            f"You'll get this reminder once a day until a login succeeds again."
        )
    except Exception as e:
        app.logger.warning(f'Shiprocket auth-failure admin email failed: {e}')


def _order_watchdog():
    """Background thread: periodically checks for orders stuck unaccepted
    past 2/6/12/24 hours, polls couriers for delivery confirmation, and
    checks Shiprocket auth health."""
    _time.sleep(60)  # let the app finish starting up first
    while True:
        try:
            with app.app_context():
                db = get_db()
                _check_unaccepted_orders(db)
                _check_delivered_orders(db)
                _check_shiprocket_auth(db)
        except Exception as e:
            app.logger.warning(f'Order watchdog run failed: {e}')
        _time.sleep(ORDER_WATCHDOG_INTERVAL_SECONDS)


_watchdog_thread = threading.Thread(target=_order_watchdog, daemon=True)
_watchdog_thread.start()


# --- ROUTES: HOME & CATEGORY ---
@app.route('/')
@app.route('/retail')
@app.route('/wholesale')
def index():
    if request.path.startswith('/retail'):
        g.site_type = 'retail'
    elif request.path.startswith('/wholesale'):
        g.site_type = 'wholesale'
    elif request.path == '/':
        # Root domain — detect from hostname
        # narinakhre.com → retail, wholesale.narinakhre.com → wholesale
        host = request.host.lower()
        if 'wholesale' in host:
            g.site_type = 'wholesale'
        else:
            g.site_type = 'retail'

    db = get_db()
    hero_images = get_random_hero_images(db, count=4)

    if g.site_type == 'retail':
        products = db.execute(
            'SELECT * FROM products WHERE is_active=1'
            ' ORDER BY CASE WHEN stock_total > 0 THEN 0 ELSE 1 END, id DESC'
        ).fetchall()

        # Build grouped products — category sorted by TOTAL product count (most first)
        grouped_products = {}
        for p in products:
            cat = p['category'] or 'New Arrivals'
            if cat not in grouped_products:
                grouped_products[cat] = []
            p_dict = dict(p)
            p_dict['images'] = get_product_images(p_dict)
            p_dict['tiers'] = get_product_tiers(p_dict)
            if is_bangle_product(p_dict):
                p_dict['size_stock'] = get_bangle_size_stock(db, p_dict['sku'])
            grouped_products[cat].append(p_dict)

        # Sort by total number of products in category (not just in-stock)
        # so Bangles (most listings) always appears first regardless of stock levels
        grouped_products = dict(
            sorted(grouped_products.items(), key=lambda x: len(x[1]), reverse=True)
        )

        # cat_counts for trending: in-stock only
        cat_counts = {
            cat: sum(1 for p in prods if p.get('stock_total', 0) and p['stock_total'] > 0)
            for cat, prods in grouped_products.items()
        }

        # Trending section: mix of highest discount + recently added in-stock products
        all_in_stock = [p for cat_prods in grouped_products.values()
                        for p in cat_prods
                        if p.get('stock_total', 0) and p['stock_total'] > 0]

        # Score: 60% weight on discount %, 40% on recency (id)
        max_id = max((p['id'] for p in all_in_stock), default=1)
        def trend_score(p):
            mrp = float(p.get('mrp_price') or 0)
            rp  = float(p.get('retail_price') or p.get('price1') or 0)
            disc_pct = ((mrp - rp) / mrp * 100) if mrp and mrp > rp else 0
            recency = (p['id'] / max_id) * 100
            return disc_pct * 0.6 + recency * 0.4

        trending = sorted(all_in_stock, key=trend_score, reverse=True)[:8]
        import random
        random.shuffle(trending)  # shuffle so it feels fresh each load

        public_coupons = get_public_coupons(db)

        # Offers carousel also gets 3 random picks from the trending shelf --
        # a random sample, not a fixed "top 3", so the carousel varies between visits.
        discounted_trending = [
            p for p in trending
            if p.get('mrp_price') and p['mrp_price'] > (p.get('retail_price') or p.get('price1') or 0)
        ]
        carousel_products = random.sample(discounted_trending, min(3, len(discounted_trending)))
        for p in carousel_products:
            mrp = float(p['mrp_price'])
            rp = float(p.get('retail_price') or p.get('price1') or 0)
            p['disc_pct'] = round((mrp - rp) / mrp * 100)
            p['current_price'] = rp
            p['thumb'] = p['images'][0] if p.get('images') else '/static/assets/products/default.jpg'

        return render_site('index.html', grouped_products=grouped_products,
                           trending=trending, hero_images=hero_images,
                           public_coupons=public_coupons, carousel_products=carousel_products)

    products = db.execute('''
        SELECT p.*, c.name as category_name FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.is_active=1
    ''').fetchall()
    grouped_products = {}
    for p in products:
        cat = p['category_name'] or p['category'] or 'New Arrivals'
        if cat not in grouped_products:
            grouped_products[cat] = []
        p_dict = dict(p)
        p_dict['images'] = get_product_images(p_dict)
        p_dict['tiers'] = get_product_tiers(p_dict)
        grouped_products[cat].append(p_dict)
    return render_site('index.html', grouped_products=grouped_products, hero_images=hero_images)

@app.route('/category/<category>')
@app.route('/retail/category/<category>')
@app.route('/wholesale/category/<category>')
def category_products(category):
    if request.path.startswith('/retail'):
        g.site_type = 'retail'
    elif request.path.startswith('/wholesale'):
        g.site_type = 'wholesale'
        
    db = get_db()
    # For retail, filter by the 'category' column
    if request.path.startswith('/retail'):
        raw_products = db.execute('SELECT * FROM products WHERE category = ? AND is_active = 1', (category,)).fetchall()
    else:
        raw_products = db.execute('''
            SELECT p.* FROM products p
            JOIN categories c ON p.category_id = c.id
            WHERE c.name = ? AND p.is_active = 1
        ''', (category,)).fetchall()
    products = []
    for p in raw_products:
        p_dict = dict(p)
        p_dict['images'] = get_product_images(p_dict)
        p_dict['tiers'] = get_product_tiers(p_dict)
        if is_bangle_product(p_dict):
            p_dict['size_stock'] = get_bangle_size_stock(db, p_dict['sku'])
        products.append(p_dict)

    sort = request.args.get('sort') or ''
    size_filter = (request.args.get('size') or '').strip()
    in_stock_only = request.args.get('in_stock') == '1'
    products, available_sizes = apply_sort_and_filters(
        products, sort=sort, size_filter=size_filter, in_stock_only=in_stock_only
    )

    public_coupons, carousel_products = get_offers_carousel_data(db) if g.site_type == 'retail' else ([], [])

    return render_site('category_products.html', category=category, products=products,
                        sort=sort, size_filter=size_filter, in_stock_only=in_stock_only,
                        available_sizes=available_sizes,
                        public_coupons=public_coupons, carousel_products=carousel_products)

@app.route('/product/<int:product_id>')
@app.route('/retail/product/<int:product_id>', endpoint='product_detail_retail')
@app.route('/wholesale/product/<int:product_id>', endpoint='product_detail_wholesale')
def product_detail(product_id):
    if request.path.startswith('/retail'):
        g.site_type = 'retail'
    elif request.path.startswith('/wholesale'):
        g.site_type = 'wholesale'
        
    db = get_db()
    product = db.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
    if not product: return "Not Found", 404
    p_dict = dict(product)
    log_product_event(db, p_dict.get('sku'), 'view', visitor_id=getattr(g, 'visitor_id', None), source=getattr(g, 'traffic_source', None))
    image_urls = get_product_images(p_dict)
    p_dict['tiers'] = get_product_tiers(p_dict)
    if is_bangle_product(p_dict):
        p_dict['size_stock'] = get_bangle_size_stock(db, p_dict['sku'])

    related = db.execute(
        'SELECT * FROM products WHERE id != ? AND is_active = 1 ORDER BY RANDOM() LIMIT 4',
        (product_id,)
    ).fetchall()
    related_products = []
    for r in related:
        r_dict = dict(r)
        r_dict['images'] = get_product_images(r_dict)
        if is_bangle_product(r_dict):
            r_dict['size_stock'] = get_bangle_size_stock(db, r_dict['sku'])
        related_products.append(r_dict)

    public_coupons, carousel_products = get_offers_carousel_data(db) if g.site_type == 'retail' else ([], [])

    return render_site('product_detail.html', product=p_dict, image_urls=image_urls, related_products=related_products,
                        public_coupons=public_coupons, carousel_products=carousel_products)

@app.route('/favicon.ico')
def favicon():
    return app.send_static_file('assets/favicon.ico')

@app.route('/robots.txt')
def robots():
    return app.response_class(
        "User-agent: *\nAllow: /\nDisallow: /admin/\nDisallow: /checkout/\nSitemap: https://narinakhre.com/sitemap.xml\n",
        mimetype='text/plain'
    )

@app.route('/sitemap.xml')
def sitemap():
    db = get_db()
    products = db.execute("SELECT id, slug, name FROM products WHERE is_active=1").fetchall()
    categories = db.execute("SELECT DISTINCT category FROM products WHERE is_active=1").fetchall()
    base = "https://narinakhre.com"
    lines = ['<?xml version="1.0" encoding="UTF-8"?>',
             '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for path in ['/', '/retail', '/wholesale/contact', '/retail/contact']:
        lines.append(f'  <url><loc>{base}{path}</loc><changefreq>weekly</changefreq><priority>1.0</priority></url>')
    for cat in categories:
        lines.append(f'  <url><loc>{base}/retail/category/{cat["category"]}</loc><changefreq>daily</changefreq><priority>0.8</priority></url>')
        lines.append(f'  <url><loc>{base}/category/{cat["category"]}</loc><changefreq>daily</changefreq><priority>0.8</priority></url>')
    for p in products:
        slug = p["slug"] or str(p["id"])
        lines.append(f'  <url><loc>{base}/retail/product/{p["id"]}</loc><changefreq>weekly</changefreq><priority>0.7</priority></url>')
        lines.append(f'  <url><loc>{base}/wholesale/product/{p["id"]}</loc><changefreq>weekly</changefreq><priority>0.6</priority></url>')
    lines.append('</urlset>')
    return app.response_class('\n'.join(lines), mimetype='application/xml')

# --- CART & CHECKOUT ---
@app.route('/cart', methods=['GET'])
def get_cart():
    """Current session cart as JSON, keyed the same way session['cart'] is
    stored (f"{sku}_{size}"). Used by the client to know, without guessing,
    how much of a given product+size is already in the cart -- e.g. to
    restore the Add to Cart/counter UI on page load, or to tell whether
    switching a product's size selector should show that size's existing
    quantity or reset to a fresh Add to Cart button."""
    return jsonify(session.get('cart', {}))


@app.route('/update-cart', methods=['POST'])
def update_cart():
    if g.site_type != 'retail':
        return jsonify({'status': 'error', 'message': 'Cart not available on wholesale'}), 403
    data = request.get_json()
    sku = data.get('product_id')
    qty = int(data.get('qty', 1))
    price = float(data.get('price', 0))
    size = data.get('size') or ''
    
    cart = session.get('cart', {})
    cart_key = f"{sku}_{size}"
    if qty > 0:
        db = get_db()
        if cart_key not in cart:
            # This is a POST/AJAX call, not a page navigation, so
            # track_page_view() never ran for it -- read the visitor
            # cookie directly rather than relying on g.visitor_id.
            log_product_event(
                db, sku, 'add_to_cart',
                visitor_id=request.cookies.get(VISITOR_COOKIE_NAME),
                source=classify_traffic_source(request.referrer)
            )
        p = db.execute('SELECT name, category, sub_category FROM products WHERE sku = ?', (sku,)).fetchone()
        if p and is_bangle_product(dict(p)) and size:
            size_stock = get_bangle_size_stock(db, sku)
            if size_stock and size_stock.get(size, 0) <= 0:
                return jsonify({'status': 'error', 'message': f'Size {size} is out of stock for this product.'}), 409
        cart[cart_key] = {
            'sku': sku,
            'name': p['name'] if p else sku,
            'qty': qty,
            'price': price,
            'size': size
        }
    else:
        cart.pop(cart_key, None)
    session['cart'] = cart
    session.modified = True
    return jsonify({'status': 'success', 'new_total': sum(item['qty'] for item in cart.values())})

@app.route('/checkout')
@app.route('/retail/checkout')
@app.route('/wholesale/checkout')
def checkout():
    if request.path.startswith('/retail'):
        g.site_type = 'retail'
    elif request.path.startswith('/wholesale'):
        g.site_type = 'wholesale'
    
    cart = session.get('cart', {})
    display_cart = []
    out_of_stock_items = []
    db = get_db()
    for item in cart.values():
        item_dict = dict(item)
        if 'units' not in item_dict:
            item_dict['units'] = item_dict.get('qty', 1)
        # Check live stock from DB
        sku = item_dict.get('sku', '')
        live = db.execute('SELECT stock_total, name, image_field, category, sub_category FROM products WHERE sku=?', (sku,)).fetchone()
        if live:
            item_size = item_dict.get('size') or ''
            if item_size and is_bangle_product(dict(live)):
                size_stock = get_bangle_size_stock(db, sku)
                effective_stock = size_stock.get(item_size, 0) if size_stock else (live['stock_total'] or 0)
            else:
                effective_stock = live['stock_total'] or 0
            item_dict['stock_total'] = effective_stock
            item_dict['is_out_of_stock'] = effective_stock == 0
            if item_dict['is_out_of_stock']:
                out_of_stock_items.append(item_dict.get('name') or live['name'] or sku)
            # Get image
            if not item_dict.get('image_url'):
                try:
                    imgs = get_product_images(dict(live))
                    if imgs and imgs[0].startswith('http'):
                        item_dict['image_url'] = imgs[0]
                except Exception:
                    pass
        else:
            item_dict['is_out_of_stock'] = False
        display_cart.append(item_dict)
    
    subtotal = sum(item['price'] * item['units'] for item in display_cart)
    applied_coupon = session.get('applied_coupon')
    discount = applied_coupon['discount_amount'] if applied_coupon else 0.0
    coupon_code = applied_coupon['code'] if applied_coupon else ''

    # Signed-in retail customers get their last shipping address prefilled
    # on checkout instead of retyping it — same source data as profile().
    saved_address = None
    saved_addresses = []
    credit_balance = 0.0
    applied_credits_row = session.get('applied_credits')
    credits_applied = float(applied_credits_row.get('amount', 0)) if applied_credits_row else 0.0
    if g.site_type == 'retail' and session.get('user_id'):
        saved_addresses = [dict(a) for a in db.execute(
            'SELECT * FROM user_addresses WHERE user_id=? ORDER BY is_default DESC, created_at DESC',
            (session['user_id'],),
        ).fetchall()]
        if not saved_addresses:
            # No address book entries yet -- fall back to prefilling from the
            # customer's last order, same as before the address book existed.
            saved_address = db.execute(
                'SELECT * FROM order_shipping WHERE user_id=? ORDER BY created_at DESC LIMIT 1',
                (session['user_id'],),
            ).fetchone()
        credit_balance = get_credit_balance(db, session['user_id'])
        if credits_applied > credit_balance:
            # Stale session value (balance changed elsewhere) — drop it rather
            # than let the customer redeem more than they actually have.
            credits_applied = 0.0
            session.pop('applied_credits', None)
            session.modified = True

    grand_total = max(subtotal - discount - credits_applied, 0)

    public_coupons = get_public_coupons(db) if g.site_type == 'retail' else []

    # "Spend a bit more to unlock this coupon" nudge -- only relevant when no
    # coupon is applied yet, and only for coupons the customer hasn't reached
    # the minimum order amount for. Picks whichever public coupon needs the
    # smallest top-up, so the nudge is always achievable.
    next_coupon_gap = None
    if not coupon_code and public_coupons:
        candidates = [
            c for c in public_coupons
            if c.get('min_order_amount') and float(c['min_order_amount']) > subtotal
        ]
        if candidates:
            closest = min(candidates, key=lambda c: float(c['min_order_amount']) - subtotal)
            next_coupon_gap = {
                'code': closest['code'],
                'amount_needed': float(closest['min_order_amount']) - subtotal,
                'discount_percent': closest['discount_percent'],
            }

    return render_site('checkout.html', display_cart=display_cart, subtotal=subtotal, total_tax=0.0,
                        discount=discount, grand_total=grand_total, coupon_code=coupon_code,
                        out_of_stock_items=out_of_stock_items, recaptcha_site_key=RECAPTCHA_SITE_KEY,
                        saved_address=saved_address, saved_addresses=saved_addresses,
                        credit_balance=credit_balance, credits_applied=credits_applied,
                        public_coupons=public_coupons, next_coupon_gap=next_coupon_gap)

@app.route('/checkout/shipping', methods=['GET', 'POST'])
@app.route('/retail/checkout/shipping', methods=['GET', 'POST'])
def checkout_shipping():
    """Render shipping address form for checkout."""
    g.site_type = 'retail'
    if request.method == 'POST':
        return redirect(url_for('checkout_process'))
    return render_site('checkout_shipping.html')


@app.route('/checkout/process', methods=['POST'])
@app.route('/retail/checkout/process', methods=['POST'])
def checkout_process():
    g.site_type = 'retail'
    if not session.get('user_id'):
        return jsonify({'status': 'error', 'message': 'Please sign in to place your order.', 'require_login': True}), 401
    consignee_name = (request.form.get('consignee_name') or '').strip()
    consignee_phone = (request.form.get('consignee_phone') or '').strip()
    consignee_address = (request.form.get('consignee_address') or '').strip()
    consignee_city = (request.form.get('consignee_city') or '').strip()
    consignee_state = (request.form.get('consignee_state') or '').strip()
    consignee_pincode = (request.form.get('consignee_pincode') or '').strip()
    consignee_email = (request.form.get('email') or '').strip()
    payment_mode = (request.form.get('payment_mode') or 'Prepaid').strip()

    def sanitize_for_delhivery(value):
        cleaned = value or ''
        for char in ['#', '&', '%', ';']:
            cleaned = cleaned.replace(char, ' ')
        return ' '.join(cleaned.split())

    cleaned_name = sanitize_for_delhivery(consignee_name)
    cleaned_address = sanitize_for_delhivery(consignee_address)

    internal_order_id = f"NN-SHP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{consignee_phone[-4:]}"
    user_id = session.get('user_id')

    # Calculate financials from cart
    cart = session.get('cart', {})
    if not cart:
        return jsonify({'status': 'error', 'message': 'Cart is empty'}), 400

    display_cart = list(cart.values())
    subtotal_amount = sum(float(item.get('price', 0)) * int(item.get('units', item.get('qty', 1))) for item in display_cart)
    applied_coupon = session.get('applied_coupon')
    discount_amount = float(applied_coupon.get('discount_amount', 0)) if applied_coupon else 0.0
    coupon_code = applied_coupon.get('code') if applied_coupon else None

    # Nari Nakhre Credits redemption -- reserved here (not just validated)
    # against the real ledger via an atomic DB-side check-and-hold, so two
    # tabs/devices redeeming at the same instant can't both succeed against
    # the same balance (see reserve_credits_atomic). The hold is only
    # finalized into a real debit once the order is confirmed (see
    # confirm_cod_order / verify_payment) -- and released automatically if
    # this order is never confirmed -- so an abandoned checkout doesn't
    # permanently spend credits either.
    conn_for_credits = get_db()
    applied_credits_row = session.get('applied_credits')
    credits_requested = float(applied_credits_row.get('amount', 0)) if applied_credits_row else 0.0
    credits_redeemed = 0.0
    if credits_requested > 0 and user_id:
        credits_requested = round(min(credits_requested, max(subtotal_amount - discount_amount, 0)), 2)
        if credits_requested > 0:
            reserved = reserve_credits(conn_for_credits, user_id, credits_requested, internal_order_id)
            conn_for_credits.commit()
            if reserved:
                credits_redeemed = credits_requested
            else:
                app.logger.warning(
                    f"Could not reserve {credits_requested} credits for user {user_id} on {internal_order_id}: "
                    f"insufficient balance (likely already claimed by another session)."
                )

    gst_breakdown = calculate_inclusive_gst(display_cart, discount_amount, subtotal_amount)
    gst_amount = gst_breakdown['total_gst']
    cgst_amount = gst_breakdown['cgst']
    sgst_amount = gst_breakdown['sgst']

    # The cheapest quote always ships free -- actual_shipping_cost is what we
    # pay the courier for it, absorbed by the business, never shown to the
    # customer as a charge. If the customer picked a pricier courier on the
    # checkout page (see ckSelectedCourier in checkout.html), they pay the
    # delta above the cheapest quote -- shipping_upgrade_charge, folded into
    # total_amount below. courier_partner/courier_eta ARE shown to the
    # customer (see thank_you page + confirmation emails) -- rate-shopped
    # here once, at order-creation time, and reused for actual shipment
    # creation later (see create_courier_shipment) so what's shown matches
    # what ships.
    actual_shipping_cost = 0.0
    shipping_upgrade_charge = 0.0
    chosen_courier = 'delhivery'
    courier_eta = None
    try:
        cart_weight = max(sum(int(item.get('units', 1)) for item in display_cart) * 250, 250)
        quotes = get_all_courier_quotes(
            app.config.get('WAREHOUSE_PIN', '482001'), consignee_pincode, cart_weight, mode=payment_mode
        )
        # Customer may have picked a non-default courier on the checkout
        # page -- honor it if it's one of the couriers that actually quoted
        # successfully for this order; otherwise (not sent, or no longer
        # valid) fall back to cheapest, same as before this override existed.
        requested_courier = (request.form.get('courier_partner') or '').strip().lower()
        selected = next((q for q in quotes if q[0] == requested_courier), None) if requested_courier else None
        chosen_courier, _provider, rates = selected or quotes[0]
        actual_shipping_cost = float(rates.get('shipping_charge', 0) or 0)
        courier_eta = rates.get('eta')
        base_charge = _customer_shipping_charge(quotes[0][2])
        shipping_upgrade_charge = round(max(_customer_shipping_charge(rates) - base_charge, 0), 2)
    except Exception as e:
        app.logger.warning(f'Shipping rate fetch failed: {e}')

    total_amount = max(subtotal_amount - discount_amount - credits_redeemed + shipping_upgrade_charge, 0)

    # Store cart items as JSON for admin order view
    cart_items_json = json.dumps([{
        'sku': item.get('sku', ''),
        'name': item.get('name', ''),
        'price': float(item.get('price', 0)),
        'units': int(item.get('units', item.get('qty', 1))),
        'size': item.get('size', ''),
    } for item in display_cart])

    conn = get_db()
    conn.execute(
        '''INSERT INTO order_shipping
           (user_id, consignee_name, consignee_phone, consignee_email,
            consignee_address, consignee_city, consignee_state, consignee_pincode,
            internal_order_id, status, payment_mode,
            subtotal_amount, gst_amount, cgst_amount, sgst_amount,
            discount_amount, actual_shipping_cost, total_amount,
            coupon_code, cart_items_json, credits_redeemed,
            courier_partner, courier_eta, shipping_upgrade_charge)
           VALUES (?,?,?,?,?,?,?,?,?,'pending',?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (user_id, cleaned_name, consignee_phone, consignee_email,
         cleaned_address, consignee_city, consignee_state, consignee_pincode,
         internal_order_id, payment_mode,
         subtotal_amount, gst_amount, cgst_amount, sgst_amount,
         discount_amount, actual_shipping_cost, total_amount,
         coupon_code, cart_items_json, credits_redeemed,
         chosen_courier, courier_eta, shipping_upgrade_charge)
    )
    conn.commit()
    session.pop('applied_credits', None)

    log_admin_event(
        conn, 'new_order', f'New order: {internal_order_id}',
        detail=f'{cleaned_name} — ₹{total_amount:.0f} ({payment_mode})',
        related_id=internal_order_id
    )

    # Delhivery shipment created AFTER payment — not here
    waybill = None

    if waybill:
        conn.execute(
            'UPDATE order_shipping SET delhivery_waybill=? WHERE internal_order_id=?',
            (waybill, internal_order_id)
        )
        conn.commit()

    print('Sanitized shipping payload for Delhivery:', {
        'internal_order_id': internal_order_id,
        'consignee_name': cleaned_name,
        'consignee_phone': consignee_phone,
        'consignee_address': cleaned_address,
        'consignee_city': consignee_city,
        'consignee_state': consignee_state,
        'consignee_pincode': consignee_pincode,
        'waybill': waybill,
    })

    session['checkout_handover'] = {
        'internal_order_id': internal_order_id,
        'waybill': None,
    }
    session.modified = True

    # Verify the order was actually saved before returning success
    verify = conn.execute(
        'SELECT id FROM order_shipping WHERE internal_order_id=?',
        (internal_order_id,)
    ).fetchone()
    if not verify:
        app.logger.error(f'checkout_process: ORDER INSERT FAILED for {internal_order_id}')
        return jsonify({
            'status': 'error',
            'message': 'Order could not be saved. Please try again.'
        }), 500

    app.logger.info(f'checkout_process: Order {internal_order_id} saved OK')
    return jsonify({
        'status': 'success',
        'internal_order_id': internal_order_id,
    }), 200

@app.route('/payment/gateway', methods=['GET'])
@app.route('/retail/payment/gateway', methods=['GET'])
def payment_gateway_router():
    """Payment authorization gateway with session validation."""
    g.site_type = 'retail'
    checkout_handover = session.get('checkout_handover', {})
    internal_order_id = checkout_handover.get('internal_order_id')
    waybill = checkout_handover.get('waybill')
    
    if not internal_order_id:
        flash('Order ID missing. Please complete shipping details again.', 'error')
        return redirect(url_for('checkout_shipping'))
    # waybill may be None if Delhivery API was unavailable — allow checkout to proceed
    
    # Calculate amount from current cart
    cart = session.get('cart', {})
    subtotal = sum(item['price'] * item['qty'] for item in cart.values())
    # Use the persisted order total (GST-inclusive, discount applied) from the DB
    # NEVER add 18% GST here — prices are already GST-inclusive at 3%
    handover = session.get('checkout_handover', {})
    order_id_for_amount = handover.get('internal_order_id', '')
    amount_to_pay = subtotal  # fallback
    total_tax = 0.0
    if order_id_for_amount:
        conn = get_db()
        orow = conn.execute(
            'SELECT total_amount, gst_amount FROM order_shipping WHERE internal_order_id=?',
            (order_id_for_amount,)
        ).fetchone()
        if orow and orow['total_amount']:
            amount_to_pay = float(orow['total_amount'])
            total_tax = float(orow['gst_amount'] or 0)
    
    return render_site('payment_gateway.html',
        internal_order_id=internal_order_id,
        waybill=waybill,
        amount_to_pay=amount_to_pay,
        subtotal=subtotal,
        total_tax=total_tax
    )

@app.route('/payment/cancel', methods=['POST'])
@app.route('/retail/payment/cancel', methods=['POST'])
def payment_cancel():
    """Clear checkout state and return to catalogue."""
    g.site_type = 'retail'
    session.pop('checkout_handover', None)
    session.pop('cart', None)
    session.modified = True
    flash('Order cancelled. Returning to store.', 'info')
    return redirect(url_for('index'))


@app.route('/api/create-order', methods=['POST'])
def create_razorpay_order():
    if g.site_type != 'retail':
        return jsonify({'status': 'error', 'message': 'Payment not available on wholesale'}), 403
    g.site_type = 'retail'
    payload = request.get_json(silent=True) or request.form or {}

    checkout_handover = session.get('checkout_handover', {})
    internal_order_id = (payload.get('order_id') or checkout_handover.get('internal_order_id') or '').strip()
    waybill = (payload.get('waybill') or checkout_handover.get('waybill') or '').strip()

    if not RAZORPAY_KEY_ID or not RAZORPAY_KEY_SECRET:
        return jsonify({
            'status': 'error',
            'message': 'Razorpay credentials are not configured'
        }), 500

    try:
        # The DB's total_amount is the source of truth for what to charge --
        # it already reflects any coupon discount and Nari Nakhre Credits
        # redeemed at checkout_process time. The client-supplied `amount` is
        # never trusted for this (it can't account for credits, and taking
        # it at face value would let a tampered request set its own price);
        # it's only a last-resort fallback if the order can't be found.
        requested_amount = None
        if internal_order_id:
            order_row_for_amount = get_db().execute(
                'SELECT total_amount FROM order_shipping WHERE internal_order_id=?',
                (internal_order_id,)
            ).fetchone()
            if order_row_for_amount and order_row_for_amount['total_amount'] is not None:
                requested_amount = float(order_row_for_amount['total_amount'])

        if requested_amount is None:
            requested_amount = payload.get('amount')

        amount_paise = int(round(float(requested_amount) * 100))
        if amount_paise <= 0:
            return jsonify({
                'status': 'error',
                'message': 'Invalid amount for payment'
            }), 400
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'Invalid amount for payment'
        }), 400

    receipt = internal_order_id or f"NN-RZP-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    try:
        order_payload = {
            'amount': amount_paise,
            'currency': 'INR',
            'receipt': receipt,
            'payment_capture': 1,
            'notes': {
                'internal_order_id': internal_order_id,
                'waybill': waybill,
            },
        }
        razorpay_order = razorpay_client.order.create(data=order_payload)

        session['payment_pending'] = {
            'internal_order_id': internal_order_id,
            'waybill': waybill,
            'razorpay_order_id': razorpay_order.get('id'),
            'amount_paise': amount_paise,
        }
        session['razorpay_order_id'] = razorpay_order.get('id')
        session['internal_order_id'] = internal_order_id
        session['waybill'] = waybill
        session.modified = True

        return jsonify({
            'status': 'success',
            'order_id': razorpay_order.get('id'),
            'razorpay_order_id': razorpay_order.get('id'),
            'amount': razorpay_order.get('amount', amount_paise),
            'currency': razorpay_order.get('currency', 'INR'),
            'receipt': razorpay_order.get('receipt', receipt),
            'key_id': app.config.get('RAZORPAY_KEY_ID', ''),
        }), 200
    except Exception:
        return jsonify({
            'status': 'error',
            'message': 'Unable to create Razorpay order'
        }), 500


@app.route('/api/confirm-cod', methods=['POST'])
def confirm_cod_order():
    """Confirm a COD order immediately after address submission."""
    g.site_type = 'retail'
    checkout_handover = session.get('checkout_handover', {})
    internal_order_id = (checkout_handover.get('internal_order_id') or '').strip()
    app.logger.info(f'confirm_cod: handover={checkout_handover}, order_id={internal_order_id}')

    if not internal_order_id:
        app.logger.error('confirm_cod: No internal_order_id in session')
        return jsonify({'status': 'error', 'message': 'Session expired. Please go back and try again.'}), 400

    # reCAPTCHA v3 — COD orders skip Razorpay's natural fraud check (no real payment
    # happens), so this is the one anti-bot/anti-fraud gate before a shipment gets
    # created and Delhivery pickup gets scheduled for a fake order.
    req_body = request.get_json(silent=True) or {}
    if not verify_recaptcha(req_body.get('recaptcha_token'), remote_ip=request.remote_addr, expected_action='confirm_cod'):
        app.logger.warning(f'confirm_cod: reCAPTCHA rejected for order {internal_order_id}')
        return jsonify({'status': 'error', 'message': 'Verification failed. Please refresh the page and try again.'}), 400

    conn = get_db()

    # First check if order exists at all (any status)
    any_row = conn.execute(
        'SELECT id, status FROM order_shipping WHERE internal_order_id=?',
        (internal_order_id,)
    ).fetchone()
    app.logger.info(f'confirm_cod: DB lookup result={dict(any_row) if any_row else None}')

    if not any_row:
        app.logger.error(f'confirm_cod: Order {internal_order_id} not found in DB at all')
        return jsonify({
            'status': 'error',
            'message': 'Order not saved — please go back and place your order again.'
        }), 400

    if any_row['status'] != 'pending':
        app.logger.warning(f'confirm_cod: Order {internal_order_id} has status={any_row["status"]}')
        if any_row['status'] in ('cod_confirmed', 'paid'):
            # Already confirmed — just redirect to thank you
            return jsonify({'status': 'success', 'waybill': None, 'internal_order_id': internal_order_id}), 200
        return jsonify({'status': 'error', 'message': f'Order already processed (status: {any_row["status"]})'}), 400

    order_row = conn.execute(
        'SELECT * FROM order_shipping WHERE internal_order_id=? AND status=?',
        (internal_order_id, 'pending')
    ).fetchone()
    if not order_row:
        return jsonify({'status': 'error', 'message': 'Order not found or already processed'}), 400
    order_row_dict = dict(order_row)
    cart = session.get('cart', {})
    cart_items = list(cart.values()) if cart else []

    # Delhivery COD agents collect whole-rupee cash, so the collection amount
    # is rounded to the nearest ₹10. If that rounds UP, the customer earns a
    # Nari Nakhre Credits award instead of getting physical change: the
    # actual rounding excess (cod_collect_amount - original_total) PLUS a
    # flat COD_ROUNDING_CREDIT bonus on top. Guest checkouts (no account to
    # credit) never round up -- only down -- so a guest is never charged
    # more with nothing given back for it.
    original_total = float(order_row_dict.get('total_amount', 0) or 0)
    cod_user_id = order_row_dict.get('user_id')
    if cod_user_id:
        cod_collect_amount, cod_rounded_up = round_cod_amount(original_total)
    else:
        cod_collect_amount = float(int(original_total // 10) * 10)
        cod_rounded_up = False
    cod_credit_awarded = round((cod_collect_amount - original_total) + COD_ROUNDING_CREDIT, 2) if cod_rounded_up else 0

    waybill, del_error, shipment_partner, shipment_id = create_courier_shipment(order_row_dict, cart_items, cod_amount_override=cod_collect_amount)
    if waybill:
        conn.execute(
            'UPDATE order_shipping SET status=?, delhivery_waybill=?, courier_partner=?, cod_collected_amount=?, cod_credit_awarded=?, shiprocket_shipment_id=? WHERE internal_order_id=?',
            ('cod_confirmed', waybill, shipment_partner, cod_collect_amount, cod_credit_awarded, shipment_id, internal_order_id))
    else:
        app.logger.error(f"{shipment_partner.capitalize()} failed for COD {internal_order_id}: {del_error}")
        conn.execute(
            'UPDATE order_shipping SET status=?, cod_collected_amount=?, cod_credit_awarded=? WHERE internal_order_id=?',
            ('cod_confirmed', cod_collect_amount, cod_credit_awarded, internal_order_id))
    if cod_credit_awarded:
        try:
            award_credits(conn, cod_user_id, cod_credit_awarded, 'cod_rounding', internal_order_id)
        except Exception as e:
            app.logger.warning(f"Failed to award COD rounding credits for {internal_order_id}: {e}")
    # Credits redeemed toward this order were already reserved (held) at
    # checkout_process time; finalize that hold into a real debit now that
    # the order is actually confirmed.
    credits_redeemed = float(order_row_dict.get('credits_redeemed', 0) or 0)
    if credits_redeemed and cod_user_id:
        try:
            finalize_credit_redemption(conn, cod_user_id, credits_redeemed, internal_order_id)
        except Exception as e:
            app.logger.warning(f"Failed to finalize redeemed credits for {internal_order_id}: {e}")
    conn.commit()
    try:
        customer_email = order_row_dict.get('consignee_email', '')
        customer_name = order_row_dict.get('consignee_name', 'Customer')
        total = cod_collect_amount
        tracking_url = f"{request.url_root.rstrip('/')}/track/{waybill}" if waybill else ''
        invoice_url = f"{request.url_root.rstrip('/')}/invoice/{internal_order_id}"
        items_for_email = [{
            'name': item.get('name', ''),
            'size': item.get('size', ''),
            'units': int(item.get('units', item.get('qty', 1))),
            'price': float(item.get('price', 0)),
            'row_total': float(item.get('price', 0)) * int(item.get('units', item.get('qty', 1))),
        } for item in cart_items]

        courier_eta = order_row_dict.get('courier_eta')
        if customer_email:
            order_html = render_template('retail/email_order_confirmation.html',
                customer_name=customer_name, order_id=internal_order_id,
                items=items_for_email, payment_mode='COD', amount=total,
                address_name=customer_name,
                address_line=order_row_dict.get('consignee_address', ''),
                address_city=order_row_dict.get('consignee_city', ''),
                address_state=order_row_dict.get('consignee_state', ''),
                address_pincode=order_row_dict.get('consignee_pincode', ''),
                tracking_url=tracking_url or None, waybill=waybill or None,
                invoice_url=invoice_url, cod_credit_awarded=cod_credit_awarded,
                courier_partner=shipment_partner, courier_eta=courier_eta)
            order_text = (
                f"Hi {customer_name},\n\nYour COD order is confirmed!\n\nOrder ID: {internal_order_id}\n"
                f"Amount to pay on delivery: ₹{total:.2f}\n"
                f"Shipped via {shipment_partner.capitalize()}"
                + (f", estimated delivery: {courier_eta}\n" if courier_eta else "\n")
                + (
                    f"\nNote: our delivery partner requires COD amounts to be a whole number, so this "
                    f"was rounded up from the order total. Instead of collecting exact change, we've "
                    f"credited {cod_credit_awarded} Nari Nakhre Credits to your account for use on a "
                    f"future order.\n" if cod_credit_awarded else ""
                )
                + (f"Track: {tracking_url}\n" if tracking_url else "Tracking will be shared once your order is dispatched.\n")
                + f"Invoice: {invoice_url}\n\nThank you for shopping with Nari Nakhre!"
            )
            send_contact_email_async(customer_email,
                f"Order Confirmed (COD) — {internal_order_id} | Nari Nakhre",
                order_text, html_body=order_html, from_email=ORDERS_FROM_EMAIL)

        item_lines = '\n'.join(
            f"  - {it['name']}" + (f" ({it['size']})" if it['size'] else '')
            + f" x{it['units']} @ ₹{it['price']:.2f} = ₹{it['row_total']:.2f}"
            for it in items_for_email
        ) or '  (no item details)'
        subtotal = float(order_row_dict.get('subtotal_amount', 0) or 0)
        discount = float(order_row_dict.get('discount_amount', 0) or 0)
        gst = float(order_row_dict.get('gst_amount', 0) or 0)
        shipping_cost = float(order_row_dict.get('actual_shipping_cost', 0) or 0)
        coupon_code = order_row_dict.get('coupon_code') or ''
        admin_orders_url = f"{request.url_root.rstrip('/')}/admin/orders"

        admin_body = (
            f"New COD order placed.\n\n"
            f"Order ID: {internal_order_id}\n"
            f"Customer: {customer_name}\n"
            f"Phone: {order_row_dict.get('consignee_phone','')}\n"
            f"Email: {customer_email or '-'}\n\n"
            f"Items:\n{item_lines}\n\n"
            f"Subtotal: ₹{subtotal:.2f}\n"
            + (f"Coupon ({coupon_code}) discount: -₹{discount:.2f}\n" if discount else "")
            + (f"GST (incl.): ₹{gst:.2f}\n" if gst else "")
            + (f"Shipping: ₹{shipping_cost:.2f}\n" if shipping_cost else "Shipping: Free\n")
            + f"Total to collect (COD): ₹{total:.2f}\n"
            + (f"(rounded up from ₹{original_total:.2f}; {cod_credit_awarded} Nari Nakhre Credits awarded to customer)\n" if cod_credit_awarded else "")
            + "\n"
            f"Shipping Address:\n{order_row_dict.get('consignee_address','')}, "
            f"{order_row_dict.get('consignee_city','')}, {order_row_dict.get('consignee_state','')} - {order_row_dict.get('consignee_pincode','')}\n\n"
            f"Courier: {shipment_partner.capitalize()}"
            + (f" (estimated delivery: {courier_eta})\n" if courier_eta else "\n")
            + (f"Waybill: {waybill}\n" if waybill else "Waybill: pending\n")
            + f"Admin orders panel: {admin_orders_url}\n"
        )
        send_contact_email_async(ADMIN_EMAIL,
            f"🛍️ New COD Order — {internal_order_id}",
            admin_body,
            from_email=ORDERS_FROM_EMAIL)
    except Exception as e:
        app.logger.warning(f"COD email failed: {e}")
    session['checkout_handover'] = {
        'internal_order_id': internal_order_id,
        'waybill': waybill,
        'amount_paid': cod_collect_amount,
        'cod_credit_awarded': cod_credit_awarded,
        'payment_mode': 'COD',
    }
    session.pop('cart', None)
    session.pop('applied_coupon', None)
    session.modified = True
    return jsonify({'status': 'success', 'waybill': waybill, 'internal_order_id': internal_order_id}), 200


@app.route('/payment-failed')
def payment_failed():
    """Page shown when Razorpay payment fails or is cancelled."""
    g.site_type = 'retail'
    order_id = request.args.get('order_id', '')
    reason = request.args.get('reason', 'Payment was not completed')
    return render_template('retail/payment_failed.html',
                           order_id=order_id, reason=reason)


@app.route('/api/verify-payment', methods=['POST'])
def verify_payment():
    g.site_type = 'retail'
    payload = request.get_json(silent=True) or request.form or {}
    razorpay_order_id = (payload.get('razorpay_order_id') or '').strip()
    razorpay_payment_id = (payload.get('razorpay_payment_id') or '').strip()
    razorpay_signature = (payload.get('razorpay_signature') or '').strip()

    if not razorpay_order_id or not razorpay_payment_id or not razorpay_signature:
        return jsonify({
            'status': 'error',
            'message': 'Missing payment verification fields'
        }), 400

    pending_razorpay_order_id = (session.get('razorpay_order_id') or '').strip()
    if not pending_razorpay_order_id or pending_razorpay_order_id != razorpay_order_id:
        return jsonify({
            'status': 'error',
            'message': 'Order ID mismatch for pending session transaction'
        }), 400

    params_dict = {
        'razorpay_order_id': razorpay_order_id,
        'razorpay_payment_id': razorpay_payment_id,
        'razorpay_signature': razorpay_signature
    }

    try:
        razorpay_client.utility.verify_payment_signature(params_dict)
    except Exception:
        return jsonify({
            'status': 'error',
            'message': 'Invalid payment signature'
        }), 400

    try:
        internal_order_id = (session.get('internal_order_id') or '').strip()
        if not internal_order_id:
            checkout_handover = session.get('checkout_handover', {})
            internal_order_id = (checkout_handover.get('internal_order_id') or '').strip()

        if not internal_order_id:
            return jsonify({
                'status': 'error',
                'message': 'No active order found for this payment'
            }), 400

        conn = get_db()
        row = conn.execute(
            'SELECT * FROM order_shipping WHERE internal_order_id=?',
            (internal_order_id,)
        ).fetchone()
        if row is None:
            return jsonify({
                'status': 'error',
                'message': 'No active order found for this payment'
            }), 400
        order_row_dict = dict(row)

        conn.execute(
            'UPDATE order_shipping SET status=? WHERE internal_order_id=?',
            ('paid', internal_order_id)
        )
        # Credits redeemed toward this order were already reserved (held) at
        # checkout_process time; finalize that hold into a real debit now
        # that payment is actually confirmed.
        credits_redeemed = float(order_row_dict.get('credits_redeemed', 0) or 0)
        if credits_redeemed and order_row_dict.get('user_id'):
            try:
                finalize_credit_redemption(conn, order_row_dict['user_id'], credits_redeemed, internal_order_id)
            except Exception as e:
                app.logger.warning(f"Failed to finalize redeemed credits for {internal_order_id}: {e}")
        conn.commit()

        # Order confirmation + admin notification emails (best-effort, never
        # blocks payment verification). Mirrors the COD confirmation emails —
        # prepaid orders previously sent none at all.
        try:
            customer_email = order_row_dict.get('consignee_email', '')
            customer_name = order_row_dict.get('consignee_name', 'Customer')
            total = float(order_row_dict.get('total_amount', 0) or 0)
            invoice_url = f"{request.url_root.rstrip('/')}/invoice/{internal_order_id}"
            cart_items = []
            if order_row_dict.get('cart_items_json'):
                try:
                    cart_items = json.loads(order_row_dict['cart_items_json'])
                except (ValueError, TypeError):
                    cart_items = []
            items_for_email = [{
                'name': item.get('name', ''),
                'size': item.get('size', ''),
                'units': int(item.get('units', item.get('qty', 1))),
                'price': float(item.get('price', 0)),
                'row_total': float(item.get('price', 0)) * int(item.get('units', item.get('qty', 1))),
            } for item in cart_items]

            courier_partner = order_row_dict.get('courier_partner')
            courier_eta = order_row_dict.get('courier_eta')
            if customer_email:
                order_html = render_template('retail/email_order_confirmation.html',
                    customer_name=customer_name, order_id=internal_order_id,
                    items=items_for_email, payment_mode='Prepaid', amount=total,
                    address_name=customer_name,
                    address_line=order_row_dict.get('consignee_address', ''),
                    address_city=order_row_dict.get('consignee_city', ''),
                    address_state=order_row_dict.get('consignee_state', ''),
                    address_pincode=order_row_dict.get('consignee_pincode', ''),
                    tracking_url=None, waybill=None,
                    invoice_url=invoice_url,
                    courier_partner=courier_partner, courier_eta=courier_eta)
                order_text = (
                    f"Hi {customer_name},\n\nYour payment for order {internal_order_id} was successful!\n\n"
                    f"Amount paid: ₹{total:.2f}\n"
                    + (f"Shipped via {courier_partner.capitalize()}" + (f", estimated delivery: {courier_eta}\n" if courier_eta else "\n") if courier_partner else "")
                    + f"Tracking details will be shared once your order is dispatched.\n"
                    f"Invoice: {invoice_url}\n\nThank you for shopping with Nari Nakhre!"
                )
                send_contact_email_async(customer_email,
                    f"Order Confirmed — {internal_order_id} | Nari Nakhre",
                    order_text, html_body=order_html, from_email=ORDERS_FROM_EMAIL)

            item_lines = '\n'.join(
                f"  - {it['name']}" + (f" ({it['size']})" if it['size'] else '')
                + f" x{it['units']} @ ₹{it['price']:.2f} = ₹{it['row_total']:.2f}"
                for it in items_for_email
            ) or '  (no item details)'
            subtotal = float(order_row_dict.get('subtotal_amount', 0) or 0)
            discount = float(order_row_dict.get('discount_amount', 0) or 0)
            gst = float(order_row_dict.get('gst_amount', 0) or 0)
            shipping_cost = float(order_row_dict.get('actual_shipping_cost', 0) or 0)
            coupon_code = order_row_dict.get('coupon_code') or ''
            admin_orders_url = f"{request.url_root.rstrip('/')}/admin/orders"
            admin_body = (
                f"New prepaid order — payment received via Razorpay.\n\n"
                f"Order ID: {internal_order_id}\n"
                f"Payment ID: {razorpay_payment_id}\n"
                f"Customer: {customer_name}\n"
                f"Phone: {order_row_dict.get('consignee_phone','')}\n"
                f"Email: {customer_email or '-'}\n\n"
                f"Items:\n{item_lines}\n\n"
                f"Subtotal: ₹{subtotal:.2f}\n"
                + (f"Coupon ({coupon_code}) discount: -₹{discount:.2f}\n" if discount else "")
                + (f"GST (incl.): ₹{gst:.2f}\n" if gst else "")
                + (f"Shipping: ₹{shipping_cost:.2f}\n" if shipping_cost else "Shipping: Free\n")
                + f"Total paid: ₹{total:.2f}\n\n"
                f"Shipping Address:\n{order_row_dict.get('consignee_address','')}, "
                f"{order_row_dict.get('consignee_city','')}, {order_row_dict.get('consignee_state','')} - {order_row_dict.get('consignee_pincode','')}\n\n"
                + (f"Courier: {courier_partner.capitalize()}" + (f" (estimated delivery: {courier_eta})\n" if courier_eta else "\n") if courier_partner else "")
                + f"Admin orders panel: {admin_orders_url}\n"
            )
            send_contact_email_async(ADMIN_EMAIL,
                f"💳 New Prepaid Order — {internal_order_id}",
                admin_body,
                from_email=ORDERS_FROM_EMAIL)
        except Exception as e:
            app.logger.warning(f"Prepaid order email failed: {e}")

        session.pop('razorpay_order_id', None)
        session.pop('payment_pending', None)
        session.pop('internal_order_id', None)
        session.pop('waybill', None)
        session.pop('checkout_handover', None)
        session.modified = True

        # If called via AJAX (fetch), return JSON; if form POST, redirect
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return jsonify({
                'status': 'success',
                'message': 'Payment verified and order finalized',
                'internal_order_id': internal_order_id,
            }), 200
        return redirect(url_for('thank_you', ref=internal_order_id))
    except Exception as e:
        app.logger.error(f'Payment verification error: {e}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return jsonify({'status': 'error', 'message': 'Unable to finalize verified payment'}), 500
        flash('Payment verification failed. Please contact support.', 'error')
        return redirect(url_for('checkout'))

# --- DELHIVERY API ROUTES (Retail Only) ---
@app.route('/api/delhivery/check/<pincode>', methods=['GET'])
def delhivery_check_pincode(pincode):
    if g.site_type != 'retail':
        return jsonify({"status": False, "msg": "Unauthorized"}), 403
    # Validate pincode format before hitting Delhivery API
    import re as _re
    if not _re.match(r'^\d{6}$', str(pincode)):
        return jsonify({"status": False, "serviceable": False, "msg": "Invalid pincode format"}), 400
    try:
        partner, provider = get_active_courier()
        pickup_pin = app.config.get('WAREHOUSE_PIN', '') if partner == 'shiprocket' else None
        result = provider.verify_pincode(pincode, pickup_pincode=pickup_pin)
        # Surface the real reason in logs for debugging (visible in Render logs)
        if not result.get('serviceable') and not result.get('status'):
            app.logger.error(f"{partner.capitalize()} pincode {pincode} check failed: {result.get('msg')}")
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'Courier pincode check exception: {type(e).__name__}: {e}')
        return jsonify({"status": False, "serviceable": False, "msg": f"Service unavailable: {type(e).__name__}"}), 200

@app.route('/api/delhivery/shipping', methods=['POST'])
def calculate_checkout_shipping():
    if g.site_type != 'retail':
        return jsonify({"status": False, "msg": "Unauthorized"}), 403
    import re as _re2
    data = request.get_json(silent=True) or {}
    pincode = str(data.get('pincode') or data.get('destination') or '').strip()
    payment_mode = data.get('mode', 'Prepaid')
    if not _re2.match(r'^[0-9]{6}$', pincode):
        return jsonify({"status": False, "shipping_charge": 0,
                        "cod_fee": 0, "msg": "Invalid pincode format"}), 400
    try:
        cart = session.get('cart', {})
        total_weight = max(sum(item.get('qty', 1) for item in cart.values()) * 250, 250)
        quotes = get_all_courier_quotes(
            app.config.get('WAREHOUSE_PIN', ''), pincode, total_weight, mode=payment_mode
        )
        # Cheapest first (get_all_courier_quotes already sorts this way) --
        # that's the default (free) selection; the customer can switch to a
        # pricier option, in which case they pay the delta above the
        # cheapest quote (extra_charge) -- see checkout_process, which
        # applies the same math server-side when the order is actually
        # placed. cod_fee/shipping_charge stay internal (never shown to the
        # customer as a price), only extra_charge (already the customer-
        # facing delta), courier name + eta are customer-facing.
        base_charge = _customer_shipping_charge(quotes[0][2])
        options = [{
            "courier_partner": name,
            "eta": rates.get('eta'),
            "shipping_charge": rates.get('rate', 0) or rates.get('shipping_charge', 0),
            "cod_fee": rates.get('cod_fee', 0) if payment_mode == 'COD' else 0,
            "extra_charge": round(max(_customer_shipping_charge(rates) - base_charge, 0), 2),
        } for name, _provider, rates in quotes]
        default = options[0]
        return jsonify({
            "status": True,
            "shipping_charge": default["shipping_charge"],
            "cod_fee": default["cod_fee"],
            "payment_mode": payment_mode,
            "courier_partner": default["courier_partner"],
            "eta": default["eta"],
            "options": options
        })
    except Exception as e:
        app.logger.error(f'Delhivery shipping calc error: {e}')
        return jsonify({"status": False, "shipping_charge": 0,
                        "cod_fee": 0, "msg": "Shipping rate unavailable"}), 200


@app.route('/retail/place_order', methods=['POST'])
def place_order():
    data = request.form if request.form else request.json
    name = data.get('name')
    phone = data.get('phone')
    email = data.get('email')
    address_line1 = data.get('address_line1')
    address_line2 = data.get('address_line2')
    city = data.get('city')
    state = data.get('state')
    pincode = data.get('pincode')
    country = data.get('country', 'IN')
    payment_mode = data.get('payment_mode')
    amount = float(data.get('amount', 0))
    order_id = f'NN{datetime.now().strftime("%Y%m%d%H%M%S")}{phone[-4:]}'
    cart = session.get('cart', {})
    db = get_db()
    total_weight = 0
    for item in cart.values():
        sku = item['sku']
        qty = item['qty']
        prod = db.execute('SELECT weight, length, breadth, height FROM products WHERE sku = ?', (sku,)).fetchone()
        if prod:
            dead_weight = (prod['weight'] or 0) * qty
            l = prod['length'] or 0
            b = prod['breadth'] or 0
            h = prod['height'] or 0
            vol_weight = get_shipping_provider(app.config['SHIPPING_PROVIDER']).calculate_volumetric_weight(l, b, h) * qty
            billable = max(dead_weight, vol_weight)
            total_weight += billable
        else:
            total_weight += qty * 250  # fallback
    provider = get_shipping_provider(
        app.config['SHIPPING_PROVIDER'],
        api_token=app.config.get('DELHIVERY_API_KEY')
    )
    shipment_data = {
        "name": name,
        "add": f"{address_line1}, {address_line2}",
        "pin": pincode,
        "phone": phone,
        "order": order_id,
        "payment_mode": payment_mode,
        "total_amount": amount,
        "weight": total_weight,
        "city": city,
        "state": state,
        "country": country,
        "email": email,
        "mobile": phone
    }
    resp = provider.create_shipment(shipment_data)
    waybill = resp.get('waybill')
    conn = get_db()
    conn.execute(
        "INSERT INTO order_shipping (internal_order_id, consignee_name, consignee_phone, consignee_address, consignee_city, consignee_state, consignee_pincode, delhivery_waybill, status) VALUES (?,?,?,?,?,?,?,?,'pending')",
        (order_id, name, phone, f"{address_line1}, {address_line2}", city, state, pincode, waybill)
    )
    conn.commit()
    log_admin_event(conn, 'new_order', f'New order: {order_id}',
                     detail=f'{name} — ₹{amount:.0f} ({payment_mode})', related_id=order_id)

    # Finalize coupon usage if one was applied to this order
    applied_coupon = session.get('applied_coupon')
    if applied_coupon and applied_coupon.get('code'):
        try:
            conn.execute(
                'UPDATE coupons SET times_used = times_used + 1 WHERE code=?',
                (applied_coupon['code'],)
            )
            conn.commit()
        except Exception as e:
            app.logger.warning(f'Could not increment coupon usage: {e}')

    session.pop('cart', None)
    session.pop('applied_coupon', None)

    # Use our own branded, shareable tracking page instead of the raw Delhivery URL
    tracking_url = url_for('track_order_page', waybill=waybill, _external=True) if waybill else None

    # Email the order confirmation + tracking link to the customer (best-effort, never blocks checkout)
    if email and waybill:
        try:
            track_body = (
                f"Hi {name},\n\n"
                f"Your Nari Nakhre order ({order_id}) has been placed successfully!\n\n"
                f"Track your order here: {tracking_url}\n\n"
                f"Thank you for shopping with us.\n- Team Nari Nakhre"
            )
            send_contact_email(email, "Your Nari Nakhre order is confirmed!", track_body)
        except Exception as e:
            app.logger.warning(f"Order confirmation email failed: {e}")

    return render_site('thank_you.html', order_id=order_id, waybill=waybill, tracking_url=tracking_url)

@app.route('/api/track/<waybill>', methods=['GET'])
def api_track_shipment(waybill):
    """Live tracking status — retail only."""
    if g.site_type != 'retail':
        return jsonify({'status': False, 'msg': 'Tracking not available'}), 403
    db = get_db()
    order = db.execute(
        'SELECT internal_order_id, courier_partner FROM order_shipping WHERE delhivery_waybill=? LIMIT 1', (waybill,)
    ).fetchone()
    partner_name = (order.get('courier_partner') if order else None) or 'delhivery'
    _partner, provider = get_courier(partner_name)
    try:
        result = provider.track_shipment(waybill)
        result['courier_partner'] = partner_name
        result['internal_order_id'] = order.get('internal_order_id') if order else None
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'Tracking error: {e}')
        return jsonify({"status": False, "msg": "Could not fetch tracking info", "courier_partner": partner_name}), 200


@app.route('/track/<waybill>')
def track_order_page(waybill):
    """Public, shareable order-tracking page — retail only.
    Wholesale is a quote-based service with no order tracking."""
    if g.site_type != 'retail':
        return redirect('/')
    conn = get_db()
    order = conn.execute(
        'SELECT * FROM order_shipping WHERE delhivery_waybill=?', (waybill,)
    ).fetchone()
    return render_template('retail/track_order.html', waybill=waybill, order=order)




@app.route('/api/search')
def api_search():
    q = (request.args.get('q') or '').strip()
    site = request.args.get('t') or getattr(g, 'site_type', 'retail')

    if len(q) < 2:
        return jsonify({'products': [], 'orders': [], 'query': q})

    conn = get_db()
    results = {'products': [], 'orders': [], 'query': q}

    try:
        matches = search_products(conn, q, limit=8)
        app.logger.info(f"Search '{q}' → {len(matches)} products")

        for p in matches:
            img = ''
            if p.get('image_field'):
                parts = p['image_field'].split(',')
                img = parts[0].strip() if parts else ''
            mrp = float(p.get('mrp_price') or 0)
            rp  = float(p.get('retail_price') or 0)
            disc = int((mrp - rp) / mrp * 100) if mrp and mrp > rp else 0
            results['products'].append({
                'id':           p['id'],
                'sku':          p.get('sku') or '',
                'model_number': p.get('model_number') or '',
                'name':         p.get('name') or '',
                'category':     p.get('category') or '',
                'price':        rp,
                'mrp':          mrp,
                'discount':     disc,
                'image':        img,
                'url': f"/retail/product/{p['id']}" if site == 'retail'
                       else f"/wholesale/product/{p['id']}",
            })
    except Exception as e:
        app.logger.error(f'Search error: {type(e).__name__}: {e}')

    if site == 'retail' and len(q) >= 6:
        try:
            like = f'%{q.lower()}%'
            rows = conn.execute(
                "SELECT internal_order_id, consignee_name, status,"
                " total_amount, delhivery_waybill"
                " FROM order_shipping"
                " WHERE LOWER(internal_order_id) LIKE ?"
                " OR LOWER(delhivery_waybill) LIKE ?"
                " LIMIT 2",
                (like, like)
            ).fetchall()
            for o in rows:
                o_dict = dict(o)
                results['orders'].append({
                    'order_id': o_dict['internal_order_id'],
                    'status':   (o_dict['status'] or 'pending').replace('_',' ').title(),
                    'waybill':  o_dict['delhivery_waybill'] or '',
                    'name':     o_dict['consignee_name'] or '',
                    'total':    float(o_dict['total_amount'] or 0),
                    'url': f"/track/{o_dict['delhivery_waybill']}"
                           if o_dict['delhivery_waybill'] else '',
                })
        except Exception as e:
            app.logger.error(f'Order search error: {e}')

    return jsonify(results)


@app.route('/search')
def search_page():
    """Full search results page for longer queries or when JS is disabled."""
    q = (request.args.get('q') or '').strip()
    site = getattr(g, 'site_type', 'retail')
    if not q:
        return redirect('/' + site)

    sort = request.args.get('sort') or ''
    size_filter = (request.args.get('size') or '').strip()
    in_stock_only = request.args.get('in_stock') == '1'

    conn = get_db()
    products = []
    try:
        matches = search_products(conn, q)
        for p in matches:
            p_dict = dict(p)
            imgs = (p_dict.get('image_field') or '').split(',')
            p_dict['image'] = imgs[0].strip() if imgs else ''
            mrp = float(p_dict.get('mrp_price') or 0)
            rp  = float(p_dict.get('retail_price') or 0)
            p_dict['discount'] = int((mrp - rp) / mrp * 100) if mrp and mrp > rp else 0
            if is_bangle_product(p_dict):
                p_dict['size_stock'] = get_bangle_size_stock(conn, p_dict['sku'])
            products.append(p_dict)
    except Exception as e:
        app.logger.error(f'Search page error: {e}')
        products = []

    products, available_sizes = apply_sort_and_filters(
        products, sort=sort, size_filter=size_filter, in_stock_only=in_stock_only
    )
    products = products[:40]

    public_coupons, carousel_products = get_offers_carousel_data(conn) if site == 'retail' else ([], [])

    return render_site('search_results.html', products=products, query=q,
                        sort=sort, size_filter=size_filter, in_stock_only=in_stock_only,
                        available_sizes=available_sizes,
                        public_coupons=public_coupons, carousel_products=carousel_products)

@app.route('/clear_oos_items', methods=['POST'])
def clear_oos_items():
    """Remove out-of-stock items from cart, then redirect back to retail checkout."""
    g.site_type = 'retail'
    cart = session.get('cart', {})
    db = get_db()
    to_remove = []
    for key, item in list(cart.items()):
        sku = item.get('sku', '')
        if not sku:
            continue
        row = db.execute(
            'SELECT stock_total, category, sub_category, name FROM products WHERE sku=?', (sku,)
        ).fetchone()
        if not row:
            continue
        item_size = item.get('size') or ''
        if item_size and is_bangle_product(dict(row)):
            size_stock = get_bangle_size_stock(db, sku)
            effective_stock = size_stock.get(item_size, 0) if size_stock else (row['stock_total'] or 0)
        else:
            effective_stock = row['stock_total'] or 0
        if effective_stock == 0:
            to_remove.append(key)
    for key in to_remove:
        cart.pop(key, None)
    session['cart'] = cart
    session.modified = True
    # Always redirect to retail checkout explicitly
    return redirect('/retail/checkout')


@app.route('/apply_coupon', methods=['POST'])
def apply_coupon():
    data = request.get_json(silent=True) or {}
    code = (data.get('coupon') or '').strip().upper()
    if not code:
        return jsonify({"status": "error", "message": "Please enter a coupon code"}), 400

    db = get_db()
    coupon = db.execute(
        "SELECT * FROM coupons WHERE code=? AND is_active=1", (code,)
    ).fetchone()

    if not coupon:
        return jsonify({"status": "error", "message": "Invalid or inactive coupon code"}), 200

    # Expiry check
    if coupon['expiry_date']:
        try:
            from datetime import date as _date
            expiry = coupon['expiry_date']
            if isinstance(expiry, str):
                expiry = datetime.strptime(expiry, '%Y-%m-%d').date()
            if expiry < _date.today():
                return jsonify({"status": "error", "message": "This coupon has expired"}), 200
        except Exception:
            pass

    # Usage limit check
    if coupon['usage_limit'] and coupon['usage_limit'] > 0:
        if coupon['times_used'] >= coupon['usage_limit']:
            return jsonify({"status": "error", "message": "This coupon has reached its usage limit"}), 200

    # Calculate cart subtotal, optionally filtered by category/sub_category
    cart = session.get('cart', {})
    if not cart:
        return jsonify({"status": "error", "message": "Your cart is empty"}), 200

    eligible_subtotal = 0.0
    full_subtotal = 0.0
    for item in cart.values():
        units = item.get('units', item.get('qty', 1))
        price = item.get('price', 0)
        line_total = price * units
        full_subtotal += line_total

        applies = True
        if coupon['category'] or coupon['sub_category']:
            prod = db.execute(
                'SELECT category, sub_category FROM products WHERE sku=?', (item.get('sku'),)
            ).fetchone()
            if prod:
                if coupon['category'] and prod['category'] != coupon['category']:
                    applies = False
                if coupon['sub_category'] and prod['sub_category'] != coupon['sub_category']:
                    applies = False
            else:
                applies = False
        if applies:
            eligible_subtotal += line_total

    # Minimum order amount check (checked against full cart subtotal)
    if coupon['min_order_amount'] and full_subtotal < coupon['min_order_amount']:
        return jsonify({
            "status": "error",
            "message": f"Minimum order amount of \u20b9{coupon['min_order_amount']:.0f} required for this coupon"
        }), 200

    if eligible_subtotal <= 0:
        return jsonify({
            "status": "error",
            "message": "This coupon doesn't apply to any items in your cart"
        }), 200

    raw_discount = round(eligible_subtotal * (coupon['discount_percent'] / 100.0), 2)
    max_disc = float(coupon.get('max_discount_amount') or 0)
    discount = round(min(raw_discount, max_disc) if max_disc > 0 else raw_discount, 2)

    session['applied_coupon'] = {
        "code": code,
        "discount_percent": coupon['discount_percent'],
        "discount_amount": discount,
        "min_order_amount": float(coupon.get('min_order_amount') or 0),
        "max_discount_amount": max_disc,
        "category": coupon['category'],
        "sub_category": coupon['sub_category']
    }
    session.modified = True

    return jsonify({
        "status": "success",
        "message": f"Coupon applied! You saved \u20b9{discount:.0f}",
        "discount": discount,
        "discount_amount": discount,
        "discount_percent": coupon['discount_percent'],
        "min_order_amount": float(coupon.get('min_order_amount') or 0),
        "max_discount_amount": max_disc,
        "code": code
    })


@app.route('/remove_coupon', methods=['POST'])
def remove_coupon():
    session.pop('applied_coupon', None)
    session.modified = True
    return jsonify({"status": "success"})


@app.route('/apply_credits', methods=['POST'])
def apply_credits():
    if not session.get('user_id'):
        return jsonify({"status": "error", "message": "Please sign in to use Nari Nakhre Credits"}), 200

    data = request.get_json(silent=True) or {}
    try:
        amount = float(data.get('amount', 0))
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({"status": "error", "message": "Enter a valid credit amount"}), 200

    db = get_db()
    balance = get_credit_balance(db, session['user_id'])
    if amount > balance:
        return jsonify({"status": "error", "message": f"You only have {balance:.0f} Nari Nakhre Credits available"}), 200

    cart = session.get('cart', {})
    subtotal = sum(item.get('price', 0) * item.get('units', item.get('qty', 1)) for item in cart.values())
    applied_coupon = session.get('applied_coupon')
    coupon_discount = float(applied_coupon.get('discount_amount', 0)) if applied_coupon else 0.0
    max_redeemable = max(subtotal - coupon_discount, 0)
    amount = round(min(amount, max_redeemable), 2)
    if amount <= 0:
        return jsonify({"status": "error", "message": "Your order total is already ₹0"}), 200

    session['applied_credits'] = {"amount": amount}
    session.modified = True
    return jsonify({
        "status": "success",
        "message": f"₹{amount:.0f} Nari Nakhre Credits applied",
        "amount": amount,
        "balance": balance,
    })


@app.route('/remove_credits', methods=['POST'])
def remove_credits():
    session.pop('applied_credits', None)
    session.modified = True
    return jsonify({"status": "success"})


@app.route('/clear_quote', methods=['POST'])
def clear_quote():
    session.pop('cart', None)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return ('', 204)
    return redirect('/retail' if g.site_type == 'retail' else '/wholesale')

@app.route('/thank_you')
def thank_you():
    # Pull order_id from query (?ref=) first, then session (set during checkout_process)
    checkout_handover = session.get('checkout_handover', {})
    order_id = (request.args.get('ref')
                or session.get('internal_order_id', '')
                or checkout_handover.get('internal_order_id', ''))

    waybill = ''
    payment_mode = ''
    amount_paid = 0.0
    cart_items = []
    cod_credit_awarded = 0

    # DB is the source of truth — session data may already be popped (e.g. after
    # a Razorpay payment verification) by the time this page loads
    if order_id:
        conn = get_db()
        row = conn.execute(
            'SELECT delhivery_waybill, payment_mode, total_amount, cart_items_json, '
            'cod_collected_amount, cod_credit_awarded '
            'FROM order_shipping WHERE internal_order_id=? ORDER BY id DESC LIMIT 1',
            (order_id,)
        ).fetchone()
        if row:
            waybill = row['delhivery_waybill'] or ''
            payment_mode = row['payment_mode'] or ''
            amount_paid = float(row['cod_collected_amount'] or row['total_amount'] or 0)
            cod_credit_awarded = float(row['cod_credit_awarded'] or 0)
            if row['cart_items_json']:
                try:
                    cart_items = json.loads(row['cart_items_json'])
                except (ValueError, TypeError):
                    cart_items = []

    # Fill any gaps from the freshly-set session handover (covers the moment right
    # after order placement, before/without a DB round trip)
    if not waybill:
        waybill = checkout_handover.get('waybill') or session.get('waybill', '') or ''
    if not payment_mode:
        payment_mode = checkout_handover.get('payment_mode', '')
    if not amount_paid:
        amount_paid = float(checkout_handover.get('amount_paid', 0) or 0)
    if not cod_credit_awarded:
        cod_credit_awarded = float(checkout_handover.get('cod_credit_awarded', 0) or 0)

    tracking_url = url_for('track_order_page', waybill=waybill, _external=True) if waybill else None

    return render_template('retail/thank_you.html',
                           order_id=order_id, waybill=waybill,
                           tracking_url=tracking_url,
                           cod_credit_awarded=cod_credit_awarded,
                           amount_paid=amount_paid,
                           payment_mode=payment_mode,
                           cart_items=cart_items)


def admin_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if session.get('is_admin') is not True:
            return redirect(url_for('admin_login'))
        return view_func(*args, **kwargs)
    return wrapped_view


@app.route('/admin/upload-images', methods=['POST'])
@admin_required
def admin_upload_images():
    sku = request.form.get('sku')
    uploaded_files = request.files.getlist('images')

    if not sku or not uploaded_files:
        flash('SKU and images are required for upload.')
        return redirect(url_for('admin_manage_images'))

    sku = sku.strip()
    if not sku:
        flash('SKU and images are required for upload.')
        return redirect(url_for('admin_manage_images'))

    conn = get_db()
    product = conn.execute('SELECT id FROM products WHERE sku=?', (sku,)).fetchone()
    if not product:
        flash('Product not found for the provided SKU.')
        return redirect(url_for('admin_manage_images'))

    first_public_cloud_url = None
    for idx, file in enumerate(uploaded_files, start=1):
        if not file or not file.filename:
            continue

        target_name = f"{sku}_{idx}.webp"
        public_cloud_url = upload_image_to_supabase(file, target_name)
        if public_cloud_url and first_public_cloud_url is None:
            first_public_cloud_url = public_cloud_url

    if first_public_cloud_url:
        conn.execute('UPDATE products SET image_field=? WHERE sku=?', (first_public_cloud_url, sku))
        conn.commit()
        flash('Product images successfully processed, scaled down, converted to WebP format, and synced to Supabase!')
    else:
        flash('No images were uploaded to cloud storage. Please try again.')

    return redirect(url_for('admin_manage_images'))


@app.route('/admin/dashboard', methods=['GET'])
@admin_required
def admin_dashboard():
    db = get_db()
    products = db.execute('SELECT * FROM products ORDER BY id DESC').fetchall()
    quotes = db.execute('SELECT * FROM quotes ORDER BY id DESC').fetchall()
    unread_row = db.execute('SELECT COUNT(*) as cnt FROM admin_events WHERE is_read=0').fetchone()
    unread_events = unread_row['cnt'] if unread_row else 0
    return render_template('admin/admin.html', products=products, quotes=quotes, unread_events=unread_events)


EVENT_TYPE_LABELS = {
    'new_user': ('👤', 'New User', '#2563eb'),
    'new_order': ('🛒', 'New Order', '#16a34a'),
    'order_delivered': ('📦', 'Delivered', '#9333ea'),
    'order_not_accepted_2h': ('⏰', 'Not Accepted 2h+', '#f59e0b'),
    'order_not_accepted_6h': ('⏰', 'Not Accepted 6h+', '#f97316'),
    'order_not_accepted_12h': ('⏰', 'Not Accepted 12h+', '#ea580c'),
    'order_not_accepted_24h': ('🚨', 'Not Accepted 24h+', '#dc2626'),
}


@app.route('/admin/events', methods=['GET'])
@admin_required
def admin_events():
    db = get_db()
    type_filter = (request.args.get('type') or '').strip()
    if type_filter:
        rows = db.execute(
            'SELECT * FROM admin_events WHERE event_type=? ORDER BY id DESC LIMIT 200', (type_filter,)
        ).fetchall()
    else:
        rows = db.execute('SELECT * FROM admin_events ORDER BY id DESC LIMIT 200').fetchall()

    events = []
    for r in rows:
        icon, label, color = EVENT_TYPE_LABELS.get(r['event_type'], ('🔔', r['event_type'], '#6b7280'))
        e = dict(r)
        e['icon'], e['label'], e['color'] = icon, label, color
        events.append(e)

    unread_row = db.execute('SELECT COUNT(*) as cnt FROM admin_events WHERE is_read=0').fetchone()
    unread_count = unread_row['cnt'] if unread_row else 0

    return render_template('admin/admin_events.html', events=events, unread_count=unread_count,
                            type_filter=type_filter, event_type_labels=EVENT_TYPE_LABELS)


@app.route('/admin/events/mark-read', methods=['POST'])
@admin_required
def admin_events_mark_read():
    db = get_db()
    db.execute('UPDATE admin_events SET is_read=1 WHERE is_read=0')
    db.commit()
    flash('All notifications marked as read.')
    return redirect(url_for('admin_events'))


def resolve_insight_range(req):
    """Shared date-range resolution for the analytics pages (Overview,
    Visitors, Product Visits) -- defaults to the last 7 days so every page
    has something sensible to show before an admin picks custom dates, and
    a custom range picked on one page carries over cleanly if reused via
    the same range/from/to query params."""
    now = datetime.now()
    range_mode = req.args.get('range', 'preset')
    from_date = req.args.get('from', '').strip()
    to_date = req.args.get('to', '').strip()
    if range_mode == 'custom' and from_date and to_date:
        insight_start = from_date + ' 00:00:00'
        insight_end = to_date + ' 23:59:59'
    else:
        range_mode = 'preset'
        insight_start = (now - timedelta(days=7)).strftime('%Y-%m-%d 00:00:00')
        insight_end = now.strftime('%Y-%m-%d 23:59:59')
        from_date = (now - timedelta(days=7)).strftime('%Y-%m-%d')
        to_date = now.strftime('%Y-%m-%d')
    return range_mode, from_date, to_date, insight_start, insight_end


@app.route('/admin/analytics', methods=['GET'])
@admin_required
def admin_analytics():
    db = get_db()
    now = datetime.now()

    def window_stats(days):
        start = (now - timedelta(days=days)).strftime('%Y-%m-%d 00:00:00')
        row = db.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(total_amount),0) as revenue "
            "FROM order_shipping WHERE created_at >= ? AND status != 'cancelled'",
            (start,)
        ).fetchone()
        return {'label': f'Last {days} Day' + ('' if days == 1 else 's'),
                'count': row['cnt'] if row else 0,
                'revenue': float(row['revenue'] or 0) if row else 0.0}

    stats_1d = window_stats(1)
    stats_7d = window_stats(7)
    stats_30d = window_stats(30)

    range_mode, from_date, to_date, insight_start, insight_end = resolve_insight_range(request)

    custom_stats = None
    if range_mode == 'custom':
        row = db.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(total_amount),0) as revenue "
            "FROM order_shipping WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled'",
            (insight_start, insight_end)
        ).fetchone()
        custom_stats = {'label': f'{from_date} to {to_date}',
                         'count': row['cnt'] if row else 0,
                         'revenue': float(row['revenue'] or 0) if row else 0.0}

    # ── Customer insights: most-ordered, most-viewed, most-added-to-cart,
    # all scoped to the same window as the custom/preset range picker above.
    orders_in_range = db.execute(
        "SELECT cart_items_json FROM order_shipping "
        "WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled'",
        (insight_start, insight_end)
    ).fetchall()
    product_orders = {}
    for o in orders_in_range:
        try:
            items = json.loads(o['cart_items_json'] or '[]')
        except Exception:
            items = []
        seen_this_order = set()
        for item in items:
            sku = item.get('sku') or ''
            if not sku:
                continue
            entry = product_orders.setdefault(sku, {'sku': sku, 'name': item.get('name', sku), 'units': 0, 'orders': 0})
            entry['units'] += int(item.get('units', 1) or 1)
            if sku not in seen_this_order:
                entry['orders'] += 1
                seen_this_order.add(sku)
    top_ordered = sorted(product_orders.values(), key=lambda e: e['units'], reverse=True)[:10]

    def top_events(event_type):
        rows = db.execute(
            "SELECT sku, COUNT(*) as cnt FROM product_events "
            "WHERE event_type=? AND created_at >= ? AND created_at <= ? "
            "GROUP BY sku ORDER BY cnt DESC LIMIT 10",
            (event_type, insight_start, insight_end)
        ).fetchall()
        out = []
        for r in rows:
            p = db.execute('SELECT name FROM products WHERE sku=?', (r['sku'],)).fetchone()
            out.append({'sku': r['sku'], 'name': p['name'] if p else r['sku'], 'count': r['cnt']})
        return out

    top_viewed = top_events('view')
    top_cart_adds = top_events('add_to_cart')

    # ── Sales insights: orders by state, colour-graded by volume (a data
    # heat-map -- ranked/shaded by intensity -- rather than a geographic
    # SVG map, which would need external map assets this app doesn't have).
    state_rows = db.execute(
        "SELECT COALESCE(NULLIF(consignee_state,''), 'Unknown') as state, "
        "COUNT(*) as cnt, COALESCE(SUM(total_amount),0) as revenue "
        "FROM order_shipping WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled' "
        "GROUP BY state ORDER BY cnt DESC",
        (insight_start, insight_end)
    ).fetchall()
    state_data = [{'state': r['state'], 'count': r['cnt'], 'revenue': float(r['revenue'] or 0)} for r in state_rows]
    max_state_count = max((s['count'] for s in state_data), default=0) or 1

    return render_template('admin/admin_analytics.html',
        stats_1d=stats_1d, stats_7d=stats_7d, stats_30d=stats_30d, custom_stats=custom_stats,
        range_mode=range_mode, from_date=from_date, to_date=to_date,
        top_ordered=top_ordered, top_viewed=top_viewed, top_cart_adds=top_cart_adds,
        state_data=state_data, max_state_count=max_state_count)


@app.route('/admin/analytics/product-visits', methods=['GET'])
@admin_required
def admin_product_visits():
    """Full "View All" drill-down for product page views: every product
    that had at least one view in range, with unique-visitor counts and a
    traffic-source breakdown per product (not just the top-10 shown on the
    main Analytics overview)."""
    db = get_db()
    range_mode, from_date, to_date, insight_start, insight_end = resolve_insight_range(request)

    view_rows = db.execute(
        "SELECT sku, COUNT(*) as views, COUNT(DISTINCT visitor_id) as unique_views "
        "FROM product_events WHERE event_type='view' AND created_at >= ? AND created_at <= ? "
        "GROUP BY sku ORDER BY views DESC",
        (insight_start, insight_end)
    ).fetchall()

    source_rows = db.execute(
        "SELECT sku, source, COUNT(*) as cnt FROM product_events "
        "WHERE event_type='view' AND created_at >= ? AND created_at <= ? AND source IS NOT NULL "
        "GROUP BY sku, source",
        (insight_start, insight_end)
    ).fetchall()
    sources_by_sku = {}
    for r in source_rows:
        sources_by_sku.setdefault(r['sku'], []).append({'source': r['source'], 'count': r['cnt']})

    products = []
    max_views = view_rows[0]['views'] if view_rows else 0
    for r in view_rows:
        p = db.execute('SELECT id, name FROM products WHERE sku=?', (r['sku'],)).fetchone()
        srcs = sorted(sources_by_sku.get(r['sku'], []), key=lambda s: -s['count'])
        products.append({
            'sku': r['sku'],
            'id': p['id'] if p else None,
            'name': p['name'] if p else r['sku'],
            'views': r['views'],
            'unique_views': r['unique_views'],
            'sources': srcs,
        })

    return render_template('admin/admin_product_visits.html',
                            products=products, max_views=max_views or 1,
                            range_mode=range_mode, from_date=from_date, to_date=to_date)


@app.route('/admin/analytics/visitors', methods=['GET'])
@admin_required
def admin_visitors():
    """Site-wide visitor analytics: how many people visited, when traffic
    peaked (which day, which hour of day), and where they came from
    (search engine, Instagram, WhatsApp, direct, etc.) -- built on the
    page_views table logged by the track_page_view before_request hook."""
    db = get_db()
    range_mode, from_date, to_date, insight_start, insight_end = resolve_insight_range(request)

    totals_row = db.execute(
        "SELECT COUNT(*) as views, COUNT(DISTINCT visitor_id) as visitors "
        "FROM page_views WHERE created_at >= ? AND created_at <= ?",
        (insight_start, insight_end)
    ).fetchone()
    totals = {'views': totals_row['views'] if totals_row else 0,
              'visitors': totals_row['visitors'] if totals_row else 0}

    daily_rows = db.execute(
        "SELECT DATE(created_at) as day, COUNT(*) as views, COUNT(DISTINCT visitor_id) as visitors "
        "FROM page_views WHERE created_at >= ? AND created_at <= ? "
        "GROUP BY DATE(created_at) ORDER BY day",
        (insight_start, insight_end)
    ).fetchall()
    daily_data = [{'day': str(r['day']), 'views': r['views'], 'visitors': r['visitors']} for r in daily_rows]
    peak_day = max(daily_data, key=lambda d: d['visitors'], default=None)
    max_daily_visitors = max((d['visitors'] for d in daily_data), default=0) or 1

    hourly_rows = db.execute(
        "SELECT EXTRACT(HOUR FROM created_at)::int as hr, COUNT(*) as views "
        "FROM page_views WHERE created_at >= ? AND created_at <= ? "
        "GROUP BY hr ORDER BY hr",
        (insight_start, insight_end)
    ).fetchall()
    hourly_by_hour = {int(r['hr']): r['views'] for r in hourly_rows}
    hourly_data = [{'hour': h, 'views': hourly_by_hour.get(h, 0)} for h in range(24)]
    peak_hour = max(hourly_data, key=lambda h: h['views'], default=None)
    max_hourly_views = max((h['views'] for h in hourly_data), default=0) or 1

    source_rows = db.execute(
        "SELECT COALESCE(source, 'Direct / Unknown') as source, COUNT(*) as views, "
        "COUNT(DISTINCT visitor_id) as visitors "
        "FROM page_views WHERE created_at >= ? AND created_at <= ? "
        "GROUP BY source ORDER BY views DESC",
        (insight_start, insight_end)
    ).fetchall()
    source_data = [{'source': r['source'], 'views': r['views'], 'visitors': r['visitors']} for r in source_rows]
    max_source_views = max((s['views'] for s in source_data), default=0) or 1

    top_pages_rows = db.execute(
        "SELECT path, COUNT(*) as views, COUNT(DISTINCT visitor_id) as visitors "
        "FROM page_views WHERE created_at >= ? AND created_at <= ? "
        "GROUP BY path ORDER BY views DESC LIMIT 10",
        (insight_start, insight_end)
    ).fetchall()
    top_pages = [{'path': r['path'], 'views': r['views'], 'visitors': r['visitors']} for r in top_pages_rows]
    max_page_views = top_pages[0]['views'] if top_pages else 1

    return render_template('admin/admin_visitors.html',
                            range_mode=range_mode, from_date=from_date, to_date=to_date,
                            totals=totals, daily_data=daily_data, peak_day=peak_day, max_daily_visitors=max_daily_visitors,
                            hourly_data=hourly_data, peak_hour=peak_hour, max_hourly_views=max_hourly_views,
                            source_data=source_data, max_source_views=max_source_views,
                            top_pages=top_pages, max_page_views=max_page_views)


def send_weekly_report_email():
    """Emails ADMIN_EMAIL a rolling-7-day snapshot: unique visitors, orders
    placed, revenue, and the week's top 10 most-viewed products. Deliberately
    reuses the exact same window definition (rolling 7 days, status !=
    'cancelled' for revenue) as window_stats()/top_events() in
    admin_analytics() above, so these numbers always agree with the
    dashboard instead of drifting into a second, slightly-different
    definition of "revenue". Returns True/False like send_contact_email().
    """
    db = get_db()
    now = datetime.now()
    start = now - timedelta(days=7)
    start_str = start.strftime('%Y-%m-%d 00:00:00')
    end_str = now.strftime('%Y-%m-%d %H:%M:%S')

    visitor_row = db.execute(
        "SELECT COUNT(DISTINCT visitor_id) as visitors, COUNT(*) as views "
        "FROM page_views WHERE created_at >= ? AND created_at <= ?",
        (start_str, end_str)
    ).fetchone()
    visitors = visitor_row['visitors'] if visitor_row else 0
    page_views = visitor_row['views'] if visitor_row else 0

    order_row = db.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(total_amount),0) as revenue "
        "FROM order_shipping WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled'",
        (start_str, end_str)
    ).fetchone()
    order_count = order_row['cnt'] if order_row else 0
    revenue = float(order_row['revenue'] or 0) if order_row else 0.0

    top_rows = db.execute(
        "SELECT sku, COUNT(*) as cnt FROM product_events "
        "WHERE event_type='view' AND created_at >= ? AND created_at <= ? "
        "GROUP BY sku ORDER BY cnt DESC LIMIT 10",
        (start_str, end_str)
    ).fetchall()
    top_products = []
    for r in top_rows:
        p = db.execute('SELECT name FROM products WHERE sku=?', (r['sku'],)).fetchone()
        top_products.append({'name': p['name'] if p else r['sku'], 'views': r['cnt']})

    date_range_label = f"{start.strftime('%d %b')} – {now.strftime('%d %b %Y')}"

    html_body = render_template('retail/email_weekly_report.html',
        date_range_label=date_range_label, visitors=visitors, page_views=page_views,
        order_count=order_count, revenue=revenue, top_products=top_products)

    text_lines = [
        f"Weekly Business Report ({date_range_label})", "",
        f"Visitors: {visitors} ({page_views} page views)",
        f"Orders: {order_count}",
        f"Revenue: Rs.{revenue:.2f}", "",
        "Top viewed products:",
    ]
    if top_products:
        text_lines += [f"  {i+1}. {p['name']} — {p['views']} views" for i, p in enumerate(top_products)]
    else:
        text_lines.append("  (no product views this week)")
    text_body = "\n".join(text_lines)

    return send_contact_email(
        ADMIN_EMAIL, f"Weekly Report — {date_range_label}",
        text_body, html_body=html_body, from_email=SUPPORT_FROM_EMAIL
    )


@app.route('/cron/weekly-report', methods=['POST'])
def cron_weekly_report():
    """Triggered by a Render Cron Job (see render.yaml + weekly_report_cron.py)
    once a week. Not under /admin/ and not @admin_required -- a cron job has
    no browser session to authenticate with, so this checks a shared secret
    header instead. Returns 403 if the secret is missing/unset/wrong, so a
    stray request (or a forgotten/unset WEEKLY_REPORT_CRON_SECRET) can never
    trigger a send."""
    provided = request.headers.get('X-Cron-Secret', '')
    if not WEEKLY_REPORT_CRON_SECRET or not hmac.compare_digest(provided, WEEKLY_REPORT_CRON_SECRET):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
    ok = send_weekly_report_email()
    return jsonify({'status': 'success' if ok else 'error'}), (200 if ok else 500)


@app.route('/admin/manage-images', methods=['GET'])
@admin_required
def admin_manage_images():
    db = get_db()
    sku_search = request.args.get('sku_search', '').strip()
    if sku_search:
        products = db.execute(
            'SELECT * FROM products WHERE sku LIKE ? ORDER BY sku',
            (f'%{sku_search}%',)
        ).fetchall()
    else:
        products = db.execute('SELECT * FROM products ORDER BY sku').fetchall()
    return render_template('admin/admin_manage_images.html', products=products, sku_search=sku_search)


@app.route('/admin/edit-product-details', methods=['GET', 'POST'])
@admin_required
def admin_edit_product_details():
    db = get_db()
    if request.method == 'POST':
        sku = request.form.get('sku', '').strip()
        name = request.form.get('name', '').strip()
        retail_price = request.form.get('retail_price', 0)
        mrp_price = request.form.get('mrp_price', 0)
        stock_alert_threshold = int(request.form.get('stock_alert_threshold', 5) or 5)
        category = request.form.get('category', '').strip()
        material = request.form.get('material', '').strip()
        slug = request.form.get('slug', '').strip()

        is_bangle_row = 'bangle_size_stock' in request.form
        if is_bangle_row:
            stock_total = 0
            for size in BANGLE_SIZES:
                field = 'stock_' + size.replace('.', '_')
                size_stock = int(request.form.get(field, 0) or 0)
                stock_total += size_stock
                variant_sku = get_variant_sku(sku, size)
                db.execute(
                    'INSERT INTO product_variants (master_sku, variant_sku, size, stock_total) '
                    'VALUES (?, ?, ?, ?) ON CONFLICT (master_sku, size) '
                    'DO UPDATE SET stock_total=EXCLUDED.stock_total',
                    (sku, variant_sku, size, size_stock)
                )
        else:
            stock_total = int(request.form.get('stock_total', 0) or 0)

        db.execute(
            '''UPDATE products SET name=?, retail_price=?, mrp_price=?,
               stock_total=?, stock_alert_threshold=?, category=?, material=?, slug=? WHERE sku=?''',
            (name, retail_price, mrp_price, stock_total, stock_alert_threshold,
             category, material, slug, sku)
        )
        db.commit()
        # Send stock alert email if stock is at or below threshold
        if stock_total <= stock_alert_threshold:
            try:
                admin_email = os.environ.get('ADMIN_EMAIL', 'mohinicosmetics.india@gmail.com')
                send_contact_email(
                    admin_email,
                    f'⚠️ Low Stock Alert: {name} ({sku})',
                    f'Product: {name}\nSKU: {sku}\n'
                    f'Current Stock: {stock_total}\n'
                    f'Alert Threshold: {stock_alert_threshold}\n\n'
                    f'Please reorder this product soon.',
                )
                flash(f'Product {sku} updated. ⚠️ Stock alert sent — only {stock_total} units left!')
            except Exception as e:
                app.logger.warning(f'Stock alert email failed: {e}')
                flash(f'Product {sku} updated. ⚠️ Stock low ({stock_total} units).')
        else:
            flash(f'Product {sku} updated successfully.')
        return redirect(url_for('admin_edit_product_details'))
    raw_products = db.execute('SELECT * FROM products ORDER BY sku').fetchall()
    products = []
    for p in raw_products:
        p_dict = dict(p)
        if is_bangle_product(p_dict):
            p_dict['size_stock'] = get_bangle_size_stock(db, p_dict['sku'])
        products.append(p_dict)
    return render_template('admin/admin_edit_product_details.html', products=products)


@app.route('/admin/edit-product', methods=['GET'])
@admin_required
def admin_edit_product_search():
    db = get_db()
    q = request.args.get('q', '').strip()
    if q:
        like = f'%{q}%'
        products = db.execute(
            'SELECT * FROM products WHERE sku LIKE ? OR name LIKE ? ORDER BY updated_at DESC',
            (like, like)
        ).fetchall()
    else:
        products = db.execute('SELECT * FROM products ORDER BY updated_at DESC').fetchall()

    drafts = db.execute("SELECT * FROM products WHERE status='draft' ORDER BY updated_at DESC").fetchall()

    published_product = None
    published_id = request.args.get('published_id', type=int)
    if published_id:
        published_product = db.execute('SELECT id, name, sku FROM products WHERE id=?', (published_id,)).fetchone()

    return render_template('admin/admin_edit_product.html', products=products, q=q,
                            drafts=drafts, published_product=published_product)


@app.route('/admin/edit-product/<int:id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_product_form(id):
    db = get_db()
    product = db.execute('SELECT * FROM products WHERE id=?', (id,)).fetchone()
    if not product:
        flash('Product not found.')
        return redirect(url_for('admin_edit_product_search'))

    if request.method == 'POST':
        intent = request.form.get('intent', 'draft')
        new_sku = request.form.get('sku', '').strip()
        name = request.form.get('name', '').strip()
        category = request.form.get('category', '').strip()
        sub_category = request.form.get('sub_category', '').strip()
        collection = request.form.get('collection', '').strip()
        retail_price = float(request.form.get('retail_price', 0) or 0)
        mrp_price = float(request.form.get('mrp_price', 0) or 0)
        wholesale_price = float(request.form.get('wholesale_price', 0) or 0)
        stock_total = int(request.form.get('stock_total', 0) or 0)
        material = request.form.get('material', '').strip()
        brand_name = request.form.get('brand_name', '').strip()
        pack_unit = request.form.get('pack_unit', '').strip()
        pack_count = int(request.form.get('pack_count', 0) or 0)
        size = request.form.get('size', '').strip()
        hsn_code = request.form.get('hsn_code', '').strip()
        gst_percent = float(request.form.get('gst_percent', 3) or 3)
        weight_grams = float(request.form.get('weight_grams', 0) or 0)
        length = float(request.form.get('length', 0) or 0)
        breadth = float(request.form.get('breadth', 0) or 0)
        height = float(request.form.get('height', 0) or 0)
        sets_count = int(request.form.get('sets_count', 1) or 1)
        min_wholesale_qty = int(request.form.get('min_wholesale_qty', 0) or 0)
        price1 = float(request.form.get('price1', 0) or 0)
        quantity1 = int(request.form.get('quantity1', 0) or 0)
        price2 = float(request.form.get('price2', 0) or 0)
        quantity2 = int(request.form.get('quantity2', 0) or 0)
        price3 = float(request.form.get('price3', 0) or 0)
        quantity3 = int(request.form.get('quantity3', 0) or 0)
        description = request.form.get('description', '').strip()
        key_features_raw = request.form.get('key_features', '')
        key_features_list = [line.strip() for line in key_features_raw.splitlines() if line.strip()]
        key_features = '\n'.join(key_features_list)
        slug = request.form.get('slug', '').strip()
        if not slug and name:
            slug = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

        if not new_sku:
            flash('SKU is required.')
            return redirect(url_for('admin_edit_product_form', id=id))

        image_files = [f for f in request.files.getlist('images') if f and f.filename][:5]

        if intent == 'publish':
            if len(key_features_list) < 4:
                flash('Please add at least 4 key features before publishing (or save as draft).')
                return redirect(url_for('admin_edit_product_form', id=id))
            if not product['image_field'] and not image_files:
                flash('At least one product image is required before publishing (or save as draft).')
                return redirect(url_for('admin_edit_product_form', id=id))

        old_sku = product['sku']
        sku_changed = new_sku != old_sku
        if sku_changed:
            clash = db.execute('SELECT id FROM products WHERE sku=? AND id<>?', (new_sku, id)).fetchone()
            if clash:
                flash(f'Another product already uses SKU {new_sku}.')
                return redirect(url_for('admin_edit_product_form', id=id))

        image_url = product['image_field']
        for idx, image_file in enumerate(image_files, start=1):
            uploaded_url = upload_image_to_supabase(image_file, f"{new_sku}_{idx}.webp")
            if idx == 1:
                image_url = uploaded_url

        new_status = 'published' if intent == 'publish' else 'draft'
        new_is_active = 1 if intent == 'publish' else 0

        db.execute(
            '''UPDATE products SET
                 sku=?, name=?, category=?, sub_category=?, collection=?,
                 retail_price=?, mrp_price=?, wholesale_price=?, stock_total=?,
                 material=?, brand_name=?, pack_unit=?, pack_count=?, size=?, hsn_code=?, gst_percent=?,
                 weight_grams=?, length=?, breadth=?, height=?,
                 sets_count=?, min_wholesale_qty=?, slug=?,
                 price1=?, quantity1=?, price2=?, quantity2=?, price3=?, quantity3=?,
                 image_field=?, description=?, key_features=?, status=?, is_active=?, updated_at=NOW()
               WHERE id=?''',
            (new_sku, name, category, sub_category, collection,
             retail_price, mrp_price, wholesale_price, stock_total,
             material, brand_name, pack_unit, pack_count, size, hsn_code, gst_percent,
             weight_grams, length, breadth, height,
             sets_count, min_wholesale_qty, slug,
             price1, quantity1, price2, quantity2, price3, quantity3,
             image_url, description, key_features, new_status, new_is_active, id)
        )

        if sku_changed:
            variant_rows = db.execute(
                'SELECT size FROM product_variants WHERE master_sku=?', (old_sku,)
            ).fetchall()
            for v in variant_rows:
                new_variant_sku = get_variant_sku(new_sku, v['size'])
                db.execute(
                    'UPDATE product_variants SET master_sku=?, variant_sku=? WHERE master_sku=? AND size=?',
                    (new_sku, new_variant_sku, old_sku, v['size'])
                )

        db.commit()

        if intent == 'publish':
            flash(f'Product {new_sku} published successfully.')
            return redirect(url_for('admin_edit_product_search', published_id=id))
        else:
            flash(f'Draft {new_sku} saved.')
            return redirect(url_for('admin_edit_product_form', id=id))

    p_dict = dict(product)
    variants = {}
    if is_bangle_product(p_dict):
        variants = get_bangle_size_stock(db, p_dict['sku'])
    drafts = db.execute(
        "SELECT * FROM products WHERE status='draft' AND id<>? ORDER BY updated_at DESC", (id,)
    ).fetchall()
    return render_template('admin/admin_edit_product_form.html', product=p_dict, variants=variants, drafts=drafts)


@app.route('/admin/delete-products', methods=['GET'])
@admin_required
def admin_delete_products():
    db = get_db()
    products = db.execute('SELECT * FROM products ORDER BY sku').fetchall()
    return render_template('admin/admin_delete_products.html', products=products)


@app.route('/admin/delete-product/<int:product_id>', methods=['GET'])
@admin_required
def admin_delete_product(product_id):
    db = get_db()
    db.execute('DELETE FROM products WHERE id=?', (product_id,))
    db.commit()
    flash('Product deleted successfully.')
    return redirect(url_for('admin_delete_products'))


@app.route('/admin/inbox', methods=['GET'])
@admin_required
def admin_inbox():
    db = get_db()
    quotes = db.execute('SELECT * FROM quotes ORDER BY id DESC').fetchall()
    return render_template('admin/admin_inbox.html', quotes=quotes)


@app.route('/admin/quote/<int:quote_id>', methods=['GET', 'POST'])
@admin_required
def admin_quote_view(quote_id):
    db = get_db()
    quote = db.execute('SELECT * FROM quotes WHERE id=?', (quote_id,)).fetchone()
    if not quote:
        flash('Quote not found.')
        return redirect(url_for('admin_inbox'))
    cart_items = []
    try:
        cart_items = json.loads(quote['items_json'] or '[]')
    except Exception:
        cart_items = []
    if request.method == 'POST' and request.form.get('mark_contacted'):
        db.execute('UPDATE quotes SET status=? WHERE id=?', ('Contacted', quote_id))
        db.commit()
        flash('Quote marked as contacted.')
        return redirect(url_for('admin_quote_view', quote_id=quote_id))
    return render_template('admin/admin_quote_view.html', quote=quote, cart_items=cart_items)


@app.route('/admin/add-product', methods=['GET', 'POST'])
@admin_required
def admin_add_product():
    db = get_db()
    if request.method == 'POST':
        intent = request.form.get('intent', 'publish')
        sku = request.form.get('sku', '').strip()
        name = request.form.get('name', '').strip()
        category = request.form.get('category', '').strip()
        sub_category = request.form.get('sub_category', '').strip()
        collection = request.form.get('collection', '').strip()
        retail_price = float(request.form.get('retail_price', 0) or 0)
        mrp_price = float(request.form.get('mrp_price', 0) or 0)
        wholesale_price = float(request.form.get('wholesale_price', 0) or 0)
        stock_total = int(request.form.get('stock_total', 0) or 0)
        material = request.form.get('material', '').strip()
        brand_name = request.form.get('brand_name', '').strip()
        pack_unit = request.form.get('pack_unit', '').strip()
        pack_count = int(request.form.get('pack_count', 0) or 0)
        size = request.form.get('size', '').strip()
        hsn_code = request.form.get('hsn_code', '').strip()
        gst_percent = float(request.form.get('gst_percent', 3) or 3)
        weight_grams = float(request.form.get('weight_grams', 0) or 0)
        length = float(request.form.get('length', 0) or 0)
        breadth = float(request.form.get('breadth', 0) or 0)
        height = float(request.form.get('height', 0) or 0)
        sets_count = int(request.form.get('sets_count', 1) or 1)
        min_wholesale_qty = int(request.form.get('min_wholesale_qty', 0) or 0)
        price1 = float(request.form.get('price1', 0) or 0)
        quantity1 = int(request.form.get('quantity1', 0) or 0)
        price2 = float(request.form.get('price2', 0) or 0)
        quantity2 = int(request.form.get('quantity2', 0) or 0)
        price3 = float(request.form.get('price3', 0) or 0)
        quantity3 = int(request.form.get('quantity3', 0) or 0)
        description = request.form.get('description', '').strip()
        key_features_raw = request.form.get('key_features', '')
        key_features_list = [line.strip() for line in key_features_raw.splitlines() if line.strip()]
        key_features = '\n'.join(key_features_list)

        # Auto-generate slug from name if not provided
        slug = request.form.get('slug', '').strip()
        if not slug and name:
            slug = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

        if not sku:
            flash('SKU is required.')
            return redirect(url_for('admin_add_product'))

        image_files = [f for f in request.files.getlist('images') if f and f.filename][:5]

        if intent == 'publish':
            if len(name) > 48:
                flash('Product name must be 48 characters or fewer (excluding brand name), or save as draft.')
                return redirect(url_for('admin_add_product'))
            if len(key_features_list) < 4:
                flash('Please add at least 4 key features (one per line), or save as draft.')
                return redirect(url_for('admin_add_product'))
            if not image_files:
                flash('At least one product image is required to publish, or save as draft.')
                return redirect(url_for('admin_add_product'))

        existing = db.execute('SELECT id FROM products WHERE sku=?', (sku,)).fetchone()
        if existing:
            flash(f'A product with SKU {sku} already exists.')
            return redirect(url_for('admin_add_product'))

        image_url = None
        for idx, image_file in enumerate(image_files, start=1):
            uploaded_url = upload_image_to_supabase(image_file, f"{sku}_{idx}.webp")
            if idx == 1:
                image_url = uploaded_url

        new_status = 'published' if intent == 'publish' else 'draft'
        new_is_active = 1 if intent == 'publish' else 0
        model_number = generate_model_number(db)

        db.execute(
            '''INSERT INTO products
               (sku, name, category, sub_category, collection,
                retail_price, mrp_price, wholesale_price,
                stock_total, material, brand_name, pack_unit, pack_count, size, hsn_code, gst_percent,
                weight_grams, length, breadth, height,
                sets_count, min_wholesale_qty,
                slug, price1, quantity1, price2, quantity2,
                price3, quantity3, image_field, description, key_features,
                status, is_active, model_number, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,NOW())''',
            (sku, name, category, sub_category, collection,
             retail_price, mrp_price, wholesale_price,
             stock_total, material, brand_name, pack_unit, pack_count, size, hsn_code, gst_percent,
             weight_grams, length, breadth, height,
             sets_count, min_wholesale_qty,
             slug, price1, quantity1, price2, quantity2,
             price3, quantity3, image_url, description, key_features,
             new_status, new_is_active, model_number)
        )
        db.commit()

        if intent == 'publish':
            new_id_row = db.execute('SELECT id FROM products WHERE sku=?', (sku,)).fetchone()
            new_id = new_id_row['id'] if new_id_row else None
            flash(f'Product {sku} published successfully.')
            return redirect(url_for('admin_add_product', published_id=new_id))
        else:
            flash(f'Draft {sku} saved. Continue editing it any time from the Drafts list below.')
            return redirect(url_for('admin_add_product'))

    drafts = db.execute("SELECT * FROM products WHERE status='draft' ORDER BY updated_at DESC").fetchall()
    published_product = None
    published_id = request.args.get('published_id', type=int)
    if published_id:
        published_product = db.execute('SELECT id, name, sku FROM products WHERE id=?', (published_id,)).fetchone()
    return render_template('admin/admin_add_product.html', drafts=drafts, published_product=published_product)


@app.route('/admin/download-users-excel', methods=['GET'])
@admin_required
def download_users_excel():
    flash('Users export coming soon.')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/download-quotes-excel', methods=['GET'])
@admin_required
def download_quotes_excel():
    flash('Quotes export coming soon.')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/download-products-excel', methods=['GET'])
@admin_required
def download_products_excel():
    import io
    try:
        import openpyxl
        from flask import send_file
        conn = get_db()
        products = conn.execute(
            'SELECT sku, name, category, sub_category, description, '
            'retail_price, mrp_price, wholesale_price, min_wholesale_qty, '
            'gst_percent, hsn_code, material, weight_grams, '
            'stock_total, is_active FROM products ORDER BY category, name'
        ).fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Products'
        headers = ['SKU', 'Name', 'Category', 'Sub Category', 'Description',
                   'Retail Price', 'MRP Price', 'Wholesale Price',
                   'Min Wholesale Qty', 'GST %', 'HSN Code',
                   'Material', 'Weight (g)', 'Stock', 'Active']
        ws.append(headers)
        # Bold header row
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for p in products:
            ws.append([
                p['sku'], p['name'], p['category'], p['sub_category'],
                p['description'] or '',
                p['retail_price'], p['mrp_price'], p['wholesale_price'],
                p['min_wholesale_qty'], p['gst_percent'], p['hsn_code'],
                p['material'], p['weight_grams'], p['stock_total'],
                'Yes' if p['is_active'] else 'No'
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        from datetime import date
        filename = f"NariNakhre_Products_{date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f'Product excel export failed: {e}')
        flash(f'Export failed: {e}')
        return redirect(url_for('admin_dashboard'))


def resolve_gst_range(req):
    """Date-range resolution for the GST report pages -- defaults to the
    current calendar month (the common unit admins file GST for), unlike
    resolve_insight_range()'s trailing-7-days default which suits day-to-day
    analytics rather than monthly filing."""
    now = datetime.now()
    from_date = req.args.get('from', '').strip()
    to_date = req.args.get('to', '').strip()
    if not (from_date and to_date):
        from_date = now.strftime('%Y-%m-01')
        to_date = now.strftime('%Y-%m-%d')
    return from_date, to_date, from_date + ' 00:00:00', to_date + ' 23:59:59'


def _compute_order_gst_lines(db, order):
    """Parse an order_shipping row's cart_items_json and return per-line GST
    detail (hsn_code, gst_percent, qty, taxable_value, cgst, sgst, igst)
    using each SKU's live products.hsn_code/gst_percent, plus whether the
    order is inter-state (IGST) vs intra-state (CGST+SGST) based on
    consignee_state vs WAREHOUSE_STATE.

    The invoice.html template hardcodes HSN 7117 / GST 3% for every line and
    always reports a 50/50 CGST+SGST split regardless of state -- fine for a
    quick on-screen invoice, but not accurate enough for GST filing, so GST
    reports recompute properly from the real per-product HSN/GST% here."""
    try:
        items = json.loads(order['cart_items_json'] or '[]')
    except Exception:
        items = []
    is_interstate = (order['consignee_state'] or '').strip().lower() != WAREHOUSE_STATE.strip().lower()
    lines = []
    for item in items:
        sku = item.get('sku') or ''
        qty = int(item.get('units') or item.get('qty') or 1)
        price = float(item.get('price') or 0)
        line_total = price * qty
        if line_total <= 0:
            continue
        hsn_code = ''
        gst_percent = 3.0
        if sku:
            prod = db.execute('SELECT hsn_code, gst_percent FROM products WHERE sku=?', (sku,)).fetchone()
            if prod:
                hsn_code = prod['hsn_code'] or ''
                if prod['gst_percent']:
                    gst_percent = float(prod['gst_percent'])
        taxable_value = round(line_total / (1 + gst_percent / 100.0), 2)
        line_gst = round(line_total - taxable_value, 2)
        cgst = sgst = igst = 0.0
        if is_interstate:
            igst = line_gst
        else:
            cgst = round(line_gst / 2, 2)
            sgst = round(line_gst - cgst, 2)
        lines.append({
            'sku': sku, 'name': item.get('name', ''), 'hsn_code': hsn_code, 'gst_percent': gst_percent,
            'qty': qty, 'line_total': line_total, 'taxable_value': taxable_value,
            'cgst': cgst, 'sgst': sgst, 'igst': igst,
        })
    return lines, is_interstate


@app.route('/admin/reports/gst', methods=['GET'])
@admin_required
def admin_gst_reports():
    from_date, to_date, range_start, range_end = resolve_gst_range(request)
    db = get_db()
    row = db.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(total_amount),0) as revenue FROM order_shipping "
        "WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled'",
        (range_start, range_end)
    ).fetchone()
    return render_template('admin/admin_gst_reports.html',
        from_date=from_date, to_date=to_date,
        order_count=row['cnt'] if row else 0,
        revenue=float(row['revenue'] or 0) if row else 0.0)


@app.route('/admin/reports/gst/invoices-pdf', methods=['GET'])
@admin_required
def admin_gst_invoices_pdf():
    from flask import send_file
    from_date, to_date, range_start, range_end = resolve_gst_range(request)
    db = get_db()
    orders = db.execute(
        "SELECT * FROM order_shipping WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled' ORDER BY created_at",
        (range_start, range_end)
    ).fetchall()

    if not orders:
        flash(f'No orders found between {from_date} and {to_date}.')
        return redirect(url_for('admin_gst_reports', **{'from': from_date, 'to': to_date}))

    MAX_INVOICES = 300
    if len(orders) > MAX_INVOICES:
        flash(f'That range has {len(orders)} orders -- please narrow it to {MAX_INVOICES} or fewer for a single PDF.')
        return redirect(url_for('admin_gst_reports', **{'from': from_date, 'to': to_date}))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm,
                                 leftMargin=15 * mm, rightMargin=15 * mm)
        styles = getSampleStyleSheet()
        rose = colors.HexColor('#be185d')
        gray = colors.HexColor('#6b7280')
        normal_style = ParagraphStyle('Normal2', parent=styles['Normal'], fontSize=9, leading=13)
        small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=gray, leading=11)

        elements = []
        for idx, order in enumerate(orders):
            lines, is_interstate = _compute_order_gst_lines(db, order)

            header_data = [[
                Paragraph('<b>Nari Nakhre&trade;</b><br/>by Mohini Cosmetics' +
                          (f'<br/><font color="#1d4ed8">GSTIN: {DELHIVERY_SELLER_GST}</font>' if DELHIVERY_SELLER_GST else ''),
                          normal_style),
                Paragraph(f'<para align="right"><b><font color="#be185d" size=15>TAX INVOICE</font></b><br/>'
                          f'Invoice No: {order["internal_order_id"]}<br/>'
                          f'Date: {str(order["created_at"])[:10] if order["created_at"] else "N/A"}</para>', normal_style),
            ]]
            header_table = Table(header_data, colWidths=[90 * mm, 90 * mm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LINEBELOW', (0, 0), (-1, -1), 1.2, rose),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 10))

            seller_text = ('<b>SOLD BY</b><br/><b>Mohini Cosmetics</b> (Trading as: Nari Nakhre)<br/>' +
                            (WAREHOUSE_ADDRESS or 'Jabalpur, Madhya Pradesh - 482001') +
                            '<br/>Email: info@narinakhre.com' +
                            (f'<br/>GSTIN: {DELHIVERY_SELLER_GST}' if DELHIVERY_SELLER_GST else ''))
            buyer_text = (f"<b>BILLED &amp; SHIPPED TO</b><br/><b>{order['consignee_name']}</b><br/>"
                          f"{order['consignee_address']}<br/>"
                          f"{order['consignee_city']}, {order['consignee_state']} - {order['consignee_pincode']}<br/>"
                          f"Phone: {order['consignee_phone']}" +
                          (f"<br/>Email: {order['consignee_email']}" if order['consignee_email'] else ''))
            parties_table = Table([[Paragraph(seller_text, normal_style), Paragraph(buyer_text, normal_style)]],
                                   colWidths=[90 * mm, 90 * mm])
            parties_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
            elements.append(parties_table)
            elements.append(Spacer(1, 14))

            item_rows = [['#', 'Description', 'HSN', 'GST%', 'Taxable', 'GST Amt', 'Total']]
            for i, line in enumerate(lines, start=1):
                item_rows.append([
                    str(i),
                    Paragraph(f"{line['name']}<br/><font size=7 color='#9f9f9f'>SKU: {line['sku']} | Qty: {line['qty']}</font>", small_style),
                    line['hsn_code'] or '-',
                    f"{line['gst_percent']:.1f}%",
                    f"Rs.{line['taxable_value']:.2f}",
                    f"Rs.{(line['cgst'] + line['sgst'] + line['igst']):.2f}",
                    f"Rs.{line['line_total']:.2f}",
                ])
            if not lines:
                item_rows.append(['1', order['internal_order_id'], '-', '-', '-', '-',
                                   f"Rs.{float(order['total_amount'] or 0):.2f}"])

            items_table = Table(item_rows, colWidths=[8 * mm, 72 * mm, 16 * mm, 14 * mm, 24 * mm, 24 * mm, 24 * mm])
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fdf2f8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#881337')),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e5e7eb')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(items_table)
            elements.append(Spacer(1, 12))

            total_taxable = sum(l['taxable_value'] for l in lines)
            total_cgst = sum(l['cgst'] for l in lines)
            total_sgst = sum(l['sgst'] for l in lines)
            total_igst = sum(l['igst'] for l in lines)
            gst_label = 'Inter-State Supply (IGST)' if is_interstate else 'Intra-State Supply (CGST + SGST)'
            if is_interstate:
                gst_rows = [['Taxable Value', 'IGST', 'Total GST'],
                            [f"Rs.{total_taxable:.2f}", f"Rs.{total_igst:.2f}", f"Rs.{total_igst:.2f}"]]
            else:
                gst_rows = [['Taxable Value', 'CGST', 'SGST', 'Total GST'],
                            [f"Rs.{total_taxable:.2f}", f"Rs.{total_cgst:.2f}", f"Rs.{total_sgst:.2f}",
                             f"Rs.{(total_cgst + total_sgst):.2f}"]]
            gst_table = Table(gst_rows)
            gst_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e5e7eb')),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(Paragraph(f"<b>{gst_label}</b>", small_style))
            elements.append(Spacer(1, 4))
            elements.append(gst_table)
            elements.append(Spacer(1, 12))

            elements.append(Paragraph(
                f"<b>Payment Mode:</b> {'Cash on Delivery' if order['payment_mode'] == 'COD' else 'Prepaid (Online)'}"
                f"&nbsp;&nbsp;&nbsp; <b>Status:</b> {(order['status'] or '').title()}", small_style))
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(
                f"<para align='right'><b><font size=13 color='#be185d'>Amount Paid: Rs.{float(order['total_amount'] or 0):.2f}</font></b></para>",
                normal_style))

            if idx < len(orders) - 1:
                elements.append(PageBreak())

        doc.build(elements)
        buf.seek(0)
        filename = f"NariNakhre_Invoices_{from_date}_to_{to_date}.pdf"
        return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f'Bulk invoice PDF export failed: {e}')
        flash(f'Invoice PDF export failed: {e}')
        return redirect(url_for('admin_gst_reports', **{'from': from_date, 'to': to_date}))


@app.route('/admin/reports/gst/excel', methods=['GET'])
@admin_required
def admin_gst_report_excel():
    import openpyxl
    from openpyxl.styles import Font
    from flask import send_file
    from_date, to_date, range_start, range_end = resolve_gst_range(request)
    db = get_db()
    orders = db.execute(
        "SELECT * FROM order_shipping WHERE created_at >= ? AND created_at <= ? AND status != 'cancelled' ORDER BY created_at",
        (range_start, range_end)
    ).fetchall()

    if not orders:
        flash(f'No orders found between {from_date} and {to_date}.')
        return redirect(url_for('admin_gst_reports', **{'from': from_date, 'to': to_date}))

    try:
        hsn_summary = {}   # (hsn_code, gst_percent) -> totals
        state_summary = {}  # consignee_state -> totals
        sales_rows = []

        for order in orders:
            lines, is_interstate = _compute_order_gst_lines(db, order)
            order_taxable = sum(l['taxable_value'] for l in lines)
            order_cgst = sum(l['cgst'] for l in lines)
            order_sgst = sum(l['sgst'] for l in lines)
            order_igst = sum(l['igst'] for l in lines)

            sales_rows.append([
                str(order['created_at'])[:10] if order['created_at'] else '',
                order['internal_order_id'], order['consignee_name'], order['consignee_state'],
                order['consignee_pincode'], order['payment_mode'],
                'Inter-State (IGST)' if is_interstate else 'Intra-State (CGST+SGST)',
                round(order_taxable, 2), round(order_cgst, 2), round(order_sgst, 2), round(order_igst, 2),
                round(order_cgst + order_sgst + order_igst, 2), float(order['total_amount'] or 0),
            ])

            for line in lines:
                key = (line['hsn_code'] or 'N/A', line['gst_percent'])
                h = hsn_summary.setdefault(key, {'qty': 0, 'taxable': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0})
                h['qty'] += line['qty']
                h['taxable'] += line['taxable_value']
                h['cgst'] += line['cgst']
                h['sgst'] += line['sgst']
                h['igst'] += line['igst']

            state_key = order['consignee_state'] or 'Unknown'
            s = state_summary.setdefault(state_key, {'orders': 0, 'taxable': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'total': 0.0})
            s['orders'] += 1
            s['taxable'] += order_taxable
            s['cgst'] += order_cgst
            s['sgst'] += order_sgst
            s['igst'] += order_igst
            s['total'] += float(order['total_amount'] or 0)

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = 'Sales Register'
        ws1.append(['Date', 'Invoice No', 'Customer', 'State', 'Pincode', 'Payment Mode', 'Supply Type',
                    'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total GST', 'Total Amount'])
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        for row in sales_rows:
            ws1.append(row)

        ws2 = wb.create_sheet('HSN Summary')
        ws2.append(['HSN Code', 'GST %', 'Total Quantity', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total Tax', 'Total Value'])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for (hsn_code, gst_percent), h in sorted(hsn_summary.items()):
            total_tax = h['cgst'] + h['sgst'] + h['igst']
            ws2.append([hsn_code, gst_percent, h['qty'], round(h['taxable'], 2), round(h['cgst'], 2),
                        round(h['sgst'], 2), round(h['igst'], 2), round(total_tax, 2), round(h['taxable'] + total_tax, 2)])

        ws3 = wb.create_sheet('State Summary')
        ws3.append(['State', 'Orders', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total Tax', 'Total Value'])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
        for state, s in sorted(state_summary.items(), key=lambda kv: -kv[1]['total']):
            total_tax = s['cgst'] + s['sgst'] + s['igst']
            ws3.append([state, s['orders'], round(s['taxable'], 2), round(s['cgst'], 2), round(s['sgst'], 2),
                        round(s['igst'], 2), round(total_tax, 2), round(s['total'], 2)])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"NariNakhre_GST_Report_{from_date}_to_{to_date}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f'GST excel report failed: {e}')
        flash(f'GST report export failed: {e}')
        return redirect(url_for('admin_gst_reports', **{'from': from_date, 'to': to_date}))


@app.route('/admin/upload-excel', methods=['POST'])
@admin_required
def admin_upload_excel():
    import pandas as pd

    file_object = request.files.get('excel_file')
    if file_object is None or not file_object.filename:
        flash('Please select an Excel or CSV catalog file to upload.')
        return redirect(url_for('admin_dashboard'))

    def normalize_value(value):
        if pd.isna(value):
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned if cleaned else None
        return value

    def to_float(value, default=0.0):
        value = normalize_value(value)
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def to_int(value, default=0):
        value = normalize_value(value)
        if value is None:
            return default
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    def to_bool(value, default=False):
        value = normalize_value(value)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().lower()
        if text in {'1', 'true', 'yes', 'y', 'active'}:
            return True
        if text in {'0', 'false', 'no', 'n', 'inactive'}:
            return False
        return default

    def row_value(row, key):
        if key not in row.index:
            return None
        return normalize_value(row.get(key))

    try:
        filename_lower = (file_object.filename or '').lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(file_object)
        elif filename_lower.endswith('.xlsx'):
            df = pd.read_excel(file_object)
        else:
            flash('Unsupported file format. Please upload a .csv or .xlsx file.')
            return redirect(url_for('admin_dashboard'))

        # Normalize column names to match exactly what the code expects.
        # Maps every header variant from the admin-exported Excel.
        col_aliases = {
            # As exported by download-products-excel
            'sub category':         'sub_category',
            'retail price':         'retail_price',
            'mrp price':            'mrp_price',
            'wholesale price':      'wholesale_price',
            'min wholesale qty':    'min_wholesale_qty',
            'gst %':                'gst_percent',
            'hsn code':             'hsn_code',
            'weight (g)':           'weight_grams',
            'stock':                'stock_total',
            'active':               'is_active',
            # Common alternates
            'mrp':                  'mrp_price',
            'selling price':        'retail_price',
            'sale price':           'retail_price',
            'price':                'retail_price',
            'discount %':           'retail_discount_percent',
            'discount percent':     'retail_discount_percent',
            'gst':                  'gst_percent',
            'gst percent':          'gst_percent',
            'qty':                  'stock_total',
            'quantity':             'stock_total',
            'weight g':             'weight_grams',
            'weight':               'weight_grams',
            'wt':                   'weight_grams',
            'sub_category':         'sub_category',  # already correct
        }
        # Lowercase + strip headers first, then apply alias map
        df.columns = df.columns.str.lower().str.strip()
        df.rename(columns=col_aliases, inplace=True)
        app.logger.info(f"Excel columns after normalization: {list(df.columns)}")

        processed_rows = 0
        created_rows = 0
        updated_rows = 0

        conn = get_db()
        for _, row in df.iterrows():
            row_sku = normalize_value(row.get('sku'))
            if not row_sku:
                continue
            row_sku = str(row_sku).strip()

            existing = conn.execute('SELECT id FROM products WHERE sku=?', (row_sku,)).fetchone()
            is_new = existing is None

            existing_image = None
            if not is_new:
                img_row = conn.execute('SELECT image_field FROM products WHERE sku=?', (row_sku,)).fetchone()
                if img_row:
                    existing_image = img_row['image_field']

            sheet_image = row_value(row, 'image_field')
            final_image = str(sheet_image) if sheet_image else existing_image

            values = (
                row_value(row, 'name'),
                row_value(row, 'slug'),
                row_value(row, 'category'),
                row_value(row, 'sub_category'),
                row_value(row, 'collection'),
                row_value(row, 'size'),
                to_float(row_value(row, 'retail_price')),
                to_float(row_value(row, 'mrp_price')),
                to_float(row_value(row, 'retail_discount_percent')),
                to_float(row_value(row, 'wholesale_price')),
                to_int(row_value(row, 'min_wholesale_qty')),
                to_int(row_value(row, 'sets_count')),
                final_image,
                to_float(row_value(row, 'price1')),
                to_int(row_value(row, 'quantity1')),
                to_float(row_value(row, 'price2')),
                to_int(row_value(row, 'quantity2')),
                to_float(row_value(row, 'price3')),
                to_int(row_value(row, 'quantity3')),
                to_float(row_value(row, 'purchase_cost')),
                to_float(row_value(row, 'making_charges')),
                to_float(row_value(row, 'weight_grams')),
                row_value(row, 'material'),
                row_value(row, 'hsn_code'),
                to_float(row_value(row, 'gst_percent')),
                to_int(row_value(row, 'stock_total'), default=0),
                row_value(row, 'box_packing_type'),
                row_value(row, 'vendor_id'),
                row_value(row, 'status'),
                1 if to_bool(row_value(row, 'is_active'), default=True) else 0,
                1 if to_bool(row_value(row, 'is_featured'), default=False) else 0,
                to_float(row_value(row, 'weight')),
                to_float(row_value(row, 'length')),
                to_float(row_value(row, 'breadth')),
                to_float(row_value(row, 'height')),
            )

            if is_new:
                model_number = generate_model_number(conn)
                conn.execute(
                    '''INSERT INTO products
                       (name, slug, category, sub_category, collection, size,
                        retail_price, mrp_price, retail_discount_percent, wholesale_price,
                        min_wholesale_qty, sets_count, image_field,
                        price1, quantity1, price2, quantity2, price3, quantity3,
                        purchase_cost, making_charges, weight_grams, material,
                        hsn_code, gst_percent, stock_total, box_packing_type,
                        vendor_id, status, is_active, is_featured,
                        weight, length, breadth, height, sku, model_number)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    values + (row_sku, model_number)
                )
                created_rows += 1
            else:
                # Only update columns present AND non-empty in this Excel row
                # Prevents zeroing prices when only stock_total column is uploaded
                def _has_val(col):
                    if col not in row.index:
                        return False
                    v = row.get(col)
                    try:
                        import math
                        if isinstance(v, float) and math.isnan(v):
                            return False
                    except Exception:
                        pass
                    return v is not None and str(v).strip() not in ('', 'nan', 'None')

                num_cols = {
                    'retail_price','mrp_price','retail_discount_percent',
                    'wholesale_price','min_wholesale_qty','sets_count',
                    'price1','price2','price3','quantity1','quantity2','quantity3',
                    'purchase_cost','making_charges','weight_grams','gst_percent',
                    'stock_total','weight','length','breadth','height'
                }
                bool_cols = {'is_active', 'is_featured'}
                col_map = [
                    ('name', row_value(row, 'name')),
                    ('slug', row_value(row, 'slug')),
                    ('category', row_value(row, 'category')),
                    ('sub_category', row_value(row, 'sub_category')),
                    ('collection', row_value(row, 'collection')),
                    ('size', row_value(row, 'size')),
                    ('retail_price', to_float(row_value(row, 'retail_price'))),
                    ('mrp_price', to_float(row_value(row, 'mrp_price'))),
                    ('retail_discount_percent', to_float(row_value(row, 'retail_discount_percent'))),
                    ('wholesale_price', to_float(row_value(row, 'wholesale_price'))),
                    ('min_wholesale_qty', to_int(row_value(row, 'min_wholesale_qty'))),
                    ('sets_count', to_int(row_value(row, 'sets_count'))),
                    ('image_field', final_image if sheet_image else None),
                    ('price1', to_float(row_value(row, 'price1'))),
                    ('quantity1', to_int(row_value(row, 'quantity1'))),
                    ('price2', to_float(row_value(row, 'price2'))),
                    ('quantity2', to_int(row_value(row, 'quantity2'))),
                    ('price3', to_float(row_value(row, 'price3'))),
                    ('quantity3', to_int(row_value(row, 'quantity3'))),
                    ('purchase_cost', to_float(row_value(row, 'purchase_cost'))),
                    ('making_charges', to_float(row_value(row, 'making_charges'))),
                    ('weight_grams', to_float(row_value(row, 'weight_grams'))),
                    ('material', row_value(row, 'material')),
                    ('hsn_code', row_value(row, 'hsn_code')),
                    ('gst_percent', to_float(row_value(row, 'gst_percent'))),
                    ('stock_total', to_int(row_value(row, 'stock_total'))),
                    ('box_packing_type', row_value(row, 'box_packing_type')),
                    ('vendor_id', row_value(row, 'vendor_id')),
                    ('status', row_value(row, 'status')),
                    ('is_active', 1 if to_bool(row_value(row, 'is_active'), default=True) else 0),
                    ('is_featured', 1 if to_bool(row_value(row, 'is_featured'), default=False) else 0),
                    ('weight', to_float(row_value(row, 'weight'))),
                    ('length', to_float(row_value(row, 'length'))),
                    ('breadth', to_float(row_value(row, 'breadth'))),
                    ('height', to_float(row_value(row, 'height'))),
                ]
                to_set = []
                for col, val in col_map:
                    if col == 'image_field' and val is None:
                        continue
                    if col in num_cols and not _has_val(col):
                        continue
                    if col in bool_cols and not _has_val(col):
                        continue
                    if col not in num_cols and col not in bool_cols and val is None:
                        continue
                    to_set.append((col, val))
                if to_set:
                    set_clause = ', '.join(f'{c}=?' for c, _ in to_set)
                    vals = [v for _, v in to_set] + [row_sku]
                    conn.execute(f'UPDATE products SET {set_clause} WHERE sku=?', vals)
                updated_rows += 1

            processed_rows += 1

        conn.commit()
        price_cols_found = [c for c in ['retail_price','mrp_price','stock_total'] if c in df.columns]
        # Send stock alerts for any products that hit the threshold during this upload
        if 'stock_total' in df.columns:
            try:
                low_stock = db.execute(
                    "SELECT sku, name, stock_total, stock_alert_threshold FROM products "
                    "WHERE stock_total <= stock_alert_threshold AND stock_total >= 0"
                ).fetchall()
                if low_stock:
                    items_str = '\n'.join(
                        f"  - {r['name']} (SKU: {r['sku']}): {r['stock_total']} units "
                        f"(alert at {r['stock_alert_threshold']})"
                        for r in low_stock
                    )
                    admin_email = os.environ.get('ADMIN_EMAIL', 'mohinicosmetics.india@gmail.com')
                    send_contact_email(
                        admin_email,
                        f'⚠️ Low Stock Alert — {len(low_stock)} product(s) need reordering',
                        f'The following products are at or below their stock alert threshold:\n\n'
                        f'{items_str}\n\nPlease reorder soon.',
                    )
            except Exception as e:
                app.logger.warning(f'Bulk stock alert email failed: {e}')
        flash(
            f'Sync complete: {processed_rows} rows processed '
            f'({created_rows} created, {updated_rows} updated). '
            f'Price columns detected: {price_cols_found or "NONE — check column headers in Excel"}'
        )
        return redirect(url_for('admin_dashboard'))
    except Exception as exc:
        flash(f'Catalog sync failed: {exc}')
        return redirect(url_for('admin_dashboard'))

@app.route('/order/<order_id>')
def customer_order_detail(order_id):
    """Public, shareable per-order details page -- retail only. Combines
    what the separate /track and /invoice pages show (items, amounts,
    shipping address, live tracking) into one Amazon-style order page,
    since neither of those alone covers everything a customer wants to see
    about a single order. Same no-login, unguessable-ID access pattern as
    /track/<waybill> and /invoice/<order_id> -- guest checkouts need to be
    able to open this link too, not just signed-in accounts."""
    if g.site_type != 'retail':
        return redirect('/')
    conn = get_db()
    order = conn.execute(
        'SELECT * FROM order_shipping WHERE internal_order_id=?', (order_id,)
    ).fetchone()
    if not order:
        return "Order not found", 404

    items = json.loads(order['cart_items_json']) if order['cart_items_json'] else []
    for item in items:
        item['image_url'] = ''
        sku = item.get('sku', '')
        if sku:
            product = conn.execute(
                'SELECT id, image_field FROM products WHERE sku=?', (sku,)
            ).fetchone()
            if product:
                item['product_id'] = product['id']
                try:
                    p_dict = dict(product)
                    p_dict['sku'] = sku
                    imgs = get_product_images(p_dict)
                    if imgs and imgs[0].startswith('http'):
                        item['image_url'] = imgs[0]
                except Exception:
                    pass

    return render_template('retail/order_detail.html', order=order, items=items)


@app.route('/invoice/<order_id>')
def customer_invoice(order_id):
    """Public invoice page for customers — link sent via email."""
    conn = get_db()
    order = conn.execute(
        'SELECT * FROM order_shipping WHERE internal_order_id=?', (order_id,)
    ).fetchone()
    if not order:
        return "Invoice not found", 404
    return render_template('admin/invoice.html', order=order,
                           seller_gst=DELHIVERY_SELLER_GST,
                           seller_name='Nari Nakhre',
                           seller_address=WAREHOUSE_ADDRESS)


@app.route('/admin/orders')
@admin_required
def admin_orders():
    conn = get_db()
    current_status = request.args.get('status', 'all')
    if current_status == 'all':
        orders = conn.execute("SELECT * FROM order_shipping ORDER BY id DESC LIMIT 200").fetchall()
    else:
        orders = conn.execute("SELECT * FROM order_shipping WHERE status=? ORDER BY id DESC", (current_status,)).fetchall()
    status_counts = conn.execute("SELECT status, COUNT(*) as count FROM order_shipping GROUP BY status").fetchall()
    count_map = {r['status']: r['count'] for r in status_counts}
    stats = [
        {'label':'All','count':sum(count_map.values()),'color':'#374151'},
        {'label':'Paid','count':count_map.get('paid',0),'color':'#059669'},
        {'label':'COD','count':count_map.get('cod_confirmed',0),'color':'#2563eb'},
        {'label':'Accepted','count':count_map.get('accepted',0),'color':'#7c3aed'},
        {'label':'Dispatched','count':count_map.get('dispatched',0),'color':'#0369a1'},
        {'label':'Delivered','count':count_map.get('delivered',0),'color':'#15803d'},
        {'label':'Cancelled','count':count_map.get('cancelled',0),'color':'#b91c1c'},
    ]
    return render_template('admin/admin_orders.html', orders=orders, stats=stats, current_status=current_status)


@app.route('/admin/api/product-search')
@admin_required
def admin_api_product_search():
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'products': []})
    db = get_db()
    like = f'%{q.lower()}%'
    rows = db.execute(
        "SELECT id, sku, name, mrp_price, retail_price, gst_percent, hsn_code, stock_total "
        "FROM products WHERE is_active=1 AND (LOWER(name) LIKE ? OR LOWER(sku) LIKE ?) "
        "ORDER BY name LIMIT 15",
        (like, like)
    ).fetchall()
    products = [{
        'id': r['id'], 'sku': r['sku'], 'name': r['name'],
        'mrp_price': float(r['mrp_price'] or 0), 'retail_price': float(r['retail_price'] or 0),
        'gst_percent': float(r['gst_percent'] or 0), 'hsn_code': r['hsn_code'] or '',
        'stock_total': r['stock_total'] or 0,
    } for r in rows]
    return jsonify({'products': products})


@app.route('/admin/api/user-search')
@admin_required
def admin_api_user_search():
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'users': []})
    db = get_db()
    like = f'%{q.lower()}%'
    rows = db.execute(
        "SELECT id, name, email FROM users WHERE LOWER(COALESCE(name,'')) LIKE ? OR LOWER(COALESCE(email,'')) LIKE ? "
        "ORDER BY name LIMIT 15",
        (like, like)
    ).fetchall()
    users = []
    for r in rows:
        addr = db.execute(
            'SELECT recipient_name, phone, address_line, city, state, pincode FROM user_addresses '
            'WHERE user_id=? ORDER BY is_default DESC, created_at DESC LIMIT 1', (r['id'],)
        ).fetchone()
        users.append({
            'id': r['id'], 'name': r['name'] or '', 'email': r['email'] or '',
            'address': {
                'name': addr['recipient_name'] or '', 'phone': addr['phone'] or '',
                'address_line': addr['address_line'] or '', 'city': addr['city'] or '',
                'state': addr['state'] or '', 'pincode': addr['pincode'] or ''
            } if addr else None
        })
    return jsonify({'users': users})


@app.route('/admin/orders/new')
@admin_required
def admin_new_order():
    enabled_couriers = get_configured_courier_names()
    return render_template('admin/admin_new_order.html', enabled_couriers=enabled_couriers)


@app.route('/admin/orders/create', methods=['POST'])
@admin_required
def admin_create_order():
    """Manual order entry for admin -- phone/in-person orders. Mirrors
    checkout_process()'s order_shipping insert exactly (same columns), but
    is driven by an explicit item list + customer payload instead of
    session['cart'], since there's no shopping-session context here."""
    db = get_db()
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []
    customer = data.get('customer') or {}
    payment_mode = (data.get('payment_mode') or 'COD').strip()
    if payment_mode not in ('COD', 'Prepaid'):
        payment_mode = 'COD'

    if not items:
        return jsonify({'status': 'error', 'message': 'Add at least one item to the order.'}), 400

    # ── Resolve items: existing products by id, or create a new product
    # record on the fly for inline "new item" entries. ──
    display_cart = []
    for item in items:
        if item.get('is_new'):
            name = (item.get('name') or '').strip()
            if not name:
                return jsonify({'status': 'error', 'message': 'New item is missing a name.'}), 400
            mrp_price = float(item.get('mrp_price') or 0)
            retail_price = float(item.get('retail_price') or 0)
            purchase_cost = float(item.get('purchase_cost') or 0)
            wholesale_price = float(item.get('wholesale_price') or 0)
            stock_total = int(item.get('stock_total') or 0)
            gst_percent = float(item.get('gst_percent') or 3)
            hsn_code = (item.get('hsn_code') or '').strip()

            sku = None
            for _ in range(5):
                candidate = 'QA' + datetime.now().strftime('%y%m%d%H%M%S%f')[:12]
                if not db.execute('SELECT id FROM products WHERE sku=?', (candidate,)).fetchone():
                    sku = candidate
                    break
            if not sku:
                return jsonify({'status': 'error', 'message': 'Could not generate a unique SKU, please try again.'}), 500

            slug = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')
            model_number = generate_model_number(db)

            db.execute(
                '''INSERT INTO products
                   (sku, name, category, retail_price, mrp_price, wholesale_price,
                    purchase_cost, stock_total, hsn_code, gst_percent, slug,
                    status, is_active, model_number, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,NOW())''',
                (sku, name, 'Uncategorized', retail_price, mrp_price, wholesale_price,
                 purchase_cost, stock_total, hsn_code, gst_percent, slug,
                 'published', 1, model_number)
            )
            db.commit()
            units = max(int(item.get('units') or 1), 1)
            display_cart.append({'sku': sku, 'name': name, 'price': retail_price, 'units': units, 'size': ''})
        else:
            prod = db.execute('SELECT sku, name, retail_price FROM products WHERE id=?', (item.get('id'),)).fetchone()
            if not prod:
                return jsonify({'status': 'error', 'message': 'One of the selected items could not be found.'}), 400
            units = max(int(item.get('units') or 1), 1)
            price = float(item.get('price')) if item.get('price') is not None else float(prod['retail_price'] or 0)
            display_cart.append({'sku': prod['sku'], 'name': prod['name'], 'price': price, 'units': units, 'size': ''})

    # ── Resolve the customer: existing user (with their saved address, or
    # address fields supplied inline if they have none saved), or a brand
    # new account created with the "narinakhre" default password. ──
    user_id = None
    if customer.get('existing_user_id'):
        user_row = db.execute('SELECT id, name, email FROM users WHERE id=?', (customer['existing_user_id'],)).fetchone()
        if not user_row:
            return jsonify({'status': 'error', 'message': 'Selected customer not found.'}), 400
        user_id = user_row['id']
        consignee_name = (customer.get('name') or user_row['name'] or '').strip()
        consignee_email = (customer.get('email') or user_row['email'] or '').strip()
        consignee_phone = (customer.get('phone') or '').strip()
        consignee_address = (customer.get('address') or '').strip()
        consignee_city = (customer.get('city') or '').strip()
        consignee_state = (customer.get('state') or '').strip()
        consignee_pincode = (customer.get('pincode') or '').strip()
    else:
        name = (customer.get('name') or '').strip()
        email = (customer.get('email') or '').strip().lower()
        phone = (customer.get('phone') or '').strip()
        address = (customer.get('address') or '').strip()
        city = (customer.get('city') or '').strip()
        state = (customer.get('state') or '').strip()
        pincode = (customer.get('pincode') or '').strip()
        if not (name and email and phone and address and pincode and city and state):
            return jsonify({'status': 'error',
                             'message': 'Please fill in all new-customer fields (name, email, mobile, address, pincode) and let the city/state resolve from the pincode.'}), 400

        existing_user = db.execute('SELECT id FROM users WHERE LOWER(email)=?', (email,)).fetchone()
        if existing_user:
            user_id = existing_user['id']
        else:
            password_hash = generate_password_hash('narinakhre')
            new_code = generate_referral_code(db)
            db.execute(
                'INSERT INTO users (name, email, password_hash, referral_code) VALUES (?,?,?,?)',
                (name, email, password_hash, new_code)
            )
            db.commit()
            new_user_row = db.execute('SELECT id FROM users WHERE LOWER(email)=?', (email,)).fetchone()
            user_id = new_user_row['id']
            notify_admin_new_user(db, name, email, 'Admin (manual order)')

            try:
                db.execute(
                    'INSERT INTO user_addresses (user_id, nickname, recipient_name, phone, email, address_line, city, state, pincode, is_default) '
                    'VALUES (?,?,?,?,?,?,?,?,?,1)',
                    (user_id, 'Home', name, phone, email, address, city, state, pincode)
                )
                db.commit()
            except Exception as e:
                app.logger.warning(f'Could not save address for manually created user {user_id}: {e}')

        consignee_name, consignee_email, consignee_phone = name, email, phone
        consignee_address, consignee_city, consignee_state, consignee_pincode = address, city, state, pincode

    if not (consignee_name and consignee_phone and consignee_address and consignee_city and consignee_state and consignee_pincode):
        return jsonify({'status': 'error',
                         'message': 'Missing shipping details for this customer -- address, city, state and pincode are all required.'}), 400

    def sanitize_for_delhivery(value):
        cleaned = value or ''
        for char in ['#', '&', '%', ';']:
            cleaned = cleaned.replace(char, ' ')
        return ' '.join(cleaned.split())

    cleaned_name = sanitize_for_delhivery(consignee_name)
    cleaned_address = sanitize_for_delhivery(consignee_address)

    # Admin explicitly picks the courier for manually-entered orders (phone/
    # in-person) rather than rate-shopping automatically, since the admin is
    # already hand-entering everything else. Falls back to Delhivery for any
    # value that isn't a known partner name.
    requested_courier = (data.get('courier_partner') or 'delhivery').strip().lower()
    chosen_courier = requested_courier if requested_courier in ('delhivery', 'shiprocket') else 'delhivery'

    internal_order_id = f"NN-SHP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{(consignee_phone or '0000')[-4:]}"
    subtotal_amount = sum(item['price'] * item['units'] for item in display_cart)
    gst_breakdown = calculate_inclusive_gst(display_cart, 0.0, subtotal_amount)
    total_amount = subtotal_amount
    cart_items_json = json.dumps(display_cart)

    courier_eta = None
    try:
        _name, chosen_provider = get_courier(chosen_courier)
        cart_weight = max(sum(item['units'] for item in display_cart) * 250, 250)
        eta_rates = chosen_provider.get_rates(
            app.config.get('WAREHOUSE_PIN', '482001'), consignee_pincode, cart_weight, mode=payment_mode
        )
        courier_eta = eta_rates.get('eta')
    except Exception as e:
        app.logger.warning(f'Manual order: courier ETA fetch failed: {e}')

    db.execute(
        '''INSERT INTO order_shipping
           (user_id, consignee_name, consignee_phone, consignee_email,
            consignee_address, consignee_city, consignee_state, consignee_pincode,
            internal_order_id, status, payment_mode,
            subtotal_amount, gst_amount, cgst_amount, sgst_amount,
            discount_amount, actual_shipping_cost, total_amount,
            coupon_code, cart_items_json, credits_redeemed,
            courier_partner, courier_eta)
           VALUES (?,?,?,?,?,?,?,?,?,'pending',?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (user_id, cleaned_name, consignee_phone, consignee_email,
         cleaned_address, consignee_city, consignee_state, consignee_pincode,
         internal_order_id, payment_mode,
         subtotal_amount, gst_breakdown['total_gst'], gst_breakdown['cgst'], gst_breakdown['sgst'],
         0.0, 0.0, total_amount,
         None, cart_items_json, 0.0,
         chosen_courier, courier_eta)
    )
    db.commit()

    log_admin_event(
        db, 'new_order', f'New order (manual): {internal_order_id}',
        detail=f'{cleaned_name} — ₹{total_amount:.0f} ({payment_mode})',
        related_id=internal_order_id
    )

    verify = db.execute('SELECT id FROM order_shipping WHERE internal_order_id=?', (internal_order_id,)).fetchone()
    if not verify:
        return jsonify({'status': 'error', 'message': 'Order could not be saved. Please try again.'}), 500

    # Confirmation emails -- customer + admin, same as checkout_process /
    # confirm_cod_order. Best-effort: an email failure must never make the
    # admin's manual order entry look like it failed.
    try:
        invoice_url = f"{request.url_root.rstrip('/')}/invoice/{internal_order_id}"
        items_for_email = [{
            'name': item.get('name', ''),
            'size': item.get('size', ''),
            'units': int(item.get('units', 1)),
            'price': float(item.get('price', 0)),
            'row_total': float(item.get('price', 0)) * int(item.get('units', 1)),
        } for item in display_cart]

        if consignee_email:
            order_html = render_template('retail/email_order_confirmation.html',
                customer_name=cleaned_name, order_id=internal_order_id,
                items=items_for_email, payment_mode=payment_mode, amount=total_amount,
                address_name=cleaned_name, address_line=cleaned_address,
                address_city=consignee_city, address_state=consignee_state,
                address_pincode=consignee_pincode,
                tracking_url=None, waybill=None, invoice_url=invoice_url,
                courier_partner=chosen_courier, courier_eta=courier_eta)
            order_text = (
                f"Hi {cleaned_name},\n\nYour order {internal_order_id} has been placed with us!\n\n"
                f"{'Amount to pay on delivery' if payment_mode == 'COD' else 'Amount due'}: ₹{total_amount:.2f}\n"
                f"Shipped via {chosen_courier.capitalize()}" + (f", estimated delivery: {courier_eta}\n" if courier_eta else "\n")
                + f"Tracking details will be shared once your order is dispatched.\n"
                f"Invoice: {invoice_url}\n\nThank you for shopping with Nari Nakhre!"
            )
            send_contact_email_async(consignee_email,
                f"Order Confirmed — {internal_order_id} | Nari Nakhre",
                order_text, html_body=order_html, from_email=ORDERS_FROM_EMAIL)

        item_lines = '\n'.join(
            f"  - {it['name']} x{it['units']} @ ₹{it['price']:.2f} = ₹{it['row_total']:.2f}"
            for it in items_for_email
        ) or '  (no item details)'
        admin_orders_url = f"{request.url_root.rstrip('/')}/admin/orders"
        admin_body = (
            f"New order entered manually by admin.\n\n"
            f"Order ID: {internal_order_id}\n"
            f"Customer: {cleaned_name}\n"
            f"Phone: {consignee_phone}\n"
            f"Email: {consignee_email or '-'}\n\n"
            f"Items:\n{item_lines}\n\n"
            f"Total ({payment_mode}): ₹{total_amount:.2f}\n\n"
            f"Shipping Address:\n{cleaned_address}, {consignee_city}, {consignee_state} - {consignee_pincode}\n\n"
            f"Courier: {chosen_courier.capitalize()}" + (f" (estimated delivery: {courier_eta})\n" if courier_eta else "\n")
            + f"Admin orders panel: {admin_orders_url}\n"
        )
        send_contact_email_async(ADMIN_EMAIL,
            f"🧾 New Manual Order — {internal_order_id}",
            admin_body, from_email=ORDERS_FROM_EMAIL)
    except Exception as e:
        app.logger.warning(f"Manual order email failed: {e}")

    return jsonify({
        'status': 'success',
        'internal_order_id': internal_order_id,
        'redirect_url': url_for('admin_orders')
    })


@app.route('/admin/orders/<int:order_id>/invoice')
@admin_required
def admin_order_invoice(order_id):
    conn = get_db()
    order = conn.execute('SELECT * FROM order_shipping WHERE id=?', (order_id,)).fetchone()
    if not order:
        flash('Order not found.', 'error')
        return redirect(url_for('admin_orders'))
    return render_template('admin/invoice.html', order=order,
                           seller_gst=DELHIVERY_SELLER_GST,
                           seller_name='Nari Nakhre',
                           seller_address=WAREHOUSE_ADDRESS)


@app.route('/admin/orders/<int:order_id>/accept', methods=['POST'])
@admin_required
def admin_order_accept(order_id):
    conn = get_db()
    order = conn.execute('SELECT * FROM order_shipping WHERE id=?', (order_id,)).fetchone()
    if not order:
        flash('Order not found.', 'error')
        return redirect(url_for('admin_orders'))
    waybill = order['delhivery_waybill']
    shipment_partner = order['courier_partner'] or 'delhivery'
    if not waybill:
        order_dict = dict(order)
        waybill, err, shipment_partner, shipment_id = create_courier_shipment(order_dict, [])
        if waybill:
            conn.execute('UPDATE order_shipping SET delhivery_waybill=?, courier_partner=?, shiprocket_shipment_id=? WHERE id=?', (waybill, shipment_partner, shipment_id, order_id))
        else:
            flash(f'Could not create {shipment_partner.capitalize()} shipment: {err}', 'error')
            return redirect(url_for('admin_orders'))
    pickup_scheduled = False
    pickup_id = None
    pickup_date = None
    if shipment_partner == 'delhivery':
        # Shiprocket's pickup scheduling isn't automated yet -- their
        # dashboard needs to be checked manually for now (see the flash
        # message below), same as when Delhivery scheduling itself fails.
        try:
            from datetime import date, timedelta
            pickup_date = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
            payload = {'pickup_time':'10:00:00','pickup_date':pickup_date,
                       'pickup_location':DELHIVERY_PICKUP_LOCATION,'expected_package_count':1}
            resp = requests.post('https://track.delhivery.com/fm/request/new/',
                json=payload, headers={'Authorization':f'Token {DELHIVERY_API_TOKEN}'}, timeout=15)
            pickup_scheduled = resp.status_code == 200
            if pickup_scheduled:
                try:
                    pr = resp.json()
                    pickup_id = str(pr.get('pickup_id') or pr.get('id') or '')
                    conn.execute('UPDATE order_shipping SET pickup_id=?, pickup_date=? WHERE id=?',
                                 (pickup_id, pickup_date, order_id))
                except Exception:
                    pass
            app.logger.info(f"Pickup for {waybill}: {resp.status_code} {resp.text[:200]}")
        except Exception as e:
            app.logger.warning(f"Pickup scheduling failed: {e}")
    conn.execute('UPDATE order_shipping SET status=? WHERE id=?', ('accepted', order_id))
    conn.commit()

    try:
        order_dict_for_email = dict(order)
        customer_email = order_dict_for_email.get('consignee_email', '')
        customer_name = order_dict_for_email.get('consignee_name', 'Customer')
        internal_order_id = order_dict_for_email.get('internal_order_id', '')
        if customer_email and waybill:
            tracking_url = f"{request.url_root.rstrip('/')}/track/{waybill}"
            courier_eta = order_dict_for_email.get('courier_eta')
            tracking_html = render_template('retail/email_tracking_update.html',
                customer_name=customer_name, order_id=internal_order_id,
                waybill=waybill, tracking_url=tracking_url,
                pickup_date=pickup_date if pickup_scheduled else None,
                courier_partner=shipment_partner, courier_eta=courier_eta)
            tracking_text = (
                f"Hi {customer_name},\n\nYour order {internal_order_id} has been accepted and is on its way!\n\n"
                f"Shipped via {shipment_partner.capitalize()}" + (f", estimated delivery: {courier_eta}\n" if courier_eta else "\n")
                + f"AWB / Tracking ID: {waybill}\nTrack: {tracking_url}\n\n"
                f"Thank you for shopping with Nari Nakhre!"
            )
            send_contact_email_async(customer_email,
                f"Your Order is On Its Way — {internal_order_id} | Nari Nakhre",
                tracking_text, html_body=tracking_html, from_email=ORDERS_FROM_EMAIL)
    except Exception as e:
        app.logger.warning(f"Order-accept tracking email failed: {e}")

    msg = f"Order accepted via {shipment_partner.capitalize()}. Waybill: {waybill}."
    if pickup_scheduled:
        msg += " Pickup scheduled for tomorrow."
    else:
        msg += f" Note: Schedule pickup manually in the {shipment_partner.capitalize()} panel."
    flash(msg, 'success')
    return redirect(url_for('admin_orders'))


@app.route('/admin/orders/<int:order_id>/dispatched', methods=['POST'])
@admin_required
def admin_order_dispatched(order_id):
    conn = get_db()
    conn.execute('UPDATE order_shipping SET status=? WHERE id=?', ('dispatched', order_id))
    conn.commit()
    flash('Order marked as dispatched.', 'success')
    return redirect(url_for('admin_orders'))


@app.route('/admin/orders/<int:order_id>/cancel', methods=['POST'])
@admin_required
def admin_order_cancel(order_id):
    conn = get_db()
    order = conn.execute('SELECT * FROM order_shipping WHERE id=?', (order_id,)).fetchone()
    if not order:
        flash('Order not found.', 'error')
        return redirect(url_for('admin_orders'))
    waybill = order['delhivery_waybill']
    courier_partner = order['courier_partner'] or 'delhivery'
    manual_cancel_note = ''
    if waybill and courier_partner == 'delhivery' and DELHIVERY_API_TOKEN:
        try:
            requests.post('https://track.delhivery.com/api/p/edit',
                json={'waybill':waybill,'cancellation':True},
                headers={'Authorization':f'Token {DELHIVERY_API_TOKEN}'}, timeout=15)
        except Exception as e:
            app.logger.warning(f"Delhivery cancellation failed: {e}")
    elif waybill and courier_partner == 'shiprocket':
        # Shiprocket cancellation needs their internal order id, which isn't
        # tracked yet (only the AWB is stored) -- cancel manually in their
        # dashboard for now rather than risk calling the wrong courier's API.
        manual_cancel_note = ' Note: cancel this shipment manually in the Shiprocket panel.'
    conn.execute('UPDATE order_shipping SET status=? WHERE id=?', ('cancelled', order_id))
    conn.commit()
    flash(f"Order {order['internal_order_id']} cancelled.{manual_cancel_note}", 'success')
    return redirect(url_for('admin_orders'))


@app.route('/admin/orders/<int:order_id>/label')
@admin_required
def admin_shipping_label(order_id):
    """
    Print Label -- courier-specific. Shiprocket hands back a ready-made PDF
    (its own official label), so that's opened directly. Delhivery's API
    only returns JSON meant to be rendered client-side, so that data (when
    fetchable) is passed into our own printable template alongside a real
    Code128 barcode of the waybill. Either courier falls back to rendering
    from locally stored order data if the live fetch fails, so Print Label
    still works even when the courier API is briefly unavailable.
    """
    conn = get_db()
    order = conn.execute('SELECT * FROM order_shipping WHERE id=?', (order_id,)).fetchone()
    if not order or not order['delhivery_waybill']:
        flash('No waybill found for this order.', 'error')
        return redirect(url_for('admin_orders'))

    courier_partner = order['courier_partner'] or 'delhivery'
    _, provider = get_courier(courier_partner)

    if courier_partner == 'shiprocket':
        result = provider.get_label(order['shiprocket_shipment_id'])
        if result.get('status') and result.get('label_url'):
            return redirect(result['label_url'])
        flash(f"Could not fetch Shiprocket's official label ({result.get('msg')}) -- showing a basic label instead.", 'error')
        return render_template('admin/shipping_label.html', order=order, delhivery_slip=None)

    slip_result = provider.get_packing_slip(order['delhivery_waybill'])
    delhivery_slip = slip_result.get('data') if slip_result.get('status') else None
    return render_template('admin/shipping_label.html', order=order, delhivery_slip=delhivery_slip)


@app.route('/admin/coupons', methods=['GET'])
@admin_required
def admin_coupons():
    db = get_db()
    coupons = db.execute('SELECT * FROM coupons ORDER BY id DESC').fetchall()
    categories = db.execute(
        "SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category != '' ORDER BY category"
    ).fetchall()
    sub_categories = db.execute(
        "SELECT DISTINCT sub_category FROM products WHERE sub_category IS NOT NULL AND sub_category != '' ORDER BY sub_category"
    ).fetchall()
    today_str = datetime.now().strftime('%Y-%m-%d')
    return render_template('admin/admin_coupons.html',
                            coupons=coupons, categories=categories, sub_categories=sub_categories,
                            today_str=today_str)


@app.route('/admin/coupons/create', methods=['POST'])
@admin_required
def admin_coupon_create():
    db = get_db()
    code = (request.form.get('code') or '').strip().upper()
    discount_percent = request.form.get('discount_percent', type=float) or 0.0
    min_order_amount = request.form.get('min_order_amount', type=float) or 0.0
    category = (request.form.get('category') or '').strip()
    sub_category = (request.form.get('sub_category') or '').strip()
    expiry_date = (request.form.get('expiry_date') or '').strip() or None
    usage_limit = request.form.get('usage_limit', type=int) or 0
    max_discount_amount = request.form.get('max_discount_amount', type=float) or None
    is_public = 1 if (request.form.get('visibility') or 'public') == 'public' else 0

    if not code:
        flash('Coupon code is required.')
        return redirect(url_for('admin_coupons'))
    if discount_percent <= 0 or discount_percent > 100:
        flash('Discount percent must be between 1 and 100.')
        return redirect(url_for('admin_coupons'))

    try:
        db.execute(
            "INSERT INTO coupons (code, discount_percent, min_order_amount, category, sub_category, expiry_date, usage_limit, max_discount_amount, is_public, is_active) VALUES (?,?,?,?,?,?,?,?,?,1)",
            (code, discount_percent, min_order_amount, category or None, sub_category or None, expiry_date, usage_limit, max_discount_amount, is_public)
        )
        db.commit()
        flash('Coupon "' + code + '" created successfully.')
    except Exception as e:
        flash('Could not create coupon - code may already exist. (' + str(e) + ')')
    return redirect(url_for('admin_coupons'))


@app.route('/admin/coupons/<int:coupon_id>/toggle', methods=['POST'])
@admin_required
def admin_coupon_toggle(coupon_id):
    db = get_db()
    row = db.execute('SELECT is_active FROM coupons WHERE id=?', (coupon_id,)).fetchone()
    if row is None:
        flash('Coupon not found.')
        return redirect(url_for('admin_coupons'))
    new_status = 0 if row['is_active'] else 1
    db.execute('UPDATE coupons SET is_active=? WHERE id=?', (new_status, coupon_id))
    db.commit()
    flash('Coupon status updated.')
    return redirect(url_for('admin_coupons'))


@app.route('/admin/coupons/<int:coupon_id>/delete', methods=['POST'])
@admin_required
def admin_coupon_delete(coupon_id):
    db = get_db()
    db.execute('DELETE FROM coupons WHERE id=?', (coupon_id,))
    db.commit()
    flash('Coupon deleted.')
    return redirect(url_for('admin_coupons'))


@app.route('/admin/coupons/<int:coupon_id>/edit', methods=['POST'])
@admin_required
def admin_coupon_edit(coupon_id):
    db = get_db()
    discount_percent = request.form.get('discount_percent', type=float) or 0.0
    min_order_amount = request.form.get('min_order_amount', type=float) or 0.0
    category = (request.form.get('category') or '').strip()
    sub_category = (request.form.get('sub_category') or '').strip()
    expiry_date = (request.form.get('expiry_date') or '').strip() or None
    usage_limit = request.form.get('usage_limit', type=int) or 0
    max_discount_amount = request.form.get('max_discount_amount', type=float) or None
    is_public = 1 if (request.form.get('visibility') or 'public') == 'public' else 0

    if discount_percent <= 0 or discount_percent > 100:
        flash('Discount percent must be between 1 and 100.')
        return redirect(url_for('admin_coupons'))

    db.execute(
        "UPDATE coupons SET discount_percent=?, min_order_amount=?, category=?, sub_category=?, expiry_date=?, usage_limit=?, max_discount_amount=?, is_public=? WHERE id=?",
        (discount_percent, min_order_amount, category or None, sub_category or None, expiry_date, usage_limit, max_discount_amount, is_public, coupon_id)
    )
    db.commit()
    flash('Coupon updated.')
    return redirect(url_for('admin_coupons'))


@app.route('/admin/delivery-partners', methods=['GET'])
@admin_required
def admin_delivery_partners():
    db = get_db()
    partners = db.execute('''
        SELECT dp.id, dp.name, dp.is_enabled, dp.updated_at AS partner_updated_at,
               dpc.environment, dpc.updated_at AS credentials_updated_at,
               CASE WHEN dpc.encrypted_credentials IS NOT NULL AND dpc.encrypted_credentials != ''
                    THEN 1 ELSE 0 END AS is_configured
        FROM delivery_partners dp
        LEFT JOIN delivery_partner_credentials dpc ON dpc.partner_id = dp.id
        ORDER BY dp.name
    ''').fetchall()
    return render_template('admin/admin_delivery_partners.html', partners=partners)


@app.route('/admin/delivery-partners/<int:partner_id>/save', methods=['POST'])
@admin_required
def admin_delivery_partner_save(partner_id):
    db = get_db()
    partner = db.execute('SELECT * FROM delivery_partners WHERE id=?', (partner_id,)).fetchone()
    if not partner:
        flash('Delivery partner not found.')
        return redirect(url_for('admin_delivery_partners'))

    name = partner['name']
    environment = 'production'

    if name == 'shiprocket':
        email = (request.form.get('email') or '').strip()
        password = (request.form.get('password') or '').strip()
        if not email or not password:
            flash('Shiprocket email and password are both required.')
            return redirect(url_for('admin_delivery_partners'))
        credentials = {'email': email, 'password': password}
    elif name == 'delhivery':
        api_token = (request.form.get('api_token') or '').strip()
        environment = (request.form.get('environment') or 'production').strip()
        if environment not in ('staging', 'production'):
            environment = 'production'
        if not api_token:
            flash('Delhivery API token is required.')
            return redirect(url_for('admin_delivery_partners'))
        credentials = {'api_token': api_token}
    else:
        flash('Unknown delivery partner.')
        return redirect(url_for('admin_delivery_partners'))

    try:
        encrypted = encrypt_credentials(credentials)
    except RuntimeError as e:
        flash(f'Could not save credentials: {e}')
        return redirect(url_for('admin_delivery_partners'))

    db.execute(
        '''INSERT INTO delivery_partner_credentials (partner_id, environment, encrypted_credentials, updated_at)
           VALUES (?, ?, ?, NOW())
           ON CONFLICT (partner_id) DO UPDATE SET
               environment = EXCLUDED.environment,
               encrypted_credentials = EXCLUDED.encrypted_credentials,
               updated_at = NOW()''',
        (partner_id, environment, encrypted)
    )
    db.commit()
    flash(f'{name.capitalize()} credentials saved.')
    return redirect(url_for('admin_delivery_partners'))


@app.route('/admin/delivery-partners/<int:partner_id>/toggle', methods=['POST'])
@admin_required
def admin_delivery_partner_toggle(partner_id):
    db = get_db()
    partner = db.execute('SELECT * FROM delivery_partners WHERE id=?', (partner_id,)).fetchone()
    if not partner:
        flash('Delivery partner not found.')
        return redirect(url_for('admin_delivery_partners'))
    new_status = 0 if partner['is_enabled'] else 1

    if new_status == 1:
        creds = db.execute(
            'SELECT encrypted_credentials FROM delivery_partner_credentials WHERE partner_id=?',
            (partner_id,)
        ).fetchone()
        if not creds or not creds['encrypted_credentials']:
            flash(f"Can't enable {partner['name'].capitalize()} — add credentials first.")
            return redirect(url_for('admin_delivery_partners'))

    db.execute('UPDATE delivery_partners SET is_enabled=?, updated_at=NOW() WHERE id=?', (new_status, partner_id))
    db.commit()
    flash('Delivery partner status updated.')
    return redirect(url_for('admin_delivery_partners'))


@app.route('/admin/stocks/sync', methods=['POST'])
@stocks_login_required
def admin_stocks_sync():
    """Manual trigger for Nari Nakhre Stocks Phase 1 ingestion -- pulls the
    latest daily candle (or backfills) for every active stock_watchlist row.
    See utils/stock_ingestion.py. Now gated by the Stocks login (any active
    account) rather than the storefront's admin_required, since it needs the
    Kite session that login owns."""
    db = get_db()
    try:
        access_token = get_kite_access_token(db)
        if not access_token:
            return jsonify({
                'status': 'error',
                'message': 'No Kite session yet -- a super_admin must log in via /admin/stocks/kite/login first.'
            }), 400
        kite_client = KiteClient(db=db, access_token=access_token)
        summary = sync_daily_data(db, kite_client=kite_client)
    except Exception as e:
        app.logger.error(f'Stock sync failed: {e}')
        return jsonify({'status': 'error', 'message': str(e)}), 500
    return jsonify({'status': 'ok', **summary})


@app.route('/admin/stocks/kite/login', methods=['GET'])
@stocks_role_required('super_admin')
def stocks_kite_login():
    """Sends the super_admin to Zerodha's login page. Kite redirects back to
    stocks_kite_callback below with a request_token once they log in there."""
    try:
        login_url = get_kite_login_url()
    except RuntimeError as e:
        flash(str(e), 'error')
        return redirect(url_for('stocks_admin_dashboard'))
    return redirect(login_url)


@app.route('/admin/stocks/kite/callback', methods=['GET'])
def stocks_kite_callback():
    """Registered as the Redirect URL on the Kite Connect app. Deliberately
    public/unauthenticated -- Zerodha's redirect is a fresh top-level GET
    from kite.zerodha.com, and this is the standard shape for an OAuth-style
    callback (the request_token itself, single-use and only exchangeable
    with our api_secret, is what proves the login happened -- not our own
    session cookie, which may or may not have survived the round trip).
    Exchanges the request_token for an access_token and stores it -- that's
    the token every subsequent Kite API call uses until it expires tomorrow
    and a super_admin repeats this flow."""
    request_token = request.args.get('request_token')
    status = request.args.get('status')
    if status != 'success' or not request_token:
        flash('Kite login was not completed successfully.', 'error')
        return redirect(url_for('stocks_admin_dashboard'))

    try:
        access_token = exchange_request_token(request_token)
    except Exception as e:
        app.logger.error(f'Kite session exchange failed: {e}')
        flash('Could not complete Kite login. Please try again.', 'error')
        return redirect(url_for('stocks_admin_dashboard'))

    db = get_db()
    expires_at = save_kite_access_token(db, access_token, session.get('stocks_admin_id'))
    expires_at_ist = expires_at.astimezone(IST).strftime('%d %b %Y, %I:%M %p IST')
    flash(f'Kite access token refreshed. Expires {expires_at_ist}.')
    return redirect(url_for('stocks_admin_dashboard'))


@app.route('/admin/stocks/kite/postback', methods=['POST'])
def stocks_kite_postback():
    """Registered as the Postback URL on the Kite Connect app. Public --
    Zerodha posts here directly, no browser session involved. Only verifies
    the checksum and logs the payload for now; nothing here updates order or
    suggestion state yet (that depends on execute_suggestion(), a later
    phase)."""
    payload = request.get_json(silent=True)
    if payload is None:
        payload = request.form.to_dict()

    if not verify_postback_checksum(payload):
        app.logger.warning('Kite postback rejected: invalid or missing checksum')
        return jsonify({'status': 'error', 'message': 'Invalid checksum'}), 400

    db = get_db()
    try:
        log_postback(db, payload)
    except Exception as e:
        app.logger.error(f'Kite postback log failed: {e}')
        return jsonify({'status': 'error'}), 500
    return jsonify({'status': 'ok'})


@app.route('/admin/stocks/login', methods=['GET', 'POST'])
def stocks_admin_login():
    """Separate login for Nari Nakhre Stocks -- shared by super_admin and
    child admins (session['stocks_admin_role'] tells them apart). Nothing to
    do with the storefront's /admin/login or session['is_admin']."""
    if request.method == 'GET':
        return render_template('admin/stocks_login.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

    if not verify_recaptcha(request.form.get('recaptcha_token'), remote_ip=request.remote_addr, expected_action='stocks_admin_login'):
        app.logger.warning('Bot caught on stocks admin login (recaptcha)')
        flash('Please try again.', 'error')
        return render_template('admin/stocks_login.html', recaptcha_site_key=RECAPTCHA_SITE_KEY), 401

    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''

    db = get_db()
    admin_row = authenticate_stocks_admin(db, username, password)
    if not admin_row:
        flash('Invalid username or password.', 'error')
        return render_template('admin/stocks_login.html', recaptcha_site_key=RECAPTCHA_SITE_KEY), 401

    session['stocks_admin_id'] = admin_row['id']
    session['stocks_admin_username'] = admin_row['username']
    session['stocks_admin_role'] = admin_row['role']
    session.modified = True
    return redirect(url_for('stocks_admin_dashboard'))


@app.route('/admin/stocks/logout', methods=['GET'])
@stocks_login_required
def stocks_admin_logout():
    session.pop('stocks_admin_id', None)
    session.pop('stocks_admin_username', None)
    session.pop('stocks_admin_role', None)
    session.modified = True
    return redirect(url_for('stocks_admin_login'))


@app.route('/admin/stocks/dashboard', methods=['GET'])
@stocks_login_required
def stocks_admin_dashboard():
    db = get_db()
    kite_status = get_kite_session_status(db)
    return render_template(
        'admin/stocks_dashboard.html',
        username=session.get('stocks_admin_username'),
        role=session.get('stocks_admin_role'),
        kite_status=kite_status,
    )


@app.route('/admin/stocks/admins', methods=['GET'])
@stocks_role_required('super_admin')
def stocks_admin_manage():
    db = get_db()
    admins = list_stocks_admin_users(db)
    return render_template('admin/stocks_admins.html', admins=admins)


@app.route('/admin/stocks/admins/create', methods=['POST'])
@stocks_role_required('super_admin')
def stocks_admin_create():
    db = get_db()
    username = request.form.get('username')
    password = request.form.get('password')
    _row, error = create_child_admin(db, username, password, session.get('stocks_admin_id'))
    if error:
        flash(error, 'error')
    else:
        flash(f'Child admin "{username.strip()}" created.')
    return redirect(url_for('stocks_admin_manage'))


@app.route('/admin/stocks/admins/<int:admin_id>/toggle', methods=['POST'])
@stocks_role_required('super_admin')
def stocks_admin_toggle(admin_id):
    db = get_db()
    if not toggle_child_admin_active(db, admin_id):
        flash('Could not update that account.', 'error')
    else:
        flash('Account status updated.')
    return redirect(url_for('stocks_admin_manage'))


def _send_welcome_backfill_batch(user_rows):
    """Runs in a background thread: sends the welcome email + private coupon
    to a batch of existing users, one at a time with a short delay between
    sends so Zeptomail doesn't see a burst. Each user is marked
    welcome_email_sent_at as soon as their email is queued, so a re-run of
    the backfill (e.g. after a crash) only picks up whoever's left.

    Needs its own app context -- a background thread has no Flask request
    context, and both get_db() (uses flask.g) and render_template() (used
    inside send_welcome_email) require an active app context to work."""
    with app.app_context():
        db = get_db()
        sent, failed = 0, 0
        for u in user_rows:
            try:
                if send_welcome_email(db, u['id'], u['name'], u['email'], async_send=False):
                    sent += 1
                else:
                    failed += 1
                    app.logger.warning(f"Welcome backfill: Zeptomail rejected send to {u['email']}")
            except Exception as e:
                failed += 1
                app.logger.warning(f"Welcome backfill failed for user {u['id']} ({u['email']}): {e}")
            _time.sleep(1)
        app.logger.info(f'Welcome email backfill finished: {sent} sent, {failed} failed.')


@app.route('/admin/users/send-welcome-backfill', methods=['POST'])
@admin_required
def admin_send_welcome_backfill():
    """One-time action: sends the welcome email + private 15% coupon to every
    existing registered user who hasn't already received one. Safe to click
    more than once -- already-sent users are skipped via welcome_email_sent_at."""
    db = get_db()
    users = db.execute(
        "SELECT id, name, email FROM users"
        " WHERE email IS NOT NULL AND email != '' AND welcome_email_sent_at IS NULL"
    ).fetchall()
    if not users:
        flash('No users left to email — everyone with an account already has a welcome coupon.')
        return redirect(url_for('admin_dashboard'))

    t = threading.Thread(target=_send_welcome_backfill_batch, args=(list(users),), daemon=True)
    t.start()
    flash(f'Sending welcome emails to {len(users)} existing user(s) in the background — this will take a few minutes.')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/email-diagnostics', methods=['GET', 'POST'])
@admin_required
def admin_email_diagnostics():
    """Shows which Zeptomail env vars are actually set on this server
    (masked, never the real secret) and which one send_contact_email()
    would pick for each sender identity -- and lets you fire a real test
    send and see Zeptomail's raw response right here, since Render log
    access isn't always available."""
    def mask(val):
        if not val:
            return None
        if len(val) <= 6:
            return '•' * len(val)
        return val[:3] + '•' * max(len(val) - 6, 3) + val[-3:]

    env_status = {
        'SMTP_SUPPORT_EMAIL_PASSWORD': mask(os.environ.get('SMTP_SUPPORT_EMAIL_PASSWORD')),
        'SMTP_ORDERS_ZEPTO_PASSWORD': mask(os.environ.get('SMTP_ORDERS_ZEPTO_PASSWORD')),
        'SMTP_support_EMAIL_FROM': os.environ.get('SMTP_support_EMAIL_FROM'),
        'SMTP_ORDERS_FROM_EMAIL': os.environ.get('SMTP_ORDERS_FROM_EMAIL'),
        'ZEPTOMAIL_API_URL': os.environ.get('ZEPTOMAIL_API_URL') or '(unset — defaults to https://api.zeptomail.in/v1.1/email)',
    }

    support_key_source = 'SMTP_SUPPORT_EMAIL_PASSWORD' if os.environ.get('SMTP_SUPPORT_EMAIL_PASSWORD') else 'NONE SET — send will fail'
    orders_key_source = 'SMTP_ORDERS_ZEPTO_PASSWORD' if os.environ.get('SMTP_ORDERS_ZEPTO_PASSWORD') else 'NONE SET — send will fail'

    test_result = None
    if request.method == 'POST':
        test_to = (request.form.get('test_email') or '').strip()
        identity = request.form.get('identity') or 'support'
        if not test_to:
            flash('Enter a test recipient email address.')
        else:
            from_email = ORDERS_FROM_EMAIL if identity == 'orders' else SUPPORT_FROM_EMAIL
            api_url = os.environ.get('ZEPTOMAIL_API_URL', 'https://api.zeptomail.in/v1.1/email')
            api_key = (os.environ.get('SMTP_ORDERS_ZEPTO_PASSWORD') if from_email == ORDERS_FROM_EMAIL
                       else os.environ.get('SMTP_SUPPORT_EMAIL_PASSWORD', ''))

            if not api_key:
                test_result = {'ok': False, 'from_email': from_email,
                                'detail': f'No API token available for sender {from_email} -- '
                                          f'the matching env var is empty or unset.'}
            else:
                payload = {
                    'from': {'address': from_email, 'name': 'Nari Nakhre'},
                    'to': [{'email_address': {'address': test_to, 'name': test_to}}],
                    'subject': 'Nari Nakhre — test email from admin diagnostics',
                    'textbody': f'This is a test send from the "{identity}" identity ({from_email}), '
                                f'triggered from the admin Email Diagnostics page.',
                }
                headers = {
                    'Authorization': f'Zoho-enczapikey {api_key}',
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                }
                try:
                    resp = requests.post(api_url, json=payload, headers=headers, timeout=10)
                    test_result = {
                        'ok': resp.status_code in (200, 201),
                        'status_code': resp.status_code,
                        'detail': resp.text[:1500],
                        'from_email': from_email,
                    }
                except requests.exceptions.Timeout:
                    test_result = {'ok': False, 'from_email': from_email, 'detail': 'Request to Zeptomail timed out after 10s.'}
                except Exception as e:
                    test_result = {'ok': False, 'from_email': from_email, 'detail': f'{type(e).__name__}: {e}'}

    return render_template('admin/admin_email_diagnostics.html',
                            env_status=env_status, support_from=SUPPORT_FROM_EMAIL, orders_from=ORDERS_FROM_EMAIL,
                            support_key_source=support_key_source, orders_key_source=orders_key_source,
                            test_result=test_result)


# ── Email Campaigns ──────────────────────────────────────────────────────

def _campaign_recipient_group(db, group):
    """Resolve a recipient-group name to a list of {id, name, email} dicts."""
    if group == 'new':
        return db.execute(
            "SELECT id, name, email FROM users WHERE email IS NOT NULL AND email != ''"
            " AND id NOT IN (SELECT DISTINCT user_id FROM order_shipping"
            "                WHERE user_id IS NOT NULL AND status != 'cancelled')"
        ).fetchall()
    if group == 'ordered':
        return db.execute(
            "SELECT id, name, email FROM users WHERE email IS NOT NULL AND email != ''"
            " AND id IN (SELECT DISTINCT user_id FROM order_shipping"
            "            WHERE user_id IS NOT NULL AND status != 'cancelled')"
        ).fetchall()
    return db.execute(
        "SELECT id, name, email FROM users WHERE email IS NOT NULL AND email != ''"
    ).fetchall()


def _run_campaign_batch(db, campaign, campaign_id, user_rows):
    """Does the actual per-recipient work for a campaign send: personal
    coupon, personalized email, synchronous send, and bookkeeping. Raises on
    any setup-level failure so the caller can mark the campaign as failed
    instead of leaving it silently stuck on "sending"."""
    try:
        product_ids = json.loads(campaign.get('product_ids') or '[]')
    except Exception:
        product_ids = []
    products = []
    for pid in product_ids:
        p = db.execute(
            'SELECT id, sku, name, mrp_price, retail_price, price1, image_field'
            ' FROM products WHERE id=?', (pid,)
        ).fetchone()
        if p:
            p = dict(p)
            imgs = get_product_images(p)
            p['image'] = imgs[0] if imgs else ''
            products.append(p)

    discount_percent = campaign['discount_percent']
    max_discount_amount = campaign['max_discount_amount']
    min_order_amount = campaign.get('min_order_amount') or 0
    sent = 0
    failed = 0
    for u in user_rows:
        recipient_id = None
        try:
            # A campaign that previously failed partway through can be
            # retried by sending it again -- skip anyone who already has a
            # successful send recorded so they don't get a second coupon
            # and a duplicate email.
            already = db.execute(
                "SELECT id FROM email_campaign_recipients WHERE campaign_id=? AND user_id=? AND status != 'failed'",
                (campaign_id, u['id'])
            ).fetchone()
            if already:
                continue

            first_name = (u.get('name') or 'there').strip().split(' ')[0] or 'there'
            code = generate_personal_coupon_code(db, first_name, discount_percent)
            expiry_date = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
            db.execute(
                "INSERT INTO coupons (code, discount_percent, min_order_amount, max_discount_amount,"
                " expiry_date, usage_limit, is_public, is_active, user_id, campaign_id)"
                " VALUES (?,?,?,?,?,?,0,1,?,?)",
                (code, discount_percent, min_order_amount, max_discount_amount, expiry_date, 1, u['id'], campaign_id)
            )
            db.commit()

            # Insert the recipient row before rendering so the email's
            # links can carry this row's id for click tracking. Marked
            # failed below if the send turns out to have failed.
            db.execute(
                "INSERT INTO email_campaign_recipients (campaign_id, user_id, email, name, coupon_code)"
                " VALUES (?,?,?,?,?)",
                (campaign_id, u['id'], u['email'], u.get('name'), code)
            )
            db.commit()
            recipient_row = db.execute(
                "SELECT id FROM email_campaign_recipients WHERE campaign_id=? AND coupon_code=?",
                (campaign_id, code)
            ).fetchone()
            recipient_id = recipient_row['id'] if recipient_row else None

            coupon = {'code': code, 'discount_percent': discount_percent,
                      'max_discount_amount': max_discount_amount,
                      'min_order_amount': min_order_amount, 'expiry_date': expiry_date}
            html = render_template('retail/email_campaign.html', campaign_name=campaign['name'],
                                    first_name=first_name, coupon=coupon, products=products,
                                    recipient_id=recipient_id)
            text = (
                f"Hi {first_name},\n\nA special offer just for you: use code {code} for "
                f"{int(discount_percent)}% off (up to Rs.{int(max_discount_amount)}) "
                f"on orders above Rs.{int(min_order_amount)}.\n"
                f"Valid until {expiry_date}.\n\nShop now: https://narinakhre.com\n"
            )
            # Send synchronously: this function already runs in its own
            # background thread (started from admin_campaign_send), so
            # there's no HTTP response to avoid blocking here. Calling the
            # async fire-and-forget wrapper from an already-backgrounded
            # thread meant we recorded every recipient as "sent" without
            # ever checking whether Zeptomail actually accepted the email --
            # a rejected send just vanished into an unwatched log line.
            ok, detail = send_contact_email(
                u['email'], f"{campaign['name']} — {int(discount_percent)}% off just for you \U0001F381",
                text, html_body=html, capture_detail=True
            )
            if ok:
                sent += 1
            else:
                failed += 1
                if recipient_id:
                    db.execute(
                        "UPDATE email_campaign_recipients SET status='failed', error_detail=? WHERE id=?",
                        (detail, recipient_id)
                    )
                    db.commit()
                app.logger.warning(f"Campaign {campaign_id}: Zeptomail rejected send to {u['email']}: {detail}")
        except Exception as e:
            failed += 1
            if recipient_id:
                db.execute(
                    "UPDATE email_campaign_recipients SET status='failed', error_detail=? WHERE id=?",
                    (f'{type(e).__name__}: {e}', recipient_id)
                )
                db.commit()
            app.logger.warning(f"Campaign {campaign_id} send failed for user {u.get('id')}: {e}")
        _time.sleep(1)

    # Count actual successful rows rather than this run's local `sent` tally,
    # so a retry after a partial failure reports the true total across both
    # attempts, not just what changed in this run.
    total_ok = db.execute(
        "SELECT COUNT(*) as cnt FROM email_campaign_recipients WHERE campaign_id=? AND status != 'failed'",
        (campaign_id,)
    ).fetchone()
    db.execute(
        "UPDATE email_campaigns SET status='sent', sent_at=CURRENT_TIMESTAMP, recipient_count=? WHERE id=?",
        (total_ok['cnt'] if total_ok else sent, campaign_id)
    )
    db.commit()
    app.logger.info(f'Campaign {campaign_id} finished: {sent} newly sent, {failed} failed.')


def _send_campaign_batch(campaign_id, user_rows):
    """Runs in a background thread: generates a personal single-use coupon
    and sends the campaign email to each recipient, one at a time. Needs its
    own app context, same reasoning as _send_welcome_backfill_batch."""
    with app.app_context():
        db = get_db()
        campaign = db.execute('SELECT * FROM email_campaigns WHERE id=?', (campaign_id,)).fetchone()
        if not campaign:
            return

        try:
            _run_campaign_batch(db, campaign, campaign_id, user_rows)
        except Exception as e:
            # Any crash in _run_campaign_batch's setup (e.g. a bad products
            # query) used to kill this thread silently, leaving the campaign
            # stuck on "sending" forever with zero recorded recipients and
            # no visible error. Surface it as a real, visible failure state.
            app.logger.error(f"Campaign {campaign_id} aborted unexpectedly: {type(e).__name__}: {e}")
            already_sent = db.execute(
                "SELECT COUNT(*) as cnt FROM email_campaign_recipients WHERE campaign_id=? AND status != 'failed'",
                (campaign_id,)
            ).fetchone()
            db.execute(
                "UPDATE email_campaigns SET status='failed', sent_at=CURRENT_TIMESTAMP, recipient_count=? WHERE id=?",
                (already_sent['cnt'] if already_sent else 0, campaign_id)
            )
            db.commit()


@app.route('/c/<int:recipient_id>')
def campaign_click(recipient_id):
    """Click-tracking redirect used by links inside campaign emails -- marks
    this recipient as having visited the site after the email (first click
    only), then forwards them on to the real destination. `to` must be a
    site-relative path (never a full URL) so this can't be abused as an
    open redirect."""
    db = get_db()
    row = db.execute(
        'SELECT id, clicked_at FROM email_campaign_recipients WHERE id=?', (recipient_id,)
    ).fetchone()
    if row and not row.get('clicked_at'):
        db.execute(
            'UPDATE email_campaign_recipients SET clicked_at=CURRENT_TIMESTAMP WHERE id=?', (recipient_id,)
        )
        db.commit()

    dest = request.args.get('to') or '/'
    if not dest.startswith('/') or dest.startswith('//'):
        dest = '/'
    return redirect(dest)


@app.route('/admin/campaigns', methods=['GET'])
@admin_required
def admin_campaigns():
    db = get_db()
    campaigns = db.execute('SELECT * FROM email_campaigns ORDER BY id DESC').fetchall()
    return render_template('admin/admin_campaigns.html', campaigns=campaigns)


@app.route('/admin/campaigns/preview', methods=['POST'])
@admin_required
def admin_campaign_preview():
    """AJAX: renders a sample of the campaign email (placeholder name/code,
    a fresh random pick of trending products) without saving anything --
    used for both the initial "Generate Email" and "Regenerate"."""
    db = get_db()
    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    discount_percent = float(data.get('discount_percent') or 0)
    max_discount_amount = float(data.get('max_discount_amount') or 0)
    min_order_amount = float(data.get('min_order_amount') or 0)

    if discount_percent <= 0 or discount_percent > 100:
        return jsonify({'status': 'error', 'message': 'Discount percent must be between 1 and 100.'}), 400

    products = get_trending_products(db, limit=4)
    if not products:
        return jsonify({'status': 'error', 'message': 'No trending products available to feature yet.'}), 400

    expiry_date = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    sample_coupon = {
        'code': 'NNSAMPLE' + str(int(discount_percent)),
        'discount_percent': discount_percent,
        'max_discount_amount': max_discount_amount,
        'min_order_amount': min_order_amount,
        'expiry_date': expiry_date,
    }
    html = render_template('retail/email_campaign.html', campaign_name=name,
                            first_name='Customer', coupon=sample_coupon, products=products)
    return jsonify({'status': 'success', 'html': html, 'product_ids': [p['id'] for p in products]})


@app.route('/admin/campaigns/save', methods=['POST'])
@admin_required
def admin_campaign_save():
    db = get_db()
    name = (request.form.get('name') or '').strip()
    discount_percent = request.form.get('discount_percent', type=float) or 0.0
    max_discount_amount = request.form.get('max_discount_amount', type=float) or 0.0
    min_order_amount = request.form.get('min_order_amount', type=float) or 0.0
    product_ids_raw = request.form.get('product_ids') or '[]'

    if not name:
        flash('Campaign name is required.')
        return redirect(url_for('admin_campaigns'))
    if discount_percent <= 0 or discount_percent > 100:
        flash('Discount percent must be between 1 and 100.')
        return redirect(url_for('admin_campaigns'))
    try:
        product_ids = json.loads(product_ids_raw)
    except Exception:
        product_ids = []
    if not product_ids:
        flash('Please generate the email before saving the campaign.')
        return redirect(url_for('admin_campaigns'))

    db.execute(
        "INSERT INTO email_campaigns (name, discount_percent, max_discount_amount, min_order_amount, product_ids, status)"
        " VALUES (?,?,?,?,?,'draft')",
        (name, discount_percent, max_discount_amount, min_order_amount, json.dumps(product_ids))
    )
    db.commit()
    flash(f'Campaign "{name}" saved as a draft.')
    return redirect(url_for('admin_campaigns'))


@app.route('/admin/campaigns/<int:campaign_id>/send', methods=['POST'])
@admin_required
def admin_campaign_send(campaign_id):
    db = get_db()
    campaign = db.execute('SELECT * FROM email_campaigns WHERE id=?', (campaign_id,)).fetchone()
    if not campaign:
        flash('Campaign not found.')
        return redirect(url_for('admin_campaigns'))
    if campaign['status'] not in ('draft', 'failed'):
        flash('This campaign has already been sent (or is currently sending).')
        return redirect(url_for('admin_campaigns'))

    group = request.form.get('recipient_group') or 'all'
    if group not in ('all', 'new', 'ordered'):
        group = 'all'
    users = _campaign_recipient_group(db, group)
    if not users:
        flash('No users found in that group.')
        return redirect(url_for('admin_campaigns'))

    db.execute("UPDATE email_campaigns SET status='sending', recipient_group=? WHERE id=?", (group, campaign_id))
    db.commit()

    t = threading.Thread(target=_send_campaign_batch, args=(campaign_id, list(users)), daemon=True)
    t.start()
    flash(f'Sending "{campaign["name"]}" to {len(users)} user(s) in the background — this will take a few minutes.')
    return redirect(url_for('admin_campaigns'))


@app.route('/admin/campaigns/<int:campaign_id>', methods=['GET'])
@admin_required
def admin_campaign_detail(campaign_id):
    db = get_db()
    campaign = db.execute('SELECT * FROM email_campaigns WHERE id=?', (campaign_id,)).fetchone()
    if not campaign:
        flash('Campaign not found.')
        return redirect(url_for('admin_campaigns'))
    recipients = [dict(r) for r in db.execute(
        'SELECT * FROM email_campaign_recipients WHERE campaign_id=? ORDER BY id', (campaign_id,)
    ).fetchall()]

    used_count = 0
    visited_count = 0
    failed_count = 0
    for r in recipients:
        coupon = db.execute('SELECT times_used FROM coupons WHERE code=?', (r['coupon_code'],)).fetchone()
        r['coupon_used'] = bool(coupon and (coupon.get('times_used') or 0) > 0)
        r['visited'] = bool(r.get('clicked_at'))
        if r.get('status') == 'failed':
            failed_count += 1
        if r['coupon_used']:
            used_count += 1
        if r['visited']:
            visited_count += 1

    sent_count = len(recipients) - failed_count
    return render_template('admin/admin_campaign_detail.html', campaign=campaign, recipients=recipients,
                            used_count=used_count, visited_count=visited_count,
                            failed_count=failed_count, sent_count=sent_count)


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'GET':
        return render_template('admin/admin_login.html')

    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''

    if not ADMIN_USERNAME or not ADMIN_PASSWORD or not ADMIN_TOTP_SECRET:
        flash('Admin authentication is not configured.', 'error')
        return render_template('admin/admin_login.html'), 500

    if hmac.compare_digest(username, ADMIN_USERNAME) and hmac.compare_digest(password, ADMIN_PASSWORD):
        session['admin_step'] = 'totp'
        session.pop('is_admin', None)
        session.modified = True
        return redirect(url_for('admin_verify_totp'))

    flash('Invalid username or password.', 'error')
    return render_template('admin/admin_login.html'), 401


@app.route('/admin/verify-totp', methods=['GET', 'POST'])
def admin_verify_totp():
    if session.get('admin_step') != 'totp':
        flash('Please complete login first.', 'error')
        return redirect(url_for('admin_login'))

    if request.method == 'GET':
        return render_template('admin/admin_totp.html')

    code = (request.form.get('totp_code') or '').strip().replace(' ', '')
    if not ADMIN_TOTP_SECRET:
        flash('TOTP is not configured.', 'error')
        return render_template('admin/admin_totp.html'), 500

    totp = pyotp.TOTP(ADMIN_TOTP_SECRET)
    if totp.verify(code, valid_window=1):
        session['is_admin'] = True
        session.pop('admin_step', None)
        session.modified = True
        return redirect(url_for('admin_dashboard'))

    flash('Invalid authentication code.', 'error')
    return render_template('admin/admin_totp.html'), 401


@app.route('/admin/logout', methods=['GET'])
@admin_required
def admin_logout():
    session.pop('is_admin', None)
    session.pop('admin_step', None)
    session.modified = True
    return redirect(url_for('admin_login'))

if __name__ == '__main__':
    app.run(debug=True)